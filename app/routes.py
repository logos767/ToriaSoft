import os
import time
import io
import json
import base64
import calendar
import secrets
from pathlib import Path
import requests
from pywebpush import webpush, WebPushException
import firebase_admin
import re
from flask import Blueprint, render_template, url_for, flash, redirect, request, jsonify, session, current_app
from flask_login import login_user, current_user, logout_user, login_required
from flask_wtf import FlaskForm
from flask_wtf.file import FileField, FileAllowed
from wtforms import StringField, SubmitField, SelectField, TextAreaField
from wtforms.validators import DataRequired, Length, Email, Optional
from sqlalchemy.exc import IntegrityError
from sqlalchemy import func, extract, or_, select, case, text
import openpyxl
from datetime import datetime, timedelta, date
from flask import Response
from weasyprint import HTML
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from babel.dates import get_month_names
from firebase_admin import messaging
from sqlalchemy.orm import joinedload, subqueryload
from .extensions import db, bcrypt, socketio
from .models import (User, Product, Client, Provider, Order, OrderItem, Purchase, PurchaseItem, Reception, Movement, 
                    CompanyInfo, CostStructure, Notification, ExchangeRate, get_current_time_ve, Bank, PointOfSale, UserActivityLog,
                    CashBox, Payment, ManualFinancialMovement, InventoryAdjustment, InventoryAdjustmentItem, VE_TIMEZONE, OrderReturn, OrderReturnItem,
                    UserDevice, Warehouse, ProductStock, WarehouseTransfer, BulkLoadLog)
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.graphics.barcode import createBarcodeDrawing
from reportlab.graphics import renderPM
from reportlab.graphics.shapes import Drawing

def get_main_calculation_currency_info():
    """Returns the main calculation currency and its symbol."""
    company_info = CompanyInfo.query.first()
    currency = company_info.calculation_currency if company_info and company_info.calculation_currency else 'USD'
    symbol = '€' if currency == 'EUR' else '$'
    return currency, symbol

# --- Helper functions for role-based access control ---
def is_superuser():
    return current_user.is_authenticated and current_user.role == 'Superusuario'

def is_gerente():
    return current_user.is_authenticated and current_user.role in ['Superusuario', 'Gerente']

def is_administrador():
    return current_user.is_authenticated and current_user.role in ['Superusuario', 'Gerente', 'Administrador']

def is_vendedor():
    return current_user.is_authenticated and current_user.role in ['Superusuario', 'Gerente', 'Administrador', 'Vendedor']

# --- End Helper functions for role-based access control ---

# --- Helper function for user activity logging ---
def log_user_activity(action, details=None, target_id=None, target_type=None):
    """
    Logs an activity for the current user if they are not a Superusuario.
    """
    if current_user.is_authenticated and current_user.role != 'Superusuario':
        try:
            log_entry = UserActivityLog(
                user_id=current_user.id,
                action=action,
                details=details,
                target_id=str(target_id) if target_id is not None else None,
                target_type=target_type
            )
            db.session.add(log_entry)
            db.session.commit()
        except Exception as e:
            current_app.logger.error(f"Error logging user activity for user {current_user.id}: {e}")
            db.session.rollback()

routes_blueprint = Blueprint('main', __name__)

# --- FORMULARIOS ---
class CompanyInfoForm(FlaskForm):
    name = StringField('Nombre de la Empresa', validators=[DataRequired()])
    rif = StringField('RIF', validators=[DataRequired()])
    address = TextAreaField('Dirección', validators=[Optional()]) # type: ignore
    phone_numbers = StringField('Teléfonos', validators=[Optional()])
    logo_file = FileField('Logo de la Empresa', validators=[FileAllowed(['jpg', 'png', 'jpeg', 'gif'], 'Solo se permiten imágenes.')])
    calculation_currency = SelectField('Moneda de Cálculo Principal', choices=[('USD', 'USD ($)'), ('EUR', 'EUR (€)')], validators=[DataRequired()])
    submit = SubmitField('Guardar Cambios')


# --- FORMULARIOS ---
class UpdateProfileForm(FlaskForm):
    first_name = StringField('Nombres', validators=[DataRequired(), Length(min=2, max=50)])
    last_name = StringField('Apellidos', validators=[DataRequired(), Length(min=2, max=50)])
    doc_type = SelectField('Tipo', choices=[('V', 'V'), ('E', 'E'), ('J', 'J'), ('G', 'G'), ('P', 'P')])
    doc_number = StringField('Nro. Documento', validators=[Length(max=20)])
    email = StringField('Correo Electrónico', validators=[DataRequired(), Email()])
    address = TextAreaField('Dirección', validators=[Length(max=255)])
    picture = FileField('Actualizar Foto de Perfil', validators=[FileAllowed(['jpg', 'png', 'jpeg'])])
    
    # Redes Sociales
    social_facebook = StringField('Facebook', validators=[Length(max=100)])
    social_instagram = StringField('Instagram', validators=[Length(max=100)])
    social_x = StringField('X (Twitter)', validators=[Length(max=100)])

    # Información Bancaria
    bank_name = StringField('Nombre del Banco', validators=[Length(max=100)])
    bank_account_number = StringField('Número de Cuenta', validators=[Length(max=25)])

    submit = SubmitField('Guardar Cambios')

def save_picture(form_picture):
    """Guarda la imagen de perfil del usuario y retorna el nombre del archivo."""
    random_hex = secrets.token_hex(8)
    _, f_ext = os.path.splitext(form_picture.filename)
    picture_fn = random_hex + f_ext
    picture_path = os.path.join(current_app.root_path, 'static/profile_pics', picture_fn)

    # Crear el directorio si no existe
    os.makedirs(os.path.dirname(picture_path), exist_ok=True)

    # Redimensionar imagen si es necesario (opcional, pero recomendado)
    # from PIL import Image
    # output_size = (125, 125)
    # i = Image.open(form_picture)
    # i.thumbnail(output_size)
    # i.save(picture_path)
    form_picture.save(picture_path)

    return picture_fn

def save_product_image(form_picture):
    """Guarda la imagen de un producto y retorna la ruta relativa para la URL."""
    random_hex = secrets.token_hex(8)
    _, f_ext = os.path.splitext(form_picture.filename)
    picture_fn = random_hex + f_ext
    
    # La ruta donde se guardará el archivo
    picture_path = os.path.join(current_app.root_path, 'static/product_images', picture_fn)

    # Crear el directorio si no existe
    os.makedirs(os.path.dirname(picture_path), exist_ok=True)

    # Guardar la imagen
    form_picture.save(picture_path)

    # Retornar la ruta que se usará en la URL
    # ej: 'product_images/abcdef12.png'
    return f'product_images/{picture_fn}'


# --- INICIO DE SECCIÓN DE TASAS DE CAMBIO ---

def obtener_tasas_exchangerate_api():
    """
    Obtiene las tasas de cambio desde exchangerate-api.com.
    Retorna un diccionario con las tasas de interés.
    """
    current_app.logger.info("Obteniendo tasas desde exchangerate-api.com...")
    api_url = "https://api.exchangerate-api.com/v4/latest/USD"
    try:
        response = requests.get(api_url, timeout=10)
        response.raise_for_status()
        data = response.json()

        if not data or 'rates' not in data:
            current_app.logger.warning("La respuesta de la API de exchangerate-api no contiene datos de tasas.")
            return None

        rates = data.get('rates')
        usd_ves_rate = rates.get('VES')
        usd_eur_rate = rates.get('EUR')

        if not usd_ves_rate or not usd_eur_rate:
            current_app.logger.error("No se pudieron encontrar las tasas VES o EUR en la respuesta.")
            return None

        # 1 EUR = (1 / usd_eur_rate) USD
        # 1 USD = usd_ves_rate VES
        # 1 EUR = (1 / usd_eur_rate) * usd_ves_rate VES
        eur_ves_rate = usd_ves_rate / usd_eur_rate

        current_app.logger.info(f"API de exchangerate-api exitosa. USD/VES: {usd_ves_rate}, EUR/VES: {eur_ves_rate}")
        return {
            'USD': usd_ves_rate,
            'EUR': eur_ves_rate
        }

    except requests.exceptions.RequestException as e:
        current_app.logger.error(f"Falló la petición a la API de exchangerate-api: {e}")
        return None
    except (ValueError, TypeError, KeyError) as e:
        current_app.logger.error(f"Error procesando la respuesta de la API de exchangerate-api: {e}")
        return None

# --- NUEVAS FUNCIONES AUXILIARES ---
def get_cached_exchange_rate(currency='USD'):
    """
    Obtiene la última tasa de cambio guardada en la base de datos para una moneda específica.
    """
    try:
        cached_rate = ExchangeRate.query.filter_by(currency=currency).order_by(ExchangeRate.date_updated.desc()).first()
        if cached_rate:
            return cached_rate.rate
    except Exception as e:
        current_app.logger.error(f"Error al obtener la tasa de cambio '{currency}' de la base de datos: {e}")
        db.session.rollback()
    
    current_app.logger.warning(f"No se encontró una tasa de cambio para '{currency}' en la base de datos.")
    return None

def fetch_and_update_exchange_rate():
    """
    Obtiene las tasas de cambio actuales VES/USD y VES/EUR de exchangerate-api.com y las guarda en la BD.
    """
    rates = obtener_tasas_exchangerate_api()

    if rates:
        try:
            for currency, rate_value in rates.items():
                exchange_rate_entry = ExchangeRate.query.filter_by(currency=currency).first()
                if exchange_rate_entry:
                    exchange_rate_entry.rate = rate_value
                    exchange_rate_entry.date_updated = get_current_time_ve()
                else:
                    exchange_rate_entry = ExchangeRate(currency=currency, rate=rate_value)
                    db.session.add(exchange_rate_entry)
            
            db.session.commit()
            current_app.logger.info(f"Tasas de cambio actualizadas en la base de datos: {rates}")
            return rates
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error al guardar la tasa de cambio en la base de datos: {e}")
    
    current_app.logger.error("No se pudo obtener ninguna tasa de cambio de las APIs externas.")
    return None

# --- FIN DE SECCIÓN DE TASAS DE CAMBIO ---


# --- Funciones del Sistema de Notificaciones ---

def create_notification_for_admins(message, link):
    """
    Crea una notificación en la BD para Superusuario y Gerente,
    y envía una notificación push a través de FCM a sus dispositivos registrados.
    """
    VAPID_PRIVATE_KEY = os.environ.get("VAPID_PRIVATE_KEY")
    VAPID_CLAIMS = {"sub": "mailto:your-email@example.com"} # Cambia esto a tu email
    current_app.logger.info(f"Attempting to create notification for admins: {message}")
    try:
        admins = User.query.filter(User.role.in_(['Superusuario', 'Gerente'])).all()
        if not admins:
            current_app.logger.warning("No se encontraron usuarios administradores para enviar la notificación.")
            return

        # 1. Guardar notificaciones en la BD y emitir por WebSocket
        for admin in admins:
            notification = Notification(
                user_id=admin.id,
                message=message,
                link=link
            )
            db.session.add(notification)
            db.session.flush() # Flush para que la notificación tenga ID y fecha antes de emitir

            # Emitir evento de WebSocket para la UI en tiempo real
            socketio.emit('new_notification', {
                'message': notification.message,
                'link': notification.link,
                'created_at': notification.created_at.strftime('%d/%m %H:%M')
            }, room=f'user_{admin.id}')
            current_app.logger.info(f"Notificación en BD y WebSocket para admin {admin.id}")

        db.session.commit()
        current_app.logger.info(f"Commit de {len(admins)} notificaciones a la BD.")

        # 2. Enviar notificaciones PUSH (Móvil y Web)
        admin_ids = [admin.id for admin in admins]
        devices = UserDevice.query.filter(UserDevice.user_id.in_(admin_ids)).all()
        
        fcm_tokens = [d.fcm_token for d in devices if d.device_type != 'web']
        web_push_subscriptions = [d for d in devices if d.device_type == 'web']

        # --- Enviar a dispositivos móviles (Android/iOS) ---
        if fcm_tokens and firebase_admin._apps:
            # CAMBIO IMPORTANTE: Enviar como mensaje de "solo datos"
            # para que la app móvil lo maneje siempre.
            fcm_message = messaging.MulticastMessage(
                data={
                    'title': 'Nueva Notificación de ToriaSoft',
                    'body': message,
                    'link': link
                },
                tokens=fcm_tokens,
                # Configuración para Android para que despierte la app
                android=messaging.AndroidConfig(
                    priority='high',
                ),
                # Configuración para APNs (iOS)
                apns=messaging.APNSConfig(
                    payload=messaging.APNSPayload(
                        aps=messaging.Aps(content_available=True)
                    )
                )
            )
            response = messaging.send_multicast(fcm_message)
            current_app.logger.info(f'Notificaciones FCM (datos) enviadas: {response.success_count} exitosas, {response.failure_count} fallidas.')

        # --- Enviar a navegadores web ---
        if web_push_subscriptions and VAPID_PRIVATE_KEY:
            push_payload = json.dumps({
                "title": "Nueva Notificación de ToriaSoft",
                "body": message,
                "icon": url_for('static', filename='images/logo.png', _external=True),
                "url": link
            })
            for sub_device in web_push_subscriptions:
                try:
                    subscription_info = json.loads(sub_device.fcm_token)
                    webpush(
                        subscription_info=subscription_info,
                        data=push_payload,
                        vapid_private_key=VAPID_PRIVATE_KEY,
                        vapid_claims=VAPID_CLAIMS.copy()
                    )
                except (WebPushException, json.JSONDecodeError, TypeError) as e:
                    current_app.logger.error(f"Error enviando web push a {sub_device.id}: {e}. Podría ser una suscripción inválida.")

    except Exception as e:
        current_app.logger.error(f"Error al crear notificaciones para administradores: {e}")
        db.session.rollback()

@routes_blueprint.context_processor
def inject_notifications():
    if not current_user.is_authenticated or not is_gerente(): # Superusuario and Gerente receive notifications
        return dict(unread_notifications=[], unread_notification_count=0)

    try:
        unread_notifications = Notification.query.filter_by(user_id=current_user.id, is_read=False).order_by(Notification.created_at.desc()).limit(10).all()
        count = Notification.query.filter_by(user_id=current_user.id, is_read=False).count()
        return dict(
            unread_notifications=unread_notifications,
            unread_notification_count=count
        )
    except Exception as e:
        current_app.logger.error(f"Error al obtener notificaciones para el usuario {current_user.id}: {e}")
        db.session.rollback()
        return dict(unread_notifications=[], unread_notification_count=0)

@routes_blueprint.context_processor
def inject_pending_withdrawals_count():
    if not current_user.is_authenticated or not is_gerente(): # Superusuario and Gerente manage withdrawals
        return dict(pending_withdrawals_count=0)
    
    try:
        count = ManualFinancialMovement.query.filter_by(status='Pendiente', movement_type='Egreso').count()
        return dict(pending_withdrawals_count=count)
    except Exception as e:
        current_app.logger.error(f"Error al obtener el conteo de retiros pendientes: {e}")
        db.session.rollback()
        return dict(pending_withdrawals_count=0)


@routes_blueprint.route('/set-display-currency', methods=['POST'])
def set_display_currency():
    """Sets the display currency in the user's session."""
    currency = request.form.get('currency')
    if currency in ['USD', 'EUR']:
        session['display_currency'] = currency
        flash(f'Moneda de cálculo cambiada a {currency}.', 'success')
    
    referrer = request.referrer
    return redirect(referrer) if referrer else redirect(url_for('main.dashboard'))

@routes_blueprint.context_processor
def inject_current_rate():
    """
    Injects the display currency, its symbol, and its exchange rate into the context for all templates.
    The currency is determined by session preference, falling back to company settings.
    """
    # 1. Determine the currency to use for display/calculation.
    # Priority: Session > Company Setting > Default 'USD'
    company_info = CompanyInfo.query.first()
    default_currency = company_info.calculation_currency if company_info and company_info.calculation_currency else 'USD'
    
    # The session stores the user's preference for this session.
    display_currency = session.get('display_currency', default_currency)

    # 2. Get the corresponding rate and symbol.
    rate = get_cached_exchange_rate(display_currency) or 0.0
    symbol = '€' if display_currency == 'EUR' else '$'
    
    # 3. Inject into context. The templates will use these variables.
    # 'calculation_currency' will be either 'USD' or 'EUR'.
    # 'current_rate' will be the rate for that currency to VES.
    # 'currency_symbol' will be '$' or '€'.
    return dict(
        current_rate=rate, 
        calculation_currency=display_currency, 
        currency_symbol=symbol
    )

@routes_blueprint.route('/notifications/mark-as-read', methods=['POST'])
@login_required
def mark_notifications_as_read():
    if not is_gerente(): # Superusuario and Gerente can mark as read
        return jsonify(success=False, message='Acceso denegado'), 403
    try:
        Notification.query.filter_by(user_id=current_user.id, is_read=False).update({'is_read': True})
        db.session.commit()
        return jsonify(success=True)
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error al marcar notificaciones como leídas para el usuario {current_user.id}: {e}")
        return jsonify(success=False, message='Error interno del servidor'), 500

@routes_blueprint.route('/notifications/list')
@login_required
def notification_list():
    """Muestra una página con todas las notificaciones del usuario."""
    if not is_gerente():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.new_order'))
    
    # Marcar todas como leídas al visitar la página
    Notification.query.filter_by(user_id=current_user.id, is_read=False).update({'is_read': True})
    db.session.commit()

    notifications = Notification.query.filter_by(user_id=current_user.id).order_by(Notification.created_at.desc()).all()
    return render_template('notificaciones/lista.html', title='Historial de Notificaciones', notifications=notifications)

@routes_blueprint.route('/notifications/clear', methods=['POST'])
@login_required
def clear_all_notifications():
    """Elimina todas las notificaciones del usuario actual."""
    if not is_gerente():
        return jsonify(success=False, message='Acceso denegado'), 403
    Notification.query.filter_by(user_id=current_user.id).delete()
    db.session.commit()
    flash('Historial de notificaciones limpiado.', 'success')
    return redirect(url_for('main.notification_list'))

@socketio.on('connect')
def handle_connect():
    if current_user.is_authenticated:
        from flask_socketio import join_room, rooms
        # Add a check to prevent joining the room if already in it.
        # This can help with reconnects.
        if f'user_{current_user.id}' not in rooms():
            join_room(f'user_{current_user.id}')

# Rutas de autenticación
@routes_blueprint.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        if current_user.role != 'Vendedor':
            return redirect(url_for('main.dashboard'))
        else:
            return redirect(url_for('main.new_order'))
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        if user and bcrypt.check_password_hash(user.password, password):
            if not user.is_active:
                flash('Tu cuenta de usuario ha sido desactivada. Por favor, contacta al administrador.', 'danger')
                return redirect(url_for('main.login'))

            login_user(user)
            
            # Actualizar la tasa de cambio al iniciar sesión
            current_app.logger.info(f"Usuario '{username}' ha iniciado sesión. Actualizando tasas de cambio...")
            rates = fetch_and_update_exchange_rate()
            if not rates:
                flash('Advertencia: No se pudo actualizar la tasa de cambio. Se usarán los últimos valores guardados.', 'warning')

            next_page = request.args.get('next') # type: ignore
            if next_page:
                return redirect(next_page)
            if user.role != 'Vendedor':
                return redirect(url_for('main.dashboard'))
            else:
                return redirect(url_for('main.new_order'))
        else:
            flash('Inicio de sesión fallido. Por favor, verifica tu nombre de usuario y contraseña.', 'danger')
    return render_template('login.html', title='Iniciar Sesión')


@routes_blueprint.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('main.login'))

# Rutas principales
@routes_blueprint.route('/')
@routes_blueprint.route('/dashboard')
@login_required
def dashboard():
    """Muestra la página principal con información de dashboard."""
    if current_user.role == 'Vendedor':
        return redirect(url_for('main.new_order'))

    # --- General Metrics ---
    # Excluir el grupo 'Ganchos' (insumos) de los conteos del dashboard.
    total_products = Product.query.filter(or_(Product.grupo != 'Ganchos', Product.grupo.is_(None))).count() # type: ignore
    total_stock_query = db.session.query(
        func.sum(ProductStock.quantity),
        func.sum(ProductStock.quantity * Product.price_usd)
    ).join(Product).filter(or_(Product.grupo != 'Ganchos', Product.grupo.is_(None))).first() # type: ignore
    total_stock = total_stock_query[0] or 0
    total_stock_value_usd = total_stock_query[1] or 0.0

    total_clients = Client.query.count()

    # --- Accounts Receivable ---
    current_rate_usd = get_cached_exchange_rate('USD') or 1.0

    # Optimized Accounts Receivable Calculation
    paid_sq = db.session.query(
        Payment.order_id.label('order_id'),
        func.sum(Payment.amount_usd_equivalent).label('total_paid_usd')
    ).group_by(Payment.order_id).subquery()

    debt_data_query = db.session.query(
        Order.client_id,
        (Order.total_amount_usd - func.coalesce(paid_sq.c.total_paid_usd, 0)).label('due_amount_usd')
    ).outerjoin(paid_sq, Order.id == paid_sq.c.order_id).subquery()

    final_debt_query = db.session.query(
        func.count(func.distinct(debt_data_query.c.client_id)),
        func.sum(debt_data_query.c.due_amount_usd)
    ).filter(debt_data_query.c.due_amount_usd > 0.01)

    debt_result = final_debt_query.first()
    clients_in_debt_count = debt_result[0] or 0
    total_due_usd = debt_result[1] or 0.0

    # --- Order Statistics ---
    today = get_current_time_ve().date()
    start_of_day = VE_TIMEZONE.localize(datetime.combine(today, datetime.min.time()))
    end_of_day = VE_TIMEZONE.localize(datetime.combine(today, datetime.max.time()))
    start_of_month = today.replace(day=1)
    start_of_month_dt = VE_TIMEZONE.localize(datetime.combine(start_of_month, datetime.min.time()))

    # Optimized Order Statistics Calculation
    def get_order_stats(start_date, end_date=None):
        query = db.session.query(
            func.count(Order.id),
            func.sum(Order.total_amount / Order.exchange_rate_at_sale)
        ).filter(
            Order.exchange_rate_at_sale.isnot(None),
            Order.exchange_rate_at_sale > 0
        )
        if end_date:
            query = query.filter(Order.date_created.between(start_date, end_date))
        else:
            query = query.filter(Order.date_created >= start_date)
        
        count, amount_usd = query.first()
        return count or 0, float(amount_usd or 0.0)

    orders_today_count, orders_today_amount_usd = get_order_stats(start_of_day, end_of_day)
    orders_month_count, orders_month_amount_usd = get_order_stats(start_of_month_dt)
    
    all_stats_query = db.session.query(
        func.count(Order.id),
        func.sum(Order.total_amount / Order.exchange_rate_at_sale)
    ).filter(
        Order.exchange_rate_at_sale.isnot(None),
        Order.exchange_rate_at_sale > 0
    )
    all_orders_count, all_orders_amount_usd = all_stats_query.first()
    all_orders_count = all_orders_count or 0
    all_orders_amount_usd = float(all_orders_amount_usd or 0.0)

    # --- Daily Cash & Bank Movements by Account ---
    from collections import defaultdict
    daily_movements_by_account = defaultdict(lambda: {'inflows': defaultdict(float), 'outflows': defaultdict(float)})

    # 1. Procesar Pagos (Ingresos)
    payments_today = Payment.query.options(
        joinedload(Payment.bank),
        joinedload(Payment.pos).joinedload(PointOfSale.bank),
        joinedload(Payment.cash_box)
    ).filter(Payment.date.between(start_of_day, end_of_day)).all()

    for payment in payments_today:
        target_account_name = None
        if payment.cash_box:
            target_account_name = payment.cash_box.name
            if payment.currency_paid == 'USD':
                daily_movements_by_account[target_account_name]['inflows']['USD'] += float(payment.amount_paid or 0.0)
            else: # Pagos en VES a caja
                daily_movements_by_account[target_account_name]['inflows']['VES'] += float(payment.amount_ves_equivalent or 0.0)
        else: # Pagos a Bancos (directo o por POS)
            bank = payment.bank or (payment.pos.bank if payment.pos else None)
            if bank:
                target_account_name = bank.name
                # Todos los ingresos a bancos se registran como VES
                daily_movements_by_account[target_account_name]['inflows']['VES'] += float(payment.amount_ves_equivalent or 0.0)

    # 2. Procesar Movimientos Manuales (Ingresos y Egresos)
    manual_movements_today = ManualFinancialMovement.query.options(
        joinedload(ManualFinancialMovement.bank),
        joinedload(ManualFinancialMovement.cash_box)
    ).filter(ManualFinancialMovement.date.between(start_of_day, end_of_day), ManualFinancialMovement.status == 'Aprobado').all()

    for m in manual_movements_today:
        account_name = None
        if m.cash_box:
            account_name = m.cash_box.name
        elif m.bank:
            account_name = m.bank.name

        if m.movement_type == 'Ingreso':
            daily_movements_by_account[account_name]['inflows'][m.currency] += float(m.amount or 0.0)
        elif m.movement_type == 'Egreso':
            daily_movements_by_account[account_name]['outflows'][m.currency] += float(m.amount or 0.0)

    # --- Credits & Reservations Summary (REMOVED as per request) ---
    # def get_credit_reservation_stats(start_date, end_date=None):
    #     query = db.session.query(func.count(Order.id), func.sum(Order.total_amount_usd)).filter(Order.order_type.in_(['credit', 'reservation']))
    #     if end_date:
    #         query = query.filter(Order.date_created.between(start_date, end_date))
    #     else:
    #         query = query.filter(Order.date_created >= start_date)
    #     count, amount_usd = query.first()
    #     return count or 0, float(amount_usd or 0.0)
    # credits_reservations_today_count, credits_reservations_today_amount_usd = get_credit_reservation_stats(start_of_day, end_of_day)
    # credits_reservations_month_count, credits_reservations_month_amount_usd = get_credit_reservation_stats(start_of_month_dt)

    # --- Accounting Donut Chart Data (Current Month) ---
    def get_accounting_data(start_date, end_date=None):
        # Sales by status
        sales_query = db.session.query(
            Order.status,
            func.sum(Order.total_amount / Order.exchange_rate_at_sale)
        ).filter(
            Order.exchange_rate_at_sale.isnot(None), Order.exchange_rate_at_sale > 0
        )
        if end_date: sales_query = sales_query.filter(Order.date_created.between(start_date, end_date))
        else: sales_query = sales_query.filter(Order.date_created >= start_date)
        
        sales_results = sales_query.group_by(Order.status).all()
        sales = {'contado': 0.0, 'credito': 0.0, 'apartado': 0.0}
        for status, amount in sales_results:
            amount = float(amount or 0.0)
            if status in ['Pagada', 'Completada']: sales['contado'] += amount
            elif status == 'Crédito': sales['credito'] += amount
            elif status == 'Apartado': sales['apartado'] += amount

        # Variable Expenses
        cost_structure = CostStructure.query.first() or CostStructure()
        var_sales_exp_pct = case((Product.variable_selling_expense_percent > 0, Product.variable_selling_expense_percent), else_=(cost_structure.default_sales_commission_percent or 0))
        var_marketing_pct = case((Product.variable_marketing_percent > 0, Product.variable_marketing_percent), else_=(cost_structure.default_marketing_percent or 0))
        
        expenses_query = db.session.query(func.sum(
            (OrderItem.quantity * (OrderItem.cost_at_sale_ves or 0)) + 
            ((OrderItem.quantity * OrderItem.price) * (var_sales_exp_pct + var_marketing_pct))
        )).join(Order).join(Product).filter(or_(Product.grupo != 'Ganchos', Product.grupo.is_(None)))

        if end_date: expenses_query = expenses_query.filter(Order.date_created.between(start_date, end_date))
        else: expenses_query = expenses_query.filter(Order.date_created >= start_date)

        variable_expenses_ves = expenses_query.scalar() or 0.0
        variable_expenses_usd = variable_expenses_ves / current_rate_usd if current_rate_usd > 0 else 0.0
        
        return sales, variable_expenses_usd

    sales_month, variable_expenses_usd_month = get_accounting_data(start_of_month_dt)
    cost_structure = CostStructure.query.first() or CostStructure()
    fixed_expenses_usd_month = (cost_structure.monthly_rent or 0) + (cost_structure.monthly_utilities or 0) + (cost_structure.monthly_fixed_taxes or 0)

    accounting_chart_data = {
        'labels': ['Ventas Contado', 'Ventas Crédito', 'Ventas Apartado', 'Gastos Fijos', 'Gastos Variables'],
        'values': [round(sales_month['contado'], 2), round(sales_month['credito'], 2), round(sales_month['apartado'], 2), round(fixed_expenses_usd_month, 2), round(variable_expenses_usd_month, 2)]
    }

    # --- Accounting Donut Chart Data (Current Day) ---
    sales_day, variable_expenses_usd_day = get_accounting_data(start_of_day, end_of_day)
    fixed_expenses_usd_day = fixed_expenses_usd_month / 30.44 # Daily prorated fixed expenses

    accounting_chart_data_day = {
        'labels': ['Ventas Contado', 'Ventas Crédito', 'Ventas Apartado', 'Gastos Fijos', 'Gastos Variables'],
        'values': [round(sales_day['contado'], 2), round(sales_day['credito'], 2), round(sales_day['apartado'], 2), round(fixed_expenses_usd_day, 2), round(variable_expenses_usd_day, 2)]
    }

    # Get current month name in Spanish
    month_names = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    current_month_name = f"{month_names[today.month - 1]} {today.year}"

    # --- Recent Activity ---
    recent_products = Product.query.filter(or_(Product.grupo != 'Ganchos', Product.grupo.is_(None))).order_by(Product.id.desc()).limit(5).all()
    recent_orders = Order.query.options(joinedload(Order.client)).order_by(Order.date_created.desc()).limit(5).all()

    return render_template('index.html', title='Dashboard',
                           total_products=total_products,
                           total_stock=total_stock,
                           total_stock_value_usd=total_stock_value_usd,
                           total_clients=total_clients,
                           clients_in_debt_count=clients_in_debt_count,
                           total_due_usd=total_due_usd,
                           orders_today_count=orders_today_count,
                           orders_today_amount_usd=orders_today_amount_usd,
                           orders_month_count=orders_month_count,
                           orders_month_amount_usd=orders_month_amount_usd,
                           all_orders_count=all_orders_count,
                           all_orders_amount_usd=all_orders_amount_usd,
                           daily_movements_by_account=daily_movements_by_account,
                           accounting_chart_data=accounting_chart_data,
                           accounting_chart_data_day=accounting_chart_data_day,
                           current_month_name=current_month_name,
                           recent_products=recent_products,
                           recent_orders=recent_orders)

@routes_blueprint.context_processor
def inject_user_profile_pic():
    """Inyecta la URL de la imagen de perfil del usuario en todas las plantillas."""
    if current_user.is_authenticated:
        image_file = url_for('static', filename='profile_pics/' + (current_user.profile_image_file or 'default.png'))
        return dict(profile_image_file=image_file)
    return dict(profile_image_file=url_for('static', filename='profile_pics/default.png'))


# --- Rutas de Perfil de Usuario ---
@routes_blueprint.route('/perfil', methods=['GET', 'POST'])
@login_required
def user_profile():
    form = UpdateProfileForm()
    if form.validate_on_submit():
        if form.picture.data:
            picture_file = save_picture(form.picture.data)
            current_user.profile_image_file = picture_file

        current_user.first_name = form.first_name.data
        current_user.last_name = form.last_name.data
        current_user.doc_type = form.doc_type.data
        current_user.doc_number = form.doc_number.data
        current_user.email = form.email.data
        current_user.address = form.address.data
        current_user.social_facebook = form.social_facebook.data
        current_user.social_instagram = form.social_instagram.data
        current_user.social_x = form.social_x.data
        current_user.bank_name = form.bank_name.data
        current_user.bank_account_number = form.bank_account_number.data
        
        db.session.commit()
        flash('¡Tu perfil ha sido actualizado!', 'success')
        return redirect(url_for('main.user_profile'))
    elif request.method == 'GET':
        form.first_name.data = current_user.first_name
        form.last_name.data = current_user.last_name
        form.doc_type.data = current_user.doc_type
        form.doc_number.data = current_user.doc_number
        form.email.data = current_user.email
        form.address.data = current_user.address
        form.social_facebook.data = current_user.social_facebook
        form.social_instagram.data = current_user.social_instagram
        form.social_x.data = current_user.social_x
        form.bank_name.data = current_user.bank_name
        form.bank_account_number.data = current_user.bank_account_number

    image_file = url_for('static', filename='profile_pics/' + (current_user.profile_image_file or 'default.png'))
    return render_template('perfil/perfil.html', title='Mi Perfil',
                           image_file=image_file, form=form)

@routes_blueprint.route('/perfil/detalles')
@login_required
def profile_details():
    """Muestra la página de solo lectura del perfil del usuario."""
    image_file = url_for('static', filename='profile_pics/' + (current_user.profile_image_file or 'default.png'))
    return render_template('perfil/detalles_perfil.html', title='Detalles de Mi Perfil', user=current_user, image_file=image_file)



# Rutas de productos (Inventario)
@routes_blueprint.route('/inventario/lista')
@login_required
def inventory_list():
    search_term = request.args.get('search', '').strip()
    group_filter = request.args.get('group', '').strip()
    sort_by = request.args.get('sort_by', 'name')
    sort_order = request.args.get('sort_order', 'asc')

    query = Product.query

    if group_filter:
        query = query.filter(Product.grupo == group_filter)

    if search_term:
        search_pattern = f'%{search_term}%'
        query = query.filter(or_(
            Product.name.ilike(search_pattern),
            Product.barcode.ilike(search_pattern),
            Product.codigo_producto.ilike(search_pattern),
            Product.marca.ilike(search_pattern),
            Product.size.ilike(search_pattern)
        ))

    # Lógica de ordenación
    valid_sort_columns = {
        'name': Product.name,
        'barcode': Product.barcode,
        'codigo_producto': Product.codigo_producto
    }
    sort_column = valid_sort_columns.get(sort_by, Product.name)

    products = query.order_by(
        sort_column.desc() if sort_order == 'desc' else sort_column.asc()
    ).all()
    
    # Obtener todos los grupos únicos para el menú desplegable de filtro
    groups = db.session.query(Product.grupo).distinct().order_by(Product.grupo).all()
    product_groups = [g[0] for g in groups if g[0]]

    user_role = current_user.role if current_user.is_authenticated else 'invitado'
    return render_template('inventario/lista.html',
                           title='Productos',
                           products=products,
                           user_role=user_role,
                           product_groups=product_groups,
                           filters={'search': search_term, 'group': group_filter},
                           sort_by=sort_by,
                           sort_order=sort_order)

@routes_blueprint.route('/inventario/codigos_barra', methods=['GET'])
@login_required
def codigos_barra():
    if not is_gerente(): # Superusuario and Gerente can access
        flash('Acceso denegado. Solo los administradores pueden ver esta sección.', 'danger')
        return redirect(url_for('main.new_order'))

    preselected_ids_str = request.args.get('product_ids', '')
    preselected_ids = preselected_ids_str.split(',') if preselected_ids_str else []

    sort_by = request.args.get('sort_by', 'id')
    sort_order = request.args.get('sort_order', 'desc')

    query = Product.query

    # Lógica de ordenación
    valid_sort_columns = {
        'name': Product.name,
        'barcode': Product.barcode,
        'codigo_producto': Product.codigo_producto,
        'id': Product.id
    }
    sort_column = valid_sort_columns.get(sort_by, Product.id)

    if sort_order == 'desc':
        query = query.order_by(sort_column.desc())
    else:
        query = query.order_by(sort_column.asc())

    products = query.all()
    groups = db.session.query(Product.grupo).distinct().order_by(Product.grupo).all()
    product_groups = [g[0] for g in groups if g[0]]
    company_info = CompanyInfo.query.first()
    _, currency_symbol = get_main_calculation_currency_info()

    return render_template('inventario/codigos_barra.html', title='Imprimir Códigos de Barra', 
                           products=products, product_groups=product_groups,
                           preselected_ids=preselected_ids,
                           company_info=company_info, currency_symbol=currency_symbol,
                           sort_by=sort_by, sort_order=sort_order)

@routes_blueprint.route('/inventario/codigos_barra_api', methods=['GET']) # type: ignore
@login_required
def codigos_barra_api():
    if not is_gerente(): # Superusuario and Gerente can access
        return jsonify(error='Acceso denegado'), 403

    search_term = request.args.get('search', '').lower()
    group_filter = request.args.get('group', '').strip()
    sort_by = request.args.get('sort_by', 'id')
    sort_order = request.args.get('sort_order', 'desc')

    query = Product.query

    if group_filter:
        query = query.filter(Product.grupo == group_filter)

    if search_term:
        query = query.filter(or_(
            Product.name.ilike(f'%{search_term}%'),
            Product.barcode.ilike(f'%{search_term}%')
        ))

    # Lógica de ordenación para la API
    valid_sort_columns = {
        'name': Product.name,
        'barcode': Product.barcode,
        'codigo_producto': Product.codigo_producto,
        'id': Product.id
    }
    sort_column = valid_sort_columns.get(sort_by, Product.id)

    if sort_order == 'desc':
        query = query.order_by(sort_column.desc())
    else:
        query = query.order_by(sort_column.asc())

    products = query.limit(500).all() # Limitar para no sobrecargar la respuesta
    
    return jsonify(products=[{
        'id': p.id, 'name': p.name, 'barcode': p.barcode,
        'codigo_producto': p.codigo_producto, 'marca': p.marca,
        'size': p.size, 'color': p.color, 'price_usd': p.price_usd
    } for p in products])

def generate_barcode_pdf_reportlab(products, company_info, currency_symbol):
    """
    Generate PDF with barcodes using ReportLab for better performance.
    Layout: 4 columns x 10 rows = 40 labels per page
    """
    # Create PDF buffer
    buffer = io.BytesIO()

    # Importar Code128 aquí para que la función sea autocontenida
    from reportlab.graphics.barcode import code128

    # Page dimensions
    page_width, page_height = A4
    margin = 3 * mm

    # Label dimensions (same as HTML template)
    label_width = 51 * mm
    label_height = 29 * mm

    # Crear el lienzo del PDF
    c = canvas.Canvas(buffer, pagesize=A4)

    # Create PDF canvas directly for more control
    c.setFont("Helvetica", 6)

    # Process products in batches of 40 (4x10 grid)
    for i in range(0, len(products), 40):
        batch = products[i:i+40]

        for j, product in enumerate(batch):
            # Calculate position in grid
            col = j % 4
            row = j // 4

            # Calculate position coordinates
            x = margin + col * label_width
            y = page_height - margin - (row + 1) * label_height

            # Company name (top right)
            if company_info and company_info.name:
                c.setFont("Helvetica-Bold", 8)
                company_name = company_info.name[:20]
                # Right align company name
                text_width = c.stringWidth(company_name, "Helvetica-Bold", 8)
                c.drawString(x + 1*mm, y + label_height - 3*mm, company_name)

            # Product name (centered, allow two lines for long names)
            c.setFont("Helvetica", 8)
            product_name = product['name'][:54]  # Allow longer names
            if len(product_name) > 27:
                # Split into two lines
                words = product_name.split()
                line1 = ""
                line2 = ""
                for word in words:
                    if len(line1 + " " + word) <= 27:
                        line1 += " " + word if line1 else word
                    else:
                        line2 += " " + word if line2 else word
                if not line2:
                    # If can't split nicely, force split
                    line1 = product_name[:27]
                    line2 = product_name[27:]
            else:
                line1 = product_name
                line2 = ""

            # Draw first line
            text_width = c.stringWidth(line1, "Helvetica", 8)
            c.drawString(x + (label_width - text_width) / 2, y + label_height - 6*mm, line1)

            # Draw second line if exists
            if line2:
                text_width = c.stringWidth(line2, "Helvetica", 8)
                c.drawString(x + (label_width - text_width) / 2, y + label_height - 9*mm, line2)

            # Price (below product name, adjust position if two lines)
            c.setFont("Helvetica-Bold", 9)
            # Cambiar el símbolo de dólar a 'ref.' para las etiquetas
            display_symbol = 'REF.' if currency_symbol == '$' else currency_symbol
            price_text = f"{display_symbol} {product['price_foreign']:.2f}"
            text_width = c.stringWidth(price_text, "Helvetica-Bold", 9)
            price_y = y + label_height - 3*mm
            c.drawString(x + label_width - text_width - 2*mm, price_y, price_text)

            # Barcode (bottom, lowered to make space)
            if product['barcode']:
                try:
                    # Calculate available width for barcode (full label width minus small margins)
                    available_width = label_width - 4*mm  # Leave 2mm margin on each side

                    # Create barcode using ReportLab with full width
                    barcode_obj = code128.Code128( # Esta línea estaba fallando por la importación faltante
                        product['barcode'],
                        barWidth=0.45*mm,  # Slightly thinner bars to fit more
                        barHeight=12*mm,    # Taller barcode
                        quiet=1
                    )

                    # Position barcode to span full width of label, lowered
                    barcode_x = x - 4*mm  # 2mm left margin
                    barcode_y = y + 6*mm  # Lowered from 6mm to 3mm to make space

                    # Draw barcode on canvas
                    barcode_obj.drawOn(c, barcode_x, barcode_y)

                    # Add barcode text below the barcode
                    c.setFont("Helvetica", 12)  # Small font for barcode text
                    barcode_text = product['barcode']
                    text_width = c.stringWidth(barcode_text, "Helvetica", 12)
                    text_x = x + (label_width - text_width) / 2  # Center the text
                    text_y = barcode_y - 4*mm  # Position below barcode

                    c.drawString(text_x, text_y, barcode_text)

                except Exception as e:
                    current_app.logger.error(f"Error generating barcode for {product['barcode']}: {e}")
                    # Draw error text instead
                    c.setFont("Helvetica-Bold", 6)
                    c.drawString(x + 2*mm, y + 4*mm, "Error")

            # Draw dashed border around the label with thinner lines and dash pattern for segmentation
            c.setLineWidth(0.5)
            c.setDash(2 * mm, 2 * mm)  # 2mm dash, 2mm gap
            c.rect(x, y, label_width, label_height, stroke=1, fill=0)
            c.setDash()  # reset dash pattern to solid

        # Start new page if there are more products
        if i + 40 < len(products):
            c.showPage()
            c.setFont("Helvetica", 6)

    # Save PDF
    c.save()

    # Get PDF data
    pdf_data = buffer.getvalue()
    buffer.close()

    return pdf_data


@routes_blueprint.route('/inventario/imprimir_codigos_barra', methods=['POST'])
@login_required
def imprimir_codigos_barra():
    if not is_gerente(): # Superusuario and Gerente can access
        flash('Acceso denegado. Solo los administradores pueden realizar esta acción.', 'danger')
        return redirect(url_for('main.new_order'))

    product_ids = request.form.getlist('product_ids')
    if not product_ids:
        flash('No se seleccionó ningún producto para imprimir.', 'warning')
        return redirect(url_for('main.codigos_barra'))

    products_to_print = Product.query.filter(Product.id.in_(product_ids)).all()
    company_info = CompanyInfo.query.first()
    
    _, currency_symbol = get_main_calculation_currency_info()

    # Preparar datos de productos para ReportLab, repitiendo por existencia.
    products_dict = []
    total_labels = 0
    MAX_LABELS = 10000  # Límite para prevenir sobrecarga del servidor.

    # Primero, calcular el número total de etiquetas para verificar el límite.
    for p in products_to_print:
        total_labels += p.stock if p.stock and p.stock > 0 else 0
    
    if total_labels > MAX_LABELS:
        flash(f'Ha intentado imprimir {total_labels} etiquetas, lo cual supera el límite de {MAX_LABELS}. Por favor, seleccione menos productos.', 'danger')
        return redirect(url_for('main.codigos_barra'))

    if total_labels == 0:
        flash('Los productos seleccionados no tienen existencia. No se generaron códigos de barra.', 'warning')
        return redirect(url_for('main.codigos_barra'))

    # Si estamos dentro del límite, construir la lista de etiquetas.
    for p in products_to_print:
        if p.stock and p.stock > 0:
            price_foreign = p.price_usd if p.price_usd else 0
            for _ in range(p.stock):
                products_dict.append({
                    'id': p.id,
                    'name': p.name,
                    'barcode': p.barcode,
                    'price_foreign': price_foreign
                })

    current_app.logger.info(f"Preparando datos para generación de PDF con {len(products_dict)} etiquetas para {len(products_to_print)} productos distintos.")

    # Generate PDF with ReportLab (more efficient)
    try:
        start_time = time.time()

        pdf_data = generate_barcode_pdf_reportlab(products_dict, company_info, currency_symbol)

        generation_time = time.time() - start_time
        current_app.logger.info(f"PDF generado exitosamente con ReportLab en {generation_time:.2f} segundos")

        # Return the PDF as a response
        return Response(pdf_data, mimetype='application/pdf', headers={'Content-Disposition': 'inline; filename=codigos_de_barra.pdf'})

    except Exception as e:
        current_app.logger.error(f"Error generating PDF with ReportLab: {str(e)}")
        error_message = f"Error generando PDF: {str(e)}. Intente con menos productos o contacte al administrador."
        log_user_activity(
            action="Intentó imprimir códigos de barra",
            details=f"Error al generar PDF para {len(product_ids)} productos.",
            target_id=None,
            target_type="BarcodePrinting"
        )

        # Fallback: try to generate a simple error PDF with ReportLab
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import A4

            buffer = io.BytesIO()
            c = canvas.Canvas(buffer, pagesize=A4)
            c.drawString(100, 750, "Error generando PDF")
            c.drawString(100, 720, error_message)
            c.drawString(100, 690, f"Número de productos seleccionados: {len(products_to_print)}")
            c.drawString(100, 660, "Recomendación: Seleccione máximo 100 productos por vez.")
            c.save()

            error_pdf = buffer.getvalue()
            buffer.close()
            return Response(error_pdf, mimetype='application/pdf', headers={'Content-Disposition': 'inline; filename=error.pdf'})

        except:
            # If even error PDF fails, return plain text
            return Response(error_message, mimetype='text/plain')
    
    log_user_activity(
        action="Imprimió códigos de barra",
        details=f"Generó PDF con {total_labels} etiquetas para {len(products_to_print)} productos.",
        target_id=None,
        target_type="BarcodePrinting")

@routes_blueprint.route('/inventario/carga_masiva/imprimir_codigos/<int:log_id>', methods=['GET'])
@login_required
def print_bulk_load_barcodes(log_id):
    """
    Generates and serves a PDF with barcodes for all products and quantities
    from a specific bulk load operation.
    """
    if not is_gerente():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.dashboard'))

    log_entry = BulkLoadLog.query.get_or_404(log_id)
    movements = log_entry.movements.options(joinedload(Movement.product)).filter(Movement.type == 'Entrada').all()

    if not movements:
        flash('Esta carga masiva no contiene productos para imprimir.', 'warning')
        return redirect(url_for('main.bulk_load_detail', log_id=log_id))

    company_info = CompanyInfo.query.first()
    _, currency_symbol = get_main_calculation_currency_info()

    products_dict = []
    for movement in movements:
        product = movement.product
        price_foreign = product.price_usd if product.price_usd else 0
        for _ in range(movement.quantity):
            products_dict.append({
                'id': product.id,
                'name': product.name,
                'barcode': product.barcode,
                'price_foreign': price_foreign
            })

    current_app.logger.info(f"Generando PDF con {len(products_dict)} etiquetas para la carga masiva #{log_id}.")

    try:
        start_time = time.time()
        pdf_data = generate_barcode_pdf_reportlab(products_dict, company_info, currency_symbol)
        generation_time = time.time() - start_time
        current_app.logger.info(f"PDF de carga masiva generado en {generation_time:.2f} segundos.")
        return Response(pdf_data, mimetype='application/pdf', headers={'Content-Disposition': f'inline; filename=codigos_carga_{log_id}.pdf'})
    except Exception as e:
        current_app.logger.error(f"Error generando PDF para carga masiva #{log_id}: {e}")
        flash(f"Error al generar el PDF: {e}", 'danger')
        return redirect(url_for('main.bulk_load_detail', log_id=log_id))

@routes_blueprint.route('/inventario/existencias')
@login_required
def inventory_stock():
    # No role check here, as per original code and user's implicit request for report visibility.
    # Permissions for specific actions (like PDF report or adjustment) are handled in their respective routes.

    selected_warehouse_id = request.args.get('warehouse_id', type=int)
    show_zero_stock = request.args.get('show_zero_stock') == 'on'

    warehouses = Warehouse.query.order_by(Warehouse.id).all()
    products_data = [] # This will hold {'product': Product_obj, 'stock': quantity}
    selected_warehouse_name = None

    if selected_warehouse_id:
        selected_warehouse_obj = next((wh for wh in warehouses if wh.id == selected_warehouse_id), None)
        if selected_warehouse_obj:
            selected_warehouse_name = selected_warehouse_obj.name

        # Query for products with stock in the selected warehouse
        query = db.session.query(Product, ProductStock.quantity) \
                          .join(ProductStock, Product.id == ProductStock.product_id) \
                          .filter(ProductStock.warehouse_id == selected_warehouse_id) \
                          .filter(or_(Product.grupo != 'Ganchos', Product.grupo.is_(None)))

        if not show_zero_stock:
            query = query.filter(ProductStock.quantity > 0)
        
        products_with_stock_tuples = query.order_by(Product.name).all()
        products_data = [{'product': p, 'stock': q} for p, q in products_with_stock_tuples]

    else:
        # If no specific warehouse is selected, show all products with their total stock.
        # We need to calculate the total stock within the query itself.
        
        # Subquery to sum quantities for each product across all warehouses
        total_stock_subquery = db.session.query(
            ProductStock.product_id,
            func.sum(ProductStock.quantity).label('total_quantity')
        ).group_by(ProductStock.product_id).subquery()

        # Main query: select Product and its total_quantity from the subquery
        query = db.session.query(
            Product,
            func.coalesce(total_stock_subquery.c.total_quantity, 0).label('total_stock')
        ).outerjoin(total_stock_subquery, Product.id == total_stock_subquery.c.product_id) \
        .filter(or_(Product.grupo != 'Ganchos', Product.grupo.is_(None)))

        if not show_zero_stock:
            query = query.filter(func.coalesce(total_stock_subquery.c.total_quantity, 0) > 0)
        
        products_with_stock_tuples = query.order_by(Product.name).all()
        products_data = [{'product': p, 'stock': q} for p, q in products_with_stock_tuples]

    return render_template('inventario/existencias.html',
                           title='Existencias por Almacén',
                           products_data=products_data,
                           warehouses=warehouses,
                           selected_warehouse_id=selected_warehouse_id,
                           selected_warehouse_name=selected_warehouse_name,
                           show_zero_stock=show_zero_stock)

@routes_blueprint.route('/inventario/existencias/reporte_pdf')
@login_required
def inventory_stock_report_pdf():
    """Genera un reporte PDF para el conteo físico del inventario. Accesible por Gerente y Superusuario."""
    if not is_administrador():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.inventory_stock'))
    
    selected_warehouse_id = request.args.get('warehouse_id', type=int)
    show_zero_stock = request.args.get('show_zero_stock') == 'True' # The boolean is passed as a string

    products_data = []
    selected_warehouse_name = "Todos los Almacenes"

    if selected_warehouse_id:
        warehouse = Warehouse.query.get(selected_warehouse_id)
        if warehouse:
            selected_warehouse_name = warehouse.name

        # Query for products with stock in the selected warehouse
        query = db.session.query(Product, ProductStock.quantity) \
                          .join(ProductStock, Product.id == ProductStock.product_id) \
                          .filter(ProductStock.warehouse_id == selected_warehouse_id) \
                          .filter(or_(Product.grupo != 'Ganchos', Product.grupo.is_(None)))

        if not show_zero_stock:
            query = query.filter(ProductStock.quantity > 0)
        
        products_with_stock_tuples = query.order_by(Product.name).all()
        products_data = [{'product': p, 'stock': q} for p, q in products_with_stock_tuples]
    else:
        # If no specific warehouse is selected, show all products with their total stock
        total_stock_subquery = db.session.query(ProductStock.product_id, func.sum(ProductStock.quantity).label('total_quantity')).group_by(ProductStock.product_id).subquery()
        query = db.session.query(Product, func.coalesce(total_stock_subquery.c.total_quantity, 0).label('total_stock')) \
                          .outerjoin(total_stock_subquery, Product.id == total_stock_subquery.c.product_id) \
                          .filter(or_(Product.grupo != 'Ganchos', Product.grupo.is_(None)))
        if not show_zero_stock:
            query = query.filter(func.coalesce(total_stock_subquery.c.total_quantity, 0) > 0)
        products_with_stock_tuples = query.order_by(Product.name).all()
        products_data = [{'product': p, 'stock': q} for p, q in products_with_stock_tuples]

    company_info = CompanyInfo.query.first()
    generation_date = get_current_time_ve().strftime('%d/%m/%Y %H:%M:%S')

    html_string = render_template('pdf/inventory_stock_report.html',
                                  products_data=products_data,
                                  company_info=company_info,
                                  generation_date=generation_date,
                                  warehouse_name=selected_warehouse_name)

    pdf_file = HTML(string=html_string, base_url=request.base_url).write_pdf()

    response = Response(pdf_file, mimetype='application/pdf', headers={'Content-Disposition': 'inline; filename=reporte_existencias.pdf'})
    return response

@routes_blueprint.route('/inventario/ajuste', methods=['GET', 'POST'])
@login_required
def inventory_adjustment():
    """Página para el ajuste digital del inventario. Accesible por Gerente y Superusuario."""
    if not is_administrador():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.new_order'))

    warehouse_id = request.args.get('warehouse_id', 1, type=int) # Por defecto, ajustar la tienda

    if request.method == 'POST':
        try:
            adjustments = request.form.getlist('adjustments')
            reason = request.form.get('reason', 'Ajuste de inventario manual')
            warehouse_id_form = request.form.get('warehouse_id', type=int)

            if not adjustments:
                flash('No se detectaron cambios para ajustar.', 'warning')
                return redirect(url_for('main.inventory_adjustment'))
            
            if not warehouse_id_form:
                flash('Debe seleccionar un almacén para realizar el ajuste.', 'danger')
                return redirect(url_for('main.inventory_adjustment'))
            warehouse_id = warehouse_id_form

            # Generate correlative code for the adjustment based on date.
            # Format: AIVyymmdd-N
            today = get_current_time_ve().date()
            date_prefix = f"AIV{today.strftime('%y%m%d')}"
            
            # Count existing adjustments for today to create a unique suffix
            count_today = InventoryAdjustment.query.filter(
                InventoryAdjustment.adjustment_code.like(f"{date_prefix}%")
            ).count()
            adjustment_code = f"{date_prefix}-{count_today + 1}"

            # Create the main adjustment record
            adjustment_record = InventoryAdjustment(
                adjustment_code=adjustment_code, reason=reason, user_id=current_user.id
            )
            db.session.add(adjustment_record)
            
            movements_to_create = []
            total_value_difference = 0.0
            
            for adj_str in adjustments:
                data = json.loads(adj_str)
                product_id = int(data['product_id'])
                real_stock = int(data['real_stock'])
                
                product = Product.query.get(product_id) # type: ignore
                product_stock_entry = ProductStock.query.filter_by(product_id=product_id, warehouse_id=warehouse_id).first()

                theoretical_stock = product_stock_entry.quantity if product_stock_entry else 0

                if product and theoretical_stock != real_stock:
                    difference = real_stock - theoretical_stock
                    
                    if not product_stock_entry:
                        product_stock_entry = ProductStock(product_id=product_id, warehouse_id=warehouse_id)
                        db.session.add(product_stock_entry)
                    product_stock_entry.quantity = real_stock

                    # Create adjustment item record
                    adj_item = InventoryAdjustmentItem(
                        adjustment_id=adjustment_record.id,
                        product_id=product.id,
                        theoretical_stock=theoretical_stock,
                        real_stock=real_stock,
                        comment=data.get('comment', '').strip() or None,
                        cost_at_adjustment_usd=product.cost_usd or 0.0
                    )
                    db.session.add(adj_item) # This line was missing
                    total_value_difference += difference * (product.cost_usd or 0.0)
                    
                    # Usar el código de ajuste en la descripción del movimiento
                    doc_identifier = adjustment_record.adjustment_code or f"Ajuste #{adjustment_record.id}"
                    # Preparar registro de movimiento
                    movement = Movement(
                        product_id=product.id,
                        type='Entrada' if difference > 0 else 'Salida',
                        warehouse_id=warehouse_id,
                        quantity=abs(difference),
                        document_id=adjustment_record.id,
                        document_type=f"Ajuste de Inventario ({doc_identifier})",
                        description=f"Motivo: {reason}. {data.get('comment', '')}".strip(),
                        date=get_current_time_ve()
                    )
                    movements_to_create.append(movement)

            # Save summary data to the main adjustment record
            adjustment_record.value_difference_usd = total_value_difference

            db.session.bulk_save_objects(movements_to_create)
            db.session.commit()
            flash(f'Ajuste de inventario completado y guardado con ID #{adjustment_record.id}.', 'success')
            
            log_user_activity(
                action="Realizó ajuste de inventario",
                details=f"Ajuste con código {adjustment_record.adjustment_code}. Motivo: {reason}",
                target_id=adjustment_record.adjustment_code,
                target_type="InventoryAdjustment"
            )
            return redirect(url_for('main.adjustment_result', adjustment_id=adjustment_record.id))

        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error al procesar el ajuste de inventario: {e}", exc_info=True)
            flash(f'Error al procesar el ajuste: {e}', 'danger')
            return redirect(url_for('main.inventory_adjustment'))

    # Cargar productos con su stock en el almacén seleccionado
    products_with_stock = db.session.query(Product, ProductStock.quantity).outerjoin(ProductStock, (Product.id == ProductStock.product_id) & (ProductStock.warehouse_id == warehouse_id)).filter(or_(Product.grupo != 'Ganchos', Product.grupo.is_(None))).order_by(Product.name).all()
    
    warehouses = Warehouse.query.order_by(Warehouse.id).all()
    selected_warehouse = Warehouse.query.get(warehouse_id)

    return render_template('inventario/ajuste_inventario.html', title='Ajuste de Inventario', products_data=products_with_stock, warehouses=warehouses, selected_warehouse=selected_warehouse)

@routes_blueprint.route('/inventario/ajustes/lista')
@login_required
def adjustment_list():
    """Displays a list of all inventory adjustments made."""
    if not is_administrador():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.new_order'))

    adjustments = InventoryAdjustment.query.options(
        joinedload(InventoryAdjustment.user)
    ).order_by(InventoryAdjustment.date.desc()).all()

    return render_template('inventario/list_ajustes.html',
                           title='Historial de Ajustes de Inventario',
                           adjustments=adjustments)

@routes_blueprint.route('/inventario/ajuste/resultado/<int:adjustment_id>')
@login_required
def adjustment_result(adjustment_id):
    """Muestra la página de resultados de un ajuste de inventario específico."""
    if not is_administrador():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.new_order'))

    adjustment = InventoryAdjustment.query.options(
        joinedload(InventoryAdjustment.items).joinedload(InventoryAdjustmentItem.product),
        joinedload(InventoryAdjustment.user)
    ).get_or_404(adjustment_id)

    # KPIs - NEW LOGIC: Calculate total inventory value before and after the adjustment.
    # 1. Calculate the total value of the entire inventory *after* the adjustment (current state).
    # We exclude 'Ganchos' as they are supplies, not for sale.
    total_inventory_value_after_query = db.session.query( # type: ignore
        func.sum(Product.stock * Product.cost_usd)
    ).filter(or_(Product.grupo != 'Ganchos', Product.grupo.is_(None))).first()
    value_after = total_inventory_value_after_query[0] or 0.0

    # 2. Calculate the value *before* by subtracting the adjustment's impact.
    value_before = value_after - adjustment.value_difference_usd
    value_diff = value_after - value_before

    # Highlights
    most_impactful_items = sorted(adjustment.items, key=lambda x: abs((x.real_stock - x.theoretical_stock) * x.cost_at_adjustment_usd), reverse=True)[:5]

    # Previous adjustments for comparison
    previous_adjustments = InventoryAdjustment.query.filter(InventoryAdjustment.id != adjustment_id).order_by(InventoryAdjustment.date.desc()).limit(5).all()

    return render_template('inventario/ajuste_resultado.html',
                           title=f'Resultado del Ajuste #{adjustment.id}',
                           adjustment=adjustment, value_before=value_before, value_after=value_after, value_diff=value_diff,
                           most_impactful_items=most_impactful_items, previous_adjustments=previous_adjustments)

@routes_blueprint.route('/inventario/producto/<int:product_id>')
@login_required
def product_detail(product_id):
    product = Product.query.options(
        joinedload(Product.stock_levels).joinedload(ProductStock.warehouse)
    ).get_or_404(product_id)
    return render_template('inventario/detalle_producto.html', title=product.name, product=product)



@routes_blueprint.route('/inventario/nuevo', methods=['GET', 'POST'])
@login_required
def new_product():
    if not is_administrador(): # Superusuario and Gerente can create products
        flash('Acceso denegado. Solo los administradores pueden realizar esta acción.', 'danger')
        return redirect(url_for('main.new_order'))

    if request.method == 'POST':
        try:
            name = request.form.get('name')
            description = request.form.get('description')
            barcode = request.form.get('barcode')
            qr_code = request.form.get('qr_code')
            image_url = request.form.get('image_url')
            size = request.form.get('size')
            color = request.form.get('color')
            codigo_producto = request.form.get('codigo_producto')
            marca = request.form.get('marca')
            grupo = request.form.get('grupo')
            cost_usd = float(request.form.get('cost_usd'))
            price_usd = float(request.form.get('price_usd'))

            new_prod = Product(
                name=name, description=description, barcode=barcode, qr_code=qr_code,
                image_url=image_url, size=size, color=color, cost_usd=cost_usd, price_usd=price_usd,
                codigo_producto=codigo_producto, marca=marca, grupo=grupo
            )
            db.session.add(new_prod)
            db.session.commit()

            log_user_activity(
                action="Creó nuevo producto",
                details=f"Producto: {new_prod.name} (Cód. Barra: {new_prod.barcode})",
                target_id=new_prod.id,
                target_type="Product"
            )
            flash('Producto creado exitosamente!', 'success')
            return redirect(url_for('main.inventory_list'))
        except (ValueError, IntegrityError) as e:
            db.session.rollback()
            flash(f'Error al crear el producto: {str(e)}', 'danger')
    return render_template('inventario/nuevo.html', title='Nuevo Producto')

@routes_blueprint.route('/inventario/editar/<int:product_id>', methods=['GET', 'POST'])
@login_required
def edit_product(product_id):
    if not is_gerente(): # Solo Superusuario y Gerente pueden editar
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.inventory_list'))

    product = Product.query.get_or_404(product_id)

    if request.method == 'POST':
        try:
            # Manejar la subida de la imagen
            if 'image_file' in request.files:
                image_file = request.files['image_file']
                if image_file.filename != '':
                    image_path = save_product_image(image_file)
                    product.image_url = url_for('static', filename=image_path)

            # Actualizar campos desde el formulario
            product.name = request.form.get('name')
            product.description = request.form.get('description')
            product.size = request.form.get('size')
            product.color = request.form.get('color')
            product.marca = request.form.get('marca')
            product.grupo = request.form.get('grupo')

            db.session.commit()
            log_user_activity(
                action="Modificó producto",
                details=f"Producto: {product.name}",
                target_id=product.id,
                target_type="Product"
            )

            flash(f'Producto "{product.name}" actualizado exitosamente.', 'success')
            return redirect(url_for('main.product_detail', product_id=product.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Ocurrió un error al actualizar el producto: {e}', 'danger')

    # Obtener todos los grupos para el dropdown
    groups = db.session.query(Product.grupo).distinct().order_by(Product.grupo).all()
    product_groups = [g[0] for g in groups if g[0]]

    return render_template('inventario/editar_producto.html', 
                           title=f'Editar {product.name}', 
                           product=product,
                           product_groups=product_groups)

# Rutas de clientes
@routes_blueprint.route('/clientes/lista')
@login_required
def client_list():
    clients = Client.query.all()
    return render_template('clientes/lista.html', title='Lista de Clientes', clients=clients)

@routes_blueprint.route('/clientes/nuevo', methods=['GET', 'POST'])
@login_required
def new_client():
    if request.method == 'POST':
        name = request.form.get('name')
        cedula_rif = request.form.get('cedula_rif')
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone')
        address = request.form.get('address')

        # Treat empty email as None
        if not email:
            email = None

        try:
            # Check for duplicate email only if it's provided
            if email and Client.query.filter_by(email=email).first():
                raise IntegrityError("El email ya está registrado.")

            new_cli = Client(name=name, cedula_rif=cedula_rif, email=email, phone=phone, address=address)
            db.session.add(new_cli)
            db.session.commit()

            log_user_activity(
                action="Creó nuevo cliente",
                details=f"Cliente: {new_cli.name} (CI/RIF: {new_cli.cedula_rif or 'N/A'})",
                target_id=new_cli.id,
                target_type="Client"
            )

            flash('Cliente creado exitosamente!', 'success')
            return redirect(url_for('main.client_list'))
        except IntegrityError as e:
            db.session.rollback()
            flash(f'Error: {e}', 'danger')
    return render_template('clientes/nuevo.html', title='Nuevo Cliente')

@routes_blueprint.route('/clientes/detalle/<int:client_id>', methods=['GET', 'POST'])
@login_required
def client_detail(client_id):
    client = Client.query.get_or_404(client_id)
    orders_query = Order.query.filter_by(client_id=client.id).options(subqueryload(Order.payments)).order_by(Order.date_created.desc())

    # Calculate total due before any potential new payment
    total_due_pre_payment = sum(order.due_amount_usd for order in orders_query.all() if order.due_amount_usd > 0)
    
    if request.method == 'POST':
        # This handles adding a payment to an order from the client detail page
        order_id = request.form.get('order_id')
        order = Order.query.get_or_404(order_id)
        
        try:
            # Para abonos, siempre usar la tasa de cambio actual para calcular el equivalente en USD. # type: ignore
            current_rate = get_cached_exchange_rate('USD') or 1.0
            payment_data_json = request.form.get('payments_data')
            payment_info = json.loads(payment_data_json)[0]
            payment_info['amount_usd_equivalent'] = float(payment_info['amount_ves_equivalent']) / current_rate if current_rate > 0 else 0

            payment_date = get_current_time_ve()
            if payment_info.get('date'):
                try:
                    naive_dt = datetime.strptime(payment_info['date'], '%Y-%m-%dT%H:%M')
                    payment_date = VE_TIMEZONE.localize(naive_dt)
                except (ValueError, TypeError):
                    current_app.logger.warning(f"Invalid payment date format: '{payment_info['date']}'. Falling back to now.")

            payment = Payment(
                order_id=order.id,
                amount_paid=payment_info['amount_paid'],
                currency_paid=payment_info['currency_paid'],
                amount_ves_equivalent=payment_info['amount_ves_equivalent'],
                amount_usd_equivalent=payment_info.get('amount_usd_equivalent', 0.0),
                method=payment_info['method'],
                reference=payment_info.get('reference'),
                issuing_bank=payment_info.get('issuing_bank'),
                sender_id=payment_info.get('sender_id'),
                date=payment_date,
                bank_id=payment_info.get('bank_id'),
                pos_id=payment_info.get('pos_id'),
                cash_box_id=payment_info.get('cash_box_id')
            )
            db.session.add(payment)

            # Update account balances
            if payment.bank_id:
                bank = Bank.query.get(payment.bank_id)
                if bank:
                    # Payments to banks are always registered as their VES equivalent for accounting
                    # but the balance update must respect the bank's currency.
                    if bank.currency == 'VES': bank.balance += payment.amount_ves_equivalent
            elif payment.pos_id:
                pos = PointOfSale.query.get(payment.pos_id)
                if pos and pos.bank: pos.bank.balance += payment.amount_ves_equivalent
            elif payment.cash_box_id:
                cash_box = CashBox.query.get(payment.cash_box_id)
                if cash_box:
                    if payment.currency_paid == 'VES': cash_box.balance_ves += payment.amount_paid
                    elif payment.currency_paid == 'USD': cash_box.balance_usd += payment.amount_paid

            db.session.flush() # Flush to calculate new due amount

            # Update order status if it's now fully paid
            if order.due_amount <= 0.01:
                order.status = 'Pagada'
            
            log_user_activity(
                action="Registró abono",
                details=f"Abono de {payment.amount_paid:.2f} {payment.currency_paid} a la orden #{order.id:09d} del cliente '{order.client.name}'",
                target_id=order.id,
                target_type="Order"
            )
            db.session.commit()
            flash(f'Abono registrado exitosamente para la orden #{order.id:09d}.', 'success')
            # After commit, the order.due_amount is updated. We can now recalculate total_due.
            total_due_post_payment = sum(o.due_amount_usd for o in orders_query.all() if o.due_amount_usd > 0)
            total_due_pre_payment = total_due_post_payment # Update the variable to be passed to the template
        except (ValueError, KeyError, IndexError, TypeError) as e:
            db.session.rollback()
            current_app.logger.error(f"Error registrando abono: {e}")
            flash(f'Error al registrar el abono: {e}', 'danger')
        return redirect(url_for('main.client_detail', client_id=client.id))

    orders = orders_query.all()

    # For the payment modal
    banks = Bank.query.order_by(Bank.name).all()
    points_of_sale = PointOfSale.query.order_by(PointOfSale.name).all()
    cash_boxes = CashBox.query.order_by(CashBox.name).all()

    return render_template('clientes/detalle_cliente.html',
                           title=f'Detalle de Cliente: {client.name}',
                           client=client,
                           orders=orders, # This now contains the updated order states
                           total_due=total_due_pre_payment, # This is now the updated total due
                           banks=banks,
                           points_of_sale=points_of_sale,
                           cash_boxes=cash_boxes)

# Rutas de proveedores
@routes_blueprint.route('/proveedores/lista')
@login_required
def provider_list():
    if not is_administrador(): # Superusuario and Gerente can view providers
        flash('Acceso denegado. Solo los administradores pueden ver esta sección.', 'danger')
        return redirect(url_for('main.new_order'))

    providers = Provider.query.all()
    return render_template('proveedores/lista.html', title='Lista de Proveedores', providers=providers)

@routes_blueprint.route('/proveedores/nuevo', methods=['GET', 'POST'])
@login_required
def new_provider():
    if not is_administrador(): # Superusuario and Gerente can create providers
        flash('Acceso denegado. Solo los administradores pueden realizar esta acción.', 'danger')
        return redirect(url_for('main.new_order'))

    if request.method == 'POST':
        try:
            new_prov = Provider(
                # Información básica
                name=request.form.get('name'),
                tax_id=request.form.get('tax_id'),
                address=request.form.get('address'),
                phone=request.form.get('phone'),
                fax=request.form.get('fax'),
                email=request.form.get('email'),
                # Contacto principal
                contact_person_name=request.form.get('contact_person_name'),
                contact_person_phone=request.form.get('contact_person_phone'),
                contact_person_email=request.form.get('contact_person_email'),
                # Información bancaria
                bank_name=request.form.get('bank_name'),
                bank_branch=request.form.get('bank_branch'),
                bank_account_number=request.form.get('bank_account_number'),
                bank_account_currency=request.form.get('bank_account_currency'),
                bank_swift_bic=request.form.get('bank_swift_bic'),
                bank_iban=request.form.get('bank_iban'),
                # Detalles del negocio y términos
                business_description=request.form.get('business_description'),
                payment_terms=request.form.get('payment_terms'),
                shipping_terms=request.form.get('shipping_terms')
            )
            db.session.add(new_prov)
            log_user_activity(
                action="Creó nuevo proveedor",
                details=f"Proveedor: {new_prov.name}",
                target_id=new_prov.id,
                target_type="Provider"
            )
            db.session.commit()
            flash('Proveedor creado exitosamente!', 'success')
            return redirect(url_for('main.provider_list'))
        except IntegrityError as e:
            db.session.rollback()
            current_app.logger.error(f"Error de integridad al crear proveedor: {e}")
            flash('Error: Ya existe un proveedor con ese nombre o RIF.', 'danger')
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error inesperado al crear proveedor: {e}")
            flash(f'Ocurrió un error inesperado: {e}', 'danger')

    return render_template('proveedores/nuevo.html', title='Nuevo Proveedor')

@routes_blueprint.route('/proveedores/detalle/<int:provider_id>')
@login_required
def provider_detail(provider_id):
    if not is_administrador():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.new_order'))
    provider = Provider.query.get_or_404(provider_id)
    return render_template('proveedores/detalle_proveedor.html', title=f'Detalle de {provider.name}', provider=provider)

# Rutas de compras
@routes_blueprint.route('/compras/lista')
@login_required
def purchase_list():
    if not is_administrador(): # Superusuario and Gerente can view purchases
        flash('Acceso denegado. Solo los administradores pueden ver esta sección.', 'danger')
        return redirect(url_for('main.new_order'))

    purchases = Purchase.query.all()
    return render_template('compras/lista.html', title='Lista de Compras', purchases=purchases)

@routes_blueprint.route('/compras/detalle/<int:purchase_id>')
@login_required
def purchase_detail(purchase_id):
    if not is_administrador(): # Superusuario and Gerente can view purchase details
        flash('Acceso denegado. Solo los administradores pueden ver esta sección.', 'danger')
        return redirect(url_for('main.new_order'))

    purchase = Purchase.query.get_or_404(purchase_id)
    return render_template('compras/detalle_compra.html', title=f'Compra #{purchase.id}', purchase=purchase)

@routes_blueprint.route('/compras/nuevo', methods=['GET', 'POST'])
@login_required
def new_purchase():
    if not is_administrador(): # Superusuario and Gerente can create purchases
        flash('Acceso denegado. Solo los administradores pueden realizar esta acción.', 'danger')
        return redirect(url_for('main.new_order'))

    # Convertir proveedores a una lista de diccionarios para que sea serializable a JSON y usable en JS
    providers_query = Provider.query.order_by(Provider.name).all()
    providers = [
        {
            'id': p.id, 'name': p.name, 'tax_id': p.tax_id,
            'phone': p.phone, 'email': p.email, 'address': p.address
        }
        for p in providers_query
    ]

    
    # Convertir productos a una lista de diccionarios para que sea serializable a JSON
    products_query = Product.query.order_by(Product.name).all()
    products = [
        {
            'id': p.id, 'name': p.name, 'barcode': p.barcode, 
            'cost_usd': p.cost_usd, 'stock': p.stock
        } 
        for p in products_query
    ]

    banks = Bank.query.order_by(Bank.name).all()
    cash_boxes = CashBox.query.order_by(CashBox.name).all()

    calculation_currency, _ = get_main_calculation_currency_info()
    current_rate = get_cached_exchange_rate(calculation_currency)

    if current_rate is None:
        flash('No se ha podido obtener la tasa de cambio. No se pueden crear compras en este momento.', 'danger')
        return redirect(url_for('main.purchase_list'))

    if request.method == 'POST':
        provider_id = request.form.get('provider_id')
        product_ids = request.form.getlist('product_id[]')
        quantities = request.form.getlist('quantity[]')
        costs_usd = request.form.getlist('cost_usd[]')
        payments_data_json = request.form.get('payments_data')
        payments_data = json.loads(payments_data_json) if payments_data_json else []

        try:
            # Create Purchase and Items
            new_purchase = Purchase(provider_id=provider_id, total_cost=0)
            db.session.add(new_purchase)
            db.session.flush()
            
            total_cost_ves = 0
            for p_id, q, c_usd in zip(product_ids, quantities, costs_usd):
                product = Product.query.get(p_id)
                quantity = int(q)
                cost_usd = float(c_usd)
                if product and quantity > 0 and cost_usd >= 0:
                    cost_ves = cost_usd * current_rate
                    item = PurchaseItem(
                        purchase_id=new_purchase.id,
                        product_id=p_id,
                        quantity=quantity,
                        cost=cost_ves
                    )
                    db.session.add(item)
                    total_cost_ves += cost_ves * quantity
            
            new_purchase.total_cost = total_cost_ves
            db.session.flush()

            # Process Payments (as ManualFinancialMovement with type 'Egreso')
            total_paid_ves = 0
            for payment_info in payments_data:
                amount_paid = float(payment_info['amount_paid'])
                currency_paid = payment_info['currency_paid']
                amount_ves_equivalent = float(payment_info['amount_ves_equivalent'])
                
                movement = ManualFinancialMovement(
                    description=f"Pago por Orden de Compra #{new_purchase.id}",
                    amount=amount_paid, currency=currency_paid, movement_type='Egreso',
                    status='Aprobado', purchase_id=new_purchase.id,
                    created_by_user_id=current_user.id, approved_by_user_id=current_user.id,
                    date_approved=get_current_time_ve(), bank_id=payment_info.get('bank_id'),
                    cash_box_id=payment_info.get('cash_box_id')
                )
                db.session.add(movement)
                total_paid_ves += amount_ves_equivalent

                # Decrease balance of the corresponding account
                if movement.bank_id:
                    bank = Bank.query.get(movement.bank_id)
                    if bank:
                        # Payments from banks are always registered as their VES equivalent for accounting
                        # but the balance update must respect the bank's currency.
                        if bank.currency == 'VES': bank.balance -= amount_ves_equivalent
                elif movement.cash_box_id:
                    cash_box = CashBox.query.get(movement.cash_box_id)
                    if cash_box:
                        if currency_paid == 'VES': cash_box.balance_ves -= amount_paid
                        elif currency_paid == 'USD': cash_box.balance_usd -= amount_paid
            
            # Update Purchase Payment Status
            if total_paid_ves >= total_cost_ves - 0.01:
                new_purchase.payment_status = 'Pagada'
            elif total_paid_ves > 0.01:
                new_purchase.payment_status = 'Abonada'
            else:
                new_purchase.payment_status = 'Pendiente de Pago'
            
            notification_message = f"Nueva Orden de Compra #{new_purchase.id} creada."
            notification_link = url_for('main.purchase_detail', purchase_id=new_purchase.id)
            create_notification_for_admins(notification_message, notification_link)

            log_user_activity(
                action="Creó orden de compra",
                details=f"Orden de compra para '{new_purchase.provider.name}' por un total de {new_purchase.total_cost:.2f} Bs.",
                target_id=new_purchase.id,
                target_type="Purchase"
            )

            db.session.commit()
            flash('Compra creada exitosamente!', 'success')
            return redirect(url_for('main.purchase_list'))
        except (ValueError, IntegrityError) as e:
            db.session.rollback()
            flash(f'Error al crear la compra: {str(e)}', 'danger')
            # Fall through to render the template again
    
    return render_template('compras/nuevo.html', 
                           title='Nueva Compra', 
                           providers=providers, 
                           products=products,
                           banks=banks,
                           cash_boxes=cash_boxes,
                           current_rate=current_rate)

# Rutas de recepciones
@routes_blueprint.route('/recepciones/lista')
@login_required
def reception_list():
    if not is_administrador(): # Superusuario and Gerente can view receptions
        flash('Acceso denegado. Solo los administradores pueden ver esta sección.', 'danger')
        return redirect(url_for('main.new_order'))

    receptions = Reception.query.order_by(Reception.date_received.desc()).all()
    return render_template('recepciones/lista.html', title='Lista de Recepciones', receptions=receptions)

@routes_blueprint.route('/api/purchase_details/<int:purchase_id>')
@login_required
def api_purchase_details(purchase_id):
    if not is_administrador(): # Superusuario, Gerente y administrador can view purchase details via API
        return jsonify({'error': 'Acceso denegado'}), 403
    
    purchase = Purchase.query.options(
        subqueryload(Purchase.items).joinedload(PurchaseItem.product) # type: ignore
    ).get(purchase_id)

    if not purchase:
        return jsonify({'error': 'Compra no encontrada'}), 404

    items_data = []
    for item in purchase.items:
        items_data.append({
            'product_id': item.product_id,
            'product_name': item.product.name,
            'quantity_ordered': item.quantity,
            'quantity_received': item.quantity_received,
            'quantity_pending': item.quantity_pending
        })
    
    return jsonify(items=items_data)

@routes_blueprint.route('/recepciones/nueva', methods=['GET', 'POST'])
@login_required
def new_reception():
    if not is_administrador(): # Superusuario, Gerente y administrador can create receptions
        flash('Acceso denegado. Solo los administradores pueden realizar esta acción.', 'danger')
        return redirect(url_for('main.new_order'))

    if request.method == 'POST':
        try:
            purchase_id = request.form.get('purchase_id')
            warehouse_id = request.form.get('warehouse_id', type=int)
            product_ids = request.form.getlist('product_id[]')
            quantities_received_str = request.form.getlist('quantity_received[]')

            if not purchase_id:
                raise ValueError("No se ha seleccionado una orden de compra.")
            if not warehouse_id:
                raise ValueError("Debe seleccionar un almacén de destino.")

            purchase = Purchase.query.get_or_404(purchase_id)
            
            reception = Reception(purchase_id=purchase.id, status='Parcial')
            db.session.add(reception)
            db.session.flush()

            total_items_received_in_this_tx = 0
            for p_id, qty_rec_str in zip(product_ids, quantities_received_str):
                qty_received = int(qty_rec_str) if qty_rec_str else 0
                if qty_received <= 0:
                    continue

                item = PurchaseItem.query.filter_by(purchase_id=purchase.id, product_id=p_id).first()
                if not item:
                    current_app.logger.warning(f"Intento de recibir producto {p_id} que no está en la compra {purchase.id}")
                    continue

                if qty_received > item.quantity_pending:
                    raise ValueError(f"Intenta recibir más unidades de '{item.product.name}' de las pendientes ({item.quantity_pending}).")

                product = item.product
                
                # Actualizar stock en el almacén de destino
                stock_entry = ProductStock.query.filter_by(product_id=product.id, warehouse_id=warehouse_id).first()
                if not stock_entry:
                    stock_entry = ProductStock(product_id=product.id, warehouse_id=warehouse_id, quantity=0)
                    db.session.add(stock_entry)
                stock_entry.quantity += qty_received

                item.quantity_received += qty_received
                total_items_received_in_this_tx += qty_received
                
                movement = Movement(
                    product_id=product.id,
                    type='Entrada', warehouse_id=warehouse_id,
                    quantity=qty_received,
                    document_id=reception.id,
                    document_type='Recepción de Compra',
                    related_party_id=purchase.provider_id, # type: ignore
                    related_party_type='Proveedor',
                    date=reception.date_received # Asegurar que el movimiento tenga la misma fecha que la recepción
                )
                db.session.add(movement)

            if total_items_received_in_this_tx == 0:
                db.session.rollback()
                flash('No se recibieron productos. No se ha creado la recepción.', 'warning')
                return redirect(url_for('main.new_reception'))

            total_ordered = sum(i.quantity for i in purchase.items)
            total_received = sum(i.quantity_received for i in purchase.items)

            if total_received >= total_ordered:
                purchase.status = 'Recibida'
                reception.status = 'Completada'
            else:
                purchase.status = 'Recibida Parcialmente'
            
            notification_message = f"Recepción para la compra #{purchase.id} procesada."
            notification_link = url_for('main.reception_list')
            create_notification_for_admins(notification_message, notification_link)

            log_user_activity(
                action="Procesó recepción de mercancía",
                details=f"Recepción para la compra #{purchase.id} del proveedor '{purchase.provider.name}'.",
                target_id=reception.id,
                target_type="Reception"
            )

            db.session.commit()
            flash('Recepción completada y stock actualizado!', 'success')
            return redirect(url_for('main.reception_list'))
        except (ValueError, IntegrityError) as e:
            db.session.rollback()
            flash(f'Error al procesar la recepción: {str(e)}', 'danger')
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error inesperado en recepción: {e}")
            flash(f'Ocurrió un error inesperado: {str(e)}', 'danger')

    pending_purchases = Purchase.query.filter(
        Purchase.status.in_(['Pendiente', 'Recibida Parcialmente'])
    ).order_by(Purchase.id.desc()).all()
    warehouses = Warehouse.query.order_by(Warehouse.id).all()

    return render_template('recepciones/nueva.html', 
                           title='Nueva Recepción', 
                           purchases=pending_purchases,
                           warehouses=warehouses)


# Rutas de órdenes
@routes_blueprint.route('/ordenes/lista')
@login_required
def order_list():
    # Get filter parameters
    search_term = request.args.get('search', '').strip()
    status_filter = request.args.get('status', '').strip()
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    # Set default date range (current month)
    today = get_current_time_ve().date()
    if not start_date_str:
        start_date_str = today.replace(day=1).strftime('%Y-%m-%d')
    if not end_date_str:
        end_date_str = today.strftime('%Y-%m-%d')

    # Base query
    query = Order.query.options(
        joinedload(Order.client),
        subqueryload(Order.payments)
    ).join(Client).order_by(Order.date_created.desc())

    # Apply search filter (Order ID or Client Name)
    if search_term:
        search_pattern = f'%{search_term}%'
        query = query.filter(or_( # The join(Client) is necessary for this filter
            Client.name.ilike(search_pattern),
            Order.id.cast(db.String).ilike(search_pattern)
        ))

    # Apply status filter
    if status_filter:
        if status_filter == 'credito':
            query = query.filter(Order.status == 'Crédito')
        elif status_filter == 'apartado':
            query = query.filter(Order.status == 'Apartado')
        elif status_filter == 'contado':
            query = query.filter(Order.status.in_(['Pagada', 'Completada']))
        elif status_filter == 'con_deuda':
            # Use a subquery to calculate paid amount and filter where due amount > 0
            paid_subquery = select(func.sum(Payment.amount_ves_equivalent)).where(Payment.order_id == Order.id).correlate(Order).as_scalar()
            query = query.filter(Order.total_amount - func.coalesce(paid_subquery, 0) > 0.01)

    # Apply date range filter
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        start_dt = VE_TIMEZONE.localize(datetime.combine(start_date, datetime.min.time()))
        end_dt = VE_TIMEZONE.localize(datetime.combine(end_date, datetime.max.time()))
        
        query = query.filter(Order.date_created.between(start_dt, end_dt))
    except (ValueError, TypeError):
        flash('Formato de fecha inválido. Usando rango por defecto.', 'warning')
        start_date_str = today.replace(day=1).strftime('%Y-%m-%d')
        end_date_str = today.strftime('%Y-%m-%d')

    orders = query.all()

    filters = { 'search': search_term, 'status': status_filter, 'start_date': start_date_str, 'end_date': end_date_str }

    return render_template('ordenes/lista.html', title='Lista de Órdenes', orders=orders, filters=filters)

@routes_blueprint.route('/ordenes/detalle/<int:order_id>')
@login_required
def order_detail(order_id):
    order = Order.query.get_or_404(order_id)
    company_info = CompanyInfo.query.first()
    # IVA desactivado
    subtotal = sum(item.price * item.quantity for item in order.items)
    iva = 0
    return render_template('ordenes/detalle_orden.html',
                           title=f'Orden #{order.id}',
                           order=order,
                           company_info=company_info,
                           subtotal=subtotal,
                           iva=iva)

@routes_blueprint.route('/ordenes/nuevo', methods=['GET', 'POST'])
@login_required
def new_order():
    clients = Client.query.all()
    products = Product.query.all()
    calculation_currency, _ = get_main_calculation_currency_info()
    banks = Bank.query.order_by(Bank.name).all()
    points_of_sale = PointOfSale.query.order_by(PointOfSale.name).all()
    cash_boxes = CashBox.query.order_by(CashBox.name).all()

    current_rate = get_cached_exchange_rate(calculation_currency)
    
    if current_rate is None:
        flash('No se ha podido obtener la tasa de cambio. No se pueden crear órdenes en este momento.', 'danger') # type: ignore
        return redirect(url_for('main.order_list'))
    
    if request.method == 'POST':
        client_id = request.form.get('client_id')
        date_created_str = request.form.get('date_created')
        special_rate_str = request.form.get('special_exchange_rate')
        product_ids = request.form.getlist('product_id[]')
        quantities = request.form.getlist('quantity[]')
        prices_usd = request.form.getlist('price_usd[]')
        payments_data_json = request.form.get('payments_data')
        sale_type = request.form.get('sale_type', 'regular')
        payments_data = json.loads(payments_data_json) if payments_data_json else []
        dispatch_reason = request.form.get('dispatch_reason', '').strip()
        
        rate_for_order = current_rate
        if is_gerente() and special_rate_str: # Only Gerente and Superuser can set special rate
            try:
                special_rate = float(special_rate_str)
                if special_rate > 0:
                    rate_for_order = special_rate
                    flash(f'¡Atención! Se está usando una tasa de cambio especial para esta orden: {rate_for_order}', 'info')
            except (ValueError, TypeError):
                flash('La tasa de cambio especial no es un número válido. Usando tasa actual.', 'warning')

        discount_enabled = request.form.get('discount_enabled') == 'on'
        discount_usd = float(request.form.get('discount_usd', 0.0)) if discount_enabled else 0.0

        try:
            sequence_map = {
                'regular': 'order_contado_seq',
                'credit': 'order_credito_seq',
                'reservation': 'order_apartado_seq',
                'special_dispatch': 'order_entrega_especial_seq'
            }
            sequence_name = sequence_map.get(sale_type)
            if not sequence_name:
                raise ValueError("Tipo de venta no válido.")
            
            if sale_type == 'special_dispatch':
                if not dispatch_reason:
                    raise ValueError("El motivo de la entrega es obligatorio para una Entrega Especial.")
                # For special dispatch, totals are zero and no payments are allowed.
                discount_usd = 0.0

            next_id = db.session.execute(text(f"SELECT nextval('{sequence_name}')")).scalar()
            
            order_total_usd_before_discount = sum(int(q) * float(p_usd) for q, p_usd in zip(quantities, prices_usd))
            final_order_total_usd = order_total_usd_before_discount - discount_usd

            order_total_ves_before_discount = order_total_usd_before_discount * rate_for_order
            discount_ves = discount_usd * rate_for_order
            final_order_total_ves = order_total_ves_before_discount - discount_ves

            # Financial validation only for sales, not for special dispatches
            if sale_type != 'special_dispatch':
                paid_total_usd = 0
                paid_total_ves = 0
                for p in payments_data:
                    if p['currency_paid'] == 'USD':
                        p['amount_ves_equivalent'] = float(p['amount_paid']) * rate_for_order
                    else:
                        p['amount_ves_equivalent'] = float(p['amount_paid'])
                    paid_total_ves += float(p['amount_ves_equivalent'])
                    # Calcular el equivalente en USD para el abono
                    p['amount_usd_equivalent'] = p['amount_ves_equivalent'] / rate_for_order
                    paid_total_usd += p['amount_usd_equivalent']
                
                if sale_type == 'regular' and paid_total_ves < final_order_total_ves - 0.01:
                    raise ValueError(f"El monto pagado (Bs. {paid_total_ves:.2f}) es menor que el total de la orden (Bs. {final_order_total_ves:.2f}).")

            order_date = get_current_time_ve()
            if date_created_str:
                try:
                    naive_dt = datetime.strptime(date_created_str, '%Y-%m-%dT%H:%M')
                    order_date = VE_TIMEZONE.localize(naive_dt)
                except (ValueError, TypeError):
                    current_app.logger.warning(f"Invalid date_created format: '{date_created_str}'. Falling back.")

            # --- Performance Optimization: Pre-fetch all products ---
            unique_product_ids = [pid for pid in product_ids if pid]
            products_from_db = Product.query.filter(Product.id.in_(unique_product_ids)).all()
            product_map = {str(p.id): p for p in products_from_db}

            # --- Stock validation before creating the order ---
            # For special dispatches, we only validate stock if the user is a manager (immediate dispatch)
            should_validate_stock_now = sale_type != 'special_dispatch' or (sale_type == 'special_dispatch' and is_gerente())
            if should_validate_stock_now:
                for p_id, q in zip(product_ids, quantities):
                    quantity = int(q)
                    product = product_map.get(p_id) # type: ignore
                    if not product or quantity <= 0: # type: ignore
                        continue
                    # La venta siempre es desde el almacén principal (ID 1)
                    if product.stock_tienda < quantity:
                        raise ValueError(f'Stock insuficiente en Tienda para "{product.name}". Solicitado: {quantity}, Disponible: {product.stock_tienda}.')

            new_order = Order(
                id=next_id, client_id=client_id, status='Pendiente', total_amount=0, 
                total_amount_usd=final_order_total_usd, discount_usd=discount_usd, 
                exchange_rate_at_sale=rate_for_order,
                date_created=order_date, order_type=sale_type
            )
            # Override totals for special dispatch
            if sale_type == 'special_dispatch':
                new_order.dispatch_reason = dispatch_reason
                new_order.total_amount_usd = 0.0
                new_order.discount_usd = 0.0
                new_order.total_amount = 0.0

            db.session.add(new_order)
            db.session.flush()

            total_amount = 0
            for p_id, q, p_usd in zip(product_ids, quantities, prices_usd):
                product = product_map.get(p_id) # type: ignore
                quantity = int(q)
                if not product or quantity <= 0:
                    continue

                price_ves = float(p_usd) * rate_for_order
                cost_ves = product.cost_usd * rate_for_order if product.cost_usd else 0
                
                item = OrderItem(order_id=new_order.id, product_id=p_id, quantity=quantity, price=price_ves, cost_at_sale_ves=cost_ves)
                db.session.add(item)
                
                # --- Inventory Movement Logic ---
                # Only move inventory if it's not a pending special dispatch
                should_move_inventory = sale_type != 'special_dispatch' or (sale_type == 'special_dispatch' and is_gerente())
                if should_move_inventory:
                    # Descontar stock del almacén principal (ID 1)
                    main_store_stock = ProductStock.query.filter_by(product_id=product.id, warehouse_id=1).first()
                    if not main_store_stock or main_store_stock.quantity < quantity:
                         raise ValueError(f"Error de consistencia de stock para {product.name}. Intente de nuevo.")
                    main_store_stock.quantity -= quantity

                    # Registrar movimiento de salida desde el almacén principal
                    document_type = 'Entrega Especial' if sale_type == 'special_dispatch' else 'Orden de Venta'
                    movement = Movement(product_id=product.id, type='Salida', warehouse_id=1, quantity=quantity, document_id=new_order.id, document_type=document_type, description=f"Venta al cliente #{new_order.client_id}", related_party_id=new_order.client_id, related_party_type='Cliente', date=order_date)
                    db.session.add(movement)
                
                total_amount += price_ves * quantity

            new_order.total_amount = total_amount - discount_ves
            db.session.flush()

            if sale_type == 'special_dispatch':
                if is_gerente():
                    new_order.status = 'Completada'
                else:
                    new_order.status = 'Pendiente de Aprobación'
            elif sale_type == 'reservation': new_order.status = 'Apartado'
            elif sale_type == 'credit': new_order.status = 'Crédito'
            else: new_order.status = 'Pagada'

            # --- Performance Optimization: Pre-fetch all financial accounts ---
            bank_ids = {p.get('bank_id') for p in payments_data if p.get('bank_id')}
            pos_ids = {p.get('pos_id') for p in payments_data if p.get('pos_id')}
            cash_box_ids = {p.get('cash_box_id') for p in payments_data if p.get('cash_box_id')}

            banks_map = {b.id: b for b in Bank.query.filter(Bank.id.in_(bank_ids))}
            pos_map = {p.id: p for p in PointOfSale.query.filter(PointOfSale.id.in_(pos_ids)).options(joinedload(PointOfSale.bank))}
            cash_box_map = {c.id: c for c in CashBox.query.filter(CashBox.id.in_(cash_box_ids))}
            # --- End Optimization ---

            if sale_type != 'special_dispatch':
                for payment_info in payments_data:
                    payment = Payment(
                        order_id=new_order.id, amount_paid=payment_info['amount_paid'], currency_paid=payment_info['currency_paid'], # type: ignore
                        amount_ves_equivalent=payment_info['amount_ves_equivalent'], amount_usd_equivalent=payment_info['amount_usd_equivalent'], method=payment_info['method'],
                        reference=payment_info.get('reference'), issuing_bank=payment_info.get('issuing_bank'),
                        sender_id=payment_info.get('sender_id'), date=order_date, # Use the order's date for the payment
                        bank_id=payment_info.get('bank_id'),
                        pos_id=payment_info.get('pos_id'), cash_box_id=payment_info.get('cash_box_id')
                    )
                    db.session.add(payment)

                    if payment.bank_id and payment.bank_id in banks_map:
                        banks_map[payment.bank_id].balance += payment.amount_ves_equivalent
                    elif payment.pos_id and payment.pos_id in pos_map:
                        pos = pos_map[payment.pos_id] # POS terminals are assumed to be in VES
                        if pos.bank: pos.bank.balance += payment.amount_ves_equivalent
                    elif payment.cash_box_id and payment.cash_box_id in cash_box_map:
                        cash_box = cash_box_map[payment.cash_box_id]
                        if payment.currency_paid == 'VES': cash_box.balance_ves += payment.amount_paid
                        elif payment.currency_paid == 'USD': cash_box.balance_usd += payment.amount_paid

            # --- Notifications ---
            if sale_type == 'special_dispatch' and not is_gerente():
                notification_message = f"El usuario {current_user.username} solicita aprobación para una Entrega Especial."
                notification_link = url_for('main.pending_dispatches')
                create_notification_for_admins(notification_message, notification_link)
            else:
                notification_message = f"Nueva Nota de Entrega #{new_order.id:09d} creada."
                notification_link = url_for('main.order_detail', order_id=new_order.id)
                create_notification_for_admins(notification_message, notification_link)

            log_user_activity(
                action="Creó orden de venta",
                details=f"Orden tipo '{sale_type}' para cliente '{new_order.client.name}' por un total de ${new_order.total_amount_usd:.2f}. Estado: {new_order.status}",
                target_id=new_order.id,
                target_type="Order"
            )

            db.session.commit()
            if sale_type == 'special_dispatch':
                flash('Entrega Especial creada exitosamente. Esperando aprobación del gerente.', 'info')
                return redirect(url_for('main.order_detail', order_id=new_order.id))
            elif sale_type == 'reservation':
                flash('Apartado creado exitosamente! Preparando para imprimir recibo...', 'success')
                return redirect(url_for('main.print_reservation_receipt', order_id=new_order.id))
            else:
                flash('Orden de venta creada exitosamente! Preparando para imprimir...', 'success')
                return redirect(url_for('main.print_delivery_note', order_id=new_order.id))
        except (ValueError, IntegrityError) as e:
            db.session.rollback()
            flash(f'Error al crear la orden: {str(e)}', 'danger')
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error inesperado al crear la orden: {e}", exc_info=True)
            flash(f'Ocurrió un error inesperado al crear la orden. Por favor, contacta al administrador.', 'danger')
            return redirect(url_for('main.new_order'))

    last_order = Order.query.order_by(Order.date_created.desc()).first()
    current_ve_time = get_current_time_ve()

    return render_template('ordenes/nuevo.html', 
                           title='Nueva Orden de Venta', 
                           clients=clients, 
                           products=products, 
                           banks=banks, 
                           points_of_sale=points_of_sale, 
                           cash_boxes=cash_boxes,
                           last_order=last_order,
                           current_ve_time=current_ve_time)

# --- Rutas de Créditos ---

@routes_blueprint.route('/creditos/lista')
@login_required
def credit_list():
    """Muestra la lista de ventas a crédito."""
    credits = Order.query.filter_by(order_type='credit').order_by(Order.date_created.desc()).all()
    return render_template('creditos/lista.html', title='Historial de Créditos', credits=credits)

@routes_blueprint.route('/creditos/detalle/<int:order_id>', methods=['GET', 'POST'])
@login_required
def credit_detail(order_id):
    """Muestra el detalle de un crédito y permite agregar abonos."""
    order = Order.query.filter_by(id=order_id, order_type='credit').first_or_404()
    if not order:
        flash('Esta orden no es un crédito válido.', 'warning')
        return redirect(url_for('main.credit_list'))

    if request.method == 'POST':
        payment_data_json = request.form.get('payments_data')
        if payment_data_json:
            try:
                payment_info = json.loads(payment_data_json)[0]
                current_rate = get_cached_exchange_rate('USD') or 1.0
                amount_usd_equivalent = float(payment_info['amount_ves_equivalent']) / current_rate if current_rate > 0 else 0

                payment_date = get_current_time_ve()
                if payment_info.get('date'):
                    try:
                        naive_dt = datetime.strptime(payment_info['date'], '%Y-%m-%dT%H:%M')
                        payment_date = VE_TIMEZONE.localize(naive_dt)
                    except (ValueError, TypeError):
                        current_app.logger.warning(f"Invalid payment date format for credit abono: '{payment_info['date']}'. Falling back to now.")

                payment = Payment(
                    order_id=order.id, amount_paid=payment_info['amount_paid'], currency_paid=payment_info['currency_paid'],
                    amount_ves_equivalent=payment_info['amount_ves_equivalent'], amount_usd_equivalent=amount_usd_equivalent, method=payment_info['method'],
                    reference=payment_info.get('reference'), issuing_bank=payment_info.get('issuing_bank'), date=payment_date,
                    sender_id=payment_info.get('sender_id'), bank_id=payment_info.get('bank_id'),
                    pos_id=payment_info.get('pos_id'), cash_box_id=payment_info.get('cash_box_id')
                )
                db.session.add(payment)

                if payment.bank_id:
                    bank = Bank.query.get(payment.bank_id)
                    if bank and bank.currency == 'VES': bank.balance += payment.amount_ves_equivalent
                elif payment.pos_id:
                    pos = PointOfSale.query.get(payment.pos_id)
                    if pos and pos.bank: pos.bank.balance += payment.amount_ves_equivalent
                elif payment.cash_box_id:
                    cash_box = CashBox.query.get(payment.cash_box_id)
                    if cash_box:
                        if payment.currency_paid == 'VES': cash_box.balance_ves += payment.amount_paid
                        elif payment.currency_paid == 'USD': cash_box.balance_usd += payment.amount_paid

                db.session.flush()
                if order.due_amount <= 0.01:
                    order.status = 'Pagada'
                log_user_activity(
                    action="Registró abono a crédito",
                    details=f"Abono de {payment.amount_paid:.2f} {payment.currency_paid} al crédito #{order.id:09d} del cliente '{order.client.name}'",
                    target_id=order.id,
                    target_type="Order"
                )
                db.session.commit()
                flash('Abono al crédito registrado exitosamente.', 'success')
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"Error registrando abono en crédito: {e}")
                flash(f'Error al registrar el abono: {e}', 'danger')
            return redirect(url_for('main.credit_detail', order_id=order.id))

    banks = Bank.query.order_by(Bank.name).all()
    points_of_sale = PointOfSale.query.order_by(PointOfSale.name).all()
    cash_boxes = CashBox.query.order_by(CashBox.name).all()
    return render_template('creditos/detalle.html', title=f'Detalle de Crédito #{order.id:09d}', order=order, banks=banks, points_of_sale=points_of_sale, cash_boxes=cash_boxes)

# --- Rutas de Apartados ---

@routes_blueprint.route('/apartados/lista')
@login_required
def reservation_list():
    """Muestra la lista de productos apartados."""
    # Muestra un historial de todas las órdenes que se crearon como 'apartado',
    # independientemente de su estado actual (Apartado, Pagada, Completada).
    # Esto permite ver tanto los apartados activos como los ya finalizados.
    reservations = Order.query.filter_by(order_type='reservation').order_by(Order.date_created.desc()).all()
    return render_template('apartados/lista.html', title='Historial de Apartados', reservations=reservations)

@routes_blueprint.route('/apartados/detalle/<int:order_id>', methods=['GET', 'POST'])
@login_required
def reservation_detail(order_id):
    """Muestra el detalle de un apartado, permite agregar abonos y marcar como entregado."""
    order = Order.query.filter_by(id=order_id, order_type='reservation').first_or_404()
    if not order:
        flash('Esta orden no es un apartado válido.', 'warning')
        return redirect(url_for('main.reservation_list'))

    if request.method == 'POST':
        action = request.form.get('action')
        
        # Acción para marcar como entregado
        if action == 'deliver':
            if order.status == 'Entregado':
                flash('Este apartado ya fue entregado.', 'info')
            elif order.due_amount <= 0.01:
                order.status = 'Entregado'
                db.session.commit()
                flash('Apartado marcado como entregado.', 'success')
            else:
                flash('El apartado debe estar totalmente pagado para poder ser entregado.', 'warning')
            return redirect(url_for('main.reservation_detail', order_id=order.id))

        # Acción para registrar un abono (pago)
        payment_data_json = request.form.get('payments_data')
        if payment_data_json:
            try:
                payment_info = json.loads(payment_data_json)[0]
                # Para abonos, siempre usar la tasa de cambio actual para calcular el equivalente en USD.
                current_rate = get_cached_exchange_rate('USD') or 1.0
                
                amount_usd_equivalent = float(payment_info['amount_ves_equivalent']) / current_rate if current_rate > 0 else 0

                payment_date = get_current_time_ve()
                if payment_info.get('date'):
                    try:
                        naive_dt = datetime.strptime(payment_info['date'], '%Y-%m-%dT%H:%M')
                        payment_date = VE_TIMEZONE.localize(naive_dt)
                    except (ValueError, TypeError):
                        current_app.logger.warning(f"Invalid payment date format for reservation abono: '{payment_info['date']}'. Falling back to now.")

                payment = Payment(
                    order_id=order.id, amount_paid=payment_info['amount_paid'], currency_paid=payment_info['currency_paid'],
                    amount_ves_equivalent=payment_info['amount_ves_equivalent'], amount_usd_equivalent=amount_usd_equivalent, method=payment_info['method'],
                    reference=payment_info.get('reference'), issuing_bank=payment_info.get('issuing_bank'), date=payment_date,
                    sender_id=payment_info.get('sender_id'), bank_id=payment_info.get('bank_id'),
                    pos_id=payment_info.get('pos_id'), cash_box_id=payment_info.get('cash_box_id')
                )
                db.session.add(payment)

                # Actualizar saldos de cuentas (ESTA LÓGICA FALTABA)
                if payment.bank_id:
                    bank = Bank.query.get(payment.bank_id)
                    # Payments to banks are always registered as their VES equivalent for accounting
                    # but the balance update must respect the bank's currency.
                    if bank and bank.currency == 'VES':
                        bank.balance += payment.amount_ves_equivalent
                elif payment.pos_id:
                    pos = PointOfSale.query.get(payment.pos_id)
                    if pos and pos.bank: pos.bank.balance += payment.amount_ves_equivalent
                elif payment.cash_box_id:
                    cash_box = CashBox.query.get(payment.cash_box_id)
                    if cash_box:
                        if payment.currency_paid == 'VES':
                            cash_box.balance_ves += payment.amount_paid
                        elif payment.currency_paid == 'USD':
                            cash_box.balance_usd += payment.amount_paid

                db.session.flush()
                if order.due_amount <= 0.01 and order.status != 'Entregado':
                    order.status = 'Pagado'
                log_user_activity(
                    action="Registró abono a apartado",
                    details=f"Abono de {payment.amount_paid:.2f} {payment.currency_paid} al apartado #{order.id:09d} del cliente '{order.client.name}'",
                    target_id=order.id,
                    target_type="Order"
                )
                db.session.commit()
                flash('Abono registrado exitosamente.', 'success')
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"Error registrando abono en apartado: {e}")
                flash(f'Error al registrar el abono: {e}', 'danger')
            return redirect(url_for('main.reservation_detail', order_id=order.id))

    banks = Bank.query.order_by(Bank.name).all()
    points_of_sale = PointOfSale.query.order_by(PointOfSale.name).all()
    cash_boxes = CashBox.query.order_by(CashBox.name).all()
    return render_template('apartados/detalle.html', title=f'Detalle de Apartado #{order.id:09d}', order=order, banks=banks, points_of_sale=points_of_sale, cash_boxes=cash_boxes)


@routes_blueprint.route('/actividad_usuarios')
@login_required
def user_activity_log():
    if not is_superuser():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.dashboard'))
    logs = UserActivityLog.query.options(joinedload(UserActivityLog.user)).order_by(UserActivityLog.timestamp.desc()).all()
    return render_template('actividad_usuarios.html', title='Actividad de Usuarios', logs=logs)


# Nueva ruta para movimientos de inventario
@routes_blueprint.route('/movimientos/lista')
@login_required
def movement_list():
    if not is_administrador(): # Superusuario, Gerente, administrador can view movements
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.new_order'))
    product_id = request.args.get('product_id', default=None, type=int) # type: ignore
    start_date_str = request.args.get('start_date', default=None)
    end_date_str = request.args.get('end_date', default=None)

    query = Movement.query

    if product_id:
        query = query.filter(Movement.product_id == product_id)

    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            query = query.filter(Movement.date >= start_date)
        except ValueError:
            flash('Formato de fecha de inicio inválido. Use AAAA-MM-DD.', 'warning')
            start_date_str = None

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            query = query.filter(Movement.date < end_date + timedelta(days=1))
        except ValueError:
            flash('Formato de fecha de fin inválido. Use AAAA-MM-DD.', 'warning')
            end_date_str = None

    movements = query.order_by(Movement.date.desc()).all()
    products = Product.query.order_by(Product.name).all()
    
    return render_template('movimientos/lista.html', 
                           title='Registro de Movimientos', 
                           movements=movements, 
                           products=products,
                           filters={'product_id': product_id, 'start_date': start_date_str, 'end_date': end_date_str})

# Nueva ruta para estadísticas (modo gerencial)
@routes_blueprint.route('/estadisticas')
@login_required
def estadisticas():
    if not is_gerente(): # Superusuario and Gerente can view statistics
        flash('Acceso denegado. Solo los administradores pueden ver esta sección.', 'danger')
        return redirect(url_for('main.new_order'))

    # --- Date Filtering ---
    today = get_current_time_ve().date()
    period = request.args.get('period', 'monthly') # 'monthly', 'daily', 'custom'
    
    if period == 'daily':
        start_date_str = request.args.get('date', today.strftime('%Y-%m-%d'))
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = start_date
        except (ValueError, TypeError):
            start_date = today
            end_date = today
        view_title = f"Estadísticas para {start_date.strftime('%d/%m/%Y')}"

    elif period == 'custom':
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else date(today.year, 1, 1)
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
        except (ValueError, TypeError):
            start_date = date(today.year, 1, 1)
            end_date = today
        view_title = f"Estadísticas de {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"

    else: # Default to 'monthly' for the current year
        start_date = datetime(today.year, 1, 1).date()
        end_date = today
        view_title = f"Estadísticas Mensuales para el Año {today.year}"

    # --- Data Queries ---
    # Excluir el grupo 'Ganchos' (insumos) de las estadísticas
    order_items_query = db.session.query(OrderItem).join(Order).join(Product).filter(
        Order.date_created >= datetime.combine(start_date, datetime.min.time()),
        Order.date_created <= datetime.combine(end_date, datetime.max.time()),
        or_(Product.grupo != 'Ganchos', Product.grupo.is_(None))
    )

    cost_structure = CostStructure.query.first()
    if not cost_structure:
        flash('Por favor, configure la estructura de costos para ver estadísticas precisas.', 'warning')
        cost_structure = CostStructure()

    # --- Calculations (in USD) ---
    stats_data = {}

    def get_period_key(dt):
        if period == 'daily' or (period == 'custom' and (end_date - start_date).days < 32):
            return dt.strftime('%Y-%m-%d')
        return dt.strftime('%Y-%m')

    for item in order_items_query.all():
        period_key = get_period_key(item.order.date_created)

        if period_key not in stats_data:
            stats_data[period_key] = {'sales': 0, 'cogs': 0, 'variable_expenses': 0}

        # All calculations will be in USD.
        rate = item.order.exchange_rate_at_sale
        if not rate or rate <= 0:
            current_app.logger.warning(f"Skipping OrderItem {item.id} in stats due to invalid exchange rate: {rate}")
            continue

        item_revenue_usd = (item.quantity * item.price) / rate

        item_cogs_usd = 0
        if item.cost_at_sale_ves is not None:
            item_cogs_usd = (item.quantity * item.cost_at_sale_ves) / rate
        elif item.product and item.product.cost_usd is not None:
            item_cogs_usd = item.quantity * item.product.cost_usd

        var_sales_exp_pct = item.product.variable_selling_expense_percent if item.product and item.product.variable_selling_expense_percent > 0 else (cost_structure.default_sales_commission_percent or 0)
        var_marketing_pct = item.product.variable_marketing_percent if item.product and item.product.variable_marketing_percent > 0 else (cost_structure.default_marketing_percent or 0)
        item_variable_expense_usd = item_revenue_usd * (var_sales_exp_pct + var_marketing_pct)

        stats_data[period_key]['sales'] += item_revenue_usd
        stats_data[period_key]['cogs'] += item_cogs_usd
        stats_data[period_key]['variable_expenses'] += item_variable_expense_usd

    monthly_fixed_costs_usd = (cost_structure.monthly_rent or 0) + (cost_structure.monthly_utilities or 0) + (cost_structure.monthly_fixed_taxes or 0)
    daily_fixed_costs_usd = monthly_fixed_costs_usd / 30.44

    total_summary = {'sales': 0, 'cogs': 0, 'variable_expenses': 0, 'fixed_expenses': 0, 'gross_profit': 0, 'net_profit': 0}
    sorted_keys = sorted(stats_data.keys())

    for key in sorted_keys:
        data = stats_data[key]
        data['gross_profit'] = data['sales'] - data['cogs']

        is_daily_view = period == 'daily' or (period == 'custom' and (end_date - start_date).days < 32)
        data['fixed_expenses'] = daily_fixed_costs_usd if is_daily_view else monthly_fixed_costs_usd

        data['net_profit'] = data['gross_profit'] - data['variable_expenses'] - data['fixed_expenses']

        for k in total_summary:
            if k in data: total_summary[k] += data[k]

    # --- Chart Data Preparation ---
    chart_labels = []
    chart_sales, chart_cogs, chart_net_profit = [], [], []
    month_names_short = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    for key in sorted_keys:
        # Check if the key is in 'YYYY-MM' format or 'YYYY-MM-DD' format
        if len(key.split('-')) == 2: # It's a monthly key
            chart_labels.append(month_names_short[int(key.split('-')[1]) - 1])
        else: # It's a daily key
            chart_labels.append(datetime.strptime(key, '%Y-%m-%d').strftime('%d/%m'))
        
        chart_sales.append(stats_data[key]['sales'])
        chart_cogs.append(stats_data[key]['cogs'])
        chart_net_profit.append(stats_data[key]['net_profit'])

    profit_loss_chart_data = {'labels': chart_labels, 'sales': chart_sales, 'cogs': chart_cogs, 'net_profit': chart_net_profit}

    # --- Other Stats (Top Products, Clients) ---
    top_products = db.session.query(
        Product.name,
        func.sum(OrderItem.quantity).label('total_sold')
    ).join(OrderItem, OrderItem.product_id == Product.id).join(Order, Order.id == OrderItem.order_id).filter( # Excluir Ganchos
        Order.date_created >= datetime.combine(start_date, datetime.min.time()),
        Order.date_created <= datetime.combine(end_date, datetime.max.time()),
        or_(Product.grupo != 'Ganchos', Product.grupo.is_(None))
    ).group_by(Product.id).order_by(func.sum(OrderItem.quantity).desc()).limit(5).all()

    frequent_clients = db.session.query(
        Client.name,
        func.count(Order.id).label('total_orders')
    ).join(Order, Client.id == Order.client_id).filter(
        Order.date_created >= datetime.combine(start_date, datetime.min.time()),
        Order.date_created <= datetime.combine(end_date, datetime.max.time())
    ).group_by(Client.id).order_by(func.count(Order.id).desc()).limit(5).all()

    top_products_data = {'labels': [p[0] for p in top_products], 'values': [float(p[1] or 0) for p in top_products]}
    frequent_clients_data = {'labels': [c[0] for c in frequent_clients], 'values': [c[1] for c in frequent_clients]}
    
    return render_template('estadisticas.html',
                           title=view_title,
                           stats_data=stats_data,
                           total_summary=total_summary,
                           profit_loss_chart_data=profit_loss_chart_data,
                           top_products_data=top_products_data,
                           frequent_clients_data=frequent_clients_data,
                           filters={'period': period, 'start_date': start_date.strftime('%Y-%m-%d'), 'end_date': end_date.strftime('%Y-%m-%d')},
                           currency_symbol='$')

def generate_pnl_chart_base64(pnl_data, currency_symbol):
    """
    Genera un gráfico de barras con el resumen de resultados (Ventas, Costos, Utilidad)
    y lo devuelve como una imagen codificada en base64.
    """
    labels = ['Ventas', 'Costos Totales', 'Utilidad Neta']
    sales = pnl_data.get('sales', 0)
    # Costos totales = CMV + Gastos (variables + fijos)
    costs = pnl_data.get('cogs', 0) + pnl_data.get('variable_expenses', 0) + pnl_data.get('fixed_expenses', 0)
    net_profit = pnl_data.get('net_profit', 0)
    
    values = [sales, costs, net_profit]
    colors = ['#3B82F6', '#F59E0B', '#22C55E' if net_profit >= 0 else '#EF4444']

    fig, ax = plt.subplots(figsize=(8, 4))
    bars = ax.bar(labels, values, color=colors)

    ax.set_ylabel(f'Monto ({currency_symbol})')
    ax.set_title('Resumen de Resultados del Mes')
    ax.yaxis.grid(True, linestyle='--', which='major', color='grey', alpha=.25)
    
    # Añadir etiquetas de valor sobre las barras
    for bar in bars:
        yval = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2.0, yval, f'{yval:,.2f}', va='bottom' if yval >= 0 else 'top', ha='center')

    plt.tight_layout()

    # Guardar el gráfico en un buffer en memoria
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=100)
    plt.close(fig)
    buf.seek(0)
    # Codificar la imagen en base64 para incrustarla en el HTML
    image_base64 = base64.b64encode(buf.read()).decode('utf-8')
    return image_base64


@routes_blueprint.route('/reporte-mensual-pdf')
@login_required
def generar_reporte_mensual_pdf():
    """
    Recopila toda la información financiera de un mes específico y genera un reporte en PDF.
    """

    try:
        month = int(request.args.get('month'))
        year = int(request.args.get('year'))
    except (ValueError, TypeError):
        return "Error: Mes y año inválidos.", 400

    # --- 1. Definir Período y Variables ---
    _, num_days = calendar.monthrange(year, month)
    start_date = date(year, month, 1)
    end_date = date(year, month, num_days)
    start_dt = VE_TIMEZONE.localize(datetime.combine(start_date, datetime.min.time()))
    end_dt = VE_TIMEZONE.localize(datetime.combine(end_date, datetime.max.time()))
    
    month_name = get_month_names('wide', locale='es_ES')[month]
    report_period = f"{month_name.capitalize()} {year}"
    currency_symbol = "$" # Las estadísticas se manejan en USD

    # --- 2. Recopilación de Datos (en USD) - Lógica unificada con el reporte diario ---
    current_fallback_rate = get_cached_exchange_rate('USD') or 1.0
    if current_fallback_rate <= 0: current_fallback_rate = 1.0 # Evitar división por cero

    # A. Estado de Resultados (P&L)
    pnl_summary = {'sales': 0, 'cogs': 0, 'variable_expenses': 0, 'fixed_expenses': 0, 'gross_profit': 0, 'net_profit': 0}
    
    orders_in_month = Order.query.filter(
        Order.date_created.between(start_dt, end_dt),
    ).options(joinedload(Order.items).joinedload(OrderItem.product)).all()

    cost_structure = CostStructure.query.first() or CostStructure()

    for order in orders_in_month:
        pnl_summary['sales'] += order.total_amount_usd
        
        rate = order.exchange_rate_at_sale or current_fallback_rate
        if rate > 0:
            for item in order.items:
                # Calcular CMV
                if item.cost_at_sale_ves is not None:
                    pnl_summary['cogs'] += (item.cost_at_sale_ves * item.quantity) / rate
                
                # Calcular Gastos Variables
                item_revenue_usd = (item.price * item.quantity) / rate
                var_sales_exp_pct = item.product.variable_selling_expense_percent if item.product and item.product.variable_selling_expense_percent > 0 else (cost_structure.default_sales_commission_percent or 0)
                var_marketing_pct = item.product.variable_marketing_percent if item.product and item.product.variable_marketing_percent > 0 else (cost_structure.default_marketing_percent or 0)
                pnl_summary['variable_expenses'] += item_revenue_usd * (var_sales_exp_pct + var_marketing_pct)

    pnl_summary['fixed_expenses'] = (cost_structure.monthly_rent or 0) + (cost_structure.monthly_utilities or 0) + (cost_structure.monthly_fixed_taxes or 0)
    pnl_summary['gross_profit'] = pnl_summary['sales'] - pnl_summary['cogs']
    pnl_summary['net_profit'] = pnl_summary['gross_profit'] - pnl_summary['variable_expenses'] - pnl_summary['fixed_expenses']

    # B. Productos más vendidos
    top_products = db.session.query(
        Product.name.label('nombre'), func.sum(OrderItem.quantity).label('total_vendido')
    ).join(OrderItem).join(Order).filter(
        Order.date_created.between(start_dt, end_dt)
    ).group_by(Product.id).order_by(func.sum(OrderItem.quantity).desc()).limit(10).all()

    # C. Ventas por tipo (status)
    sales_by_type_raw = db.session.query(
        Order.order_type,
        func.count(Order.id).label('num_ventas'),
        func.sum(Order.total_amount_usd).label('total_ventas_usd')
    ).filter(
        Order.date_created.between(start_dt, end_dt)
    ).group_by(Order.order_type).all()

    sales_by_type = {'Contado': {'num_ventas': 0, 'total_ventas': 0.0}, 'Crédito': {'num_ventas': 0, 'total_ventas': 0.0}, 'Apartado': {'num_ventas': 0, 'total_ventas': 0.0}}
    for order_type, num, total in sales_by_type_raw:
        total = float(total or 0.0)
        if order_type == 'regular': sales_by_type['Contado']['num_ventas'] += num; sales_by_type['Contado']['total_ventas'] += total
        elif order_type == 'credit': sales_by_type['Crédito']['num_ventas'] += num; sales_by_type['Crédito']['total_ventas'] += total
        elif order_type == 'reservation': sales_by_type['Apartado']['num_ventas'] += num; sales_by_type['Apartado']['total_ventas'] += total

    # D. Cuentas por cobrar pendientes (al final del mes)
    # Optimización: Filtrar en la base de datos en lugar de en Python
    paid_subquery = db.session.query(
        Payment.order_id,
        func.sum(Payment.amount_ves_equivalent).label('total_paid')
    ).group_by(Payment.order_id).subquery()

    pending_accounts_receivable = Order.query.options(
        joinedload(Order.client), subqueryload(Order.payments)
    ).outerjoin(paid_subquery, Order.id == paid_subquery.c.order_id).filter( # This logic is correct, it calculates due amount at the time of query
        Order.date_created <= end_dt, (Order.total_amount - func.coalesce(paid_subquery.c.total_paid, 0)) > 0.01
    ).order_by(Order.date_created.asc()).all()

    # E. Cobros hechos en el mes
    collections_in_month = Payment.query.options(joinedload(Payment.order).joinedload(Order.client)).filter(Payment.date.between(start_dt, end_dt)).order_by(Payment.date.asc()).all()

    # F. Flujo de Fondos por Cuenta
    from flask import make_response
    banks = Bank.query.all()
    bank_balances = []
    for bank in banks:
        inflows_ves = (db.session.query(func.sum(Payment.amount_ves_equivalent)).filter(or_(Payment.bank_id == bank.id, Payment.pos.has(bank_id=bank.id)), Payment.date.between(start_dt, end_dt)).scalar() or 0.0) + (db.session.query(func.sum(ManualFinancialMovement.amount)).filter(ManualFinancialMovement.bank_id == bank.id, ManualFinancialMovement.date.between(start_dt, end_dt), ManualFinancialMovement.movement_type == 'Ingreso', ManualFinancialMovement.currency == 'VES', ManualFinancialMovement.status == 'Aprobado').scalar() or 0.0)
        outflows_ves = db.session.query(func.sum(ManualFinancialMovement.amount)).filter(ManualFinancialMovement.bank_id == bank.id, ManualFinancialMovement.date.between(start_dt, end_dt), ManualFinancialMovement.movement_type == 'Egreso', ManualFinancialMovement.status == 'Aprobado', ManualFinancialMovement.currency == 'VES').scalar() or 0.0
        final_balance_ves = bank.balance # This is the current balance
        initial_balance_ves = final_balance_ves - inflows_ves + outflows_ves
        bank_balances.append({'name': bank.name, 'initial_balance_ves': initial_balance_ves, 'inflows_ves': inflows_ves, 'outflows_ves': outflows_ves, 'final_balance_ves': final_balance_ves})

    cash_boxes = CashBox.query.all()
    cash_box_balances = []
    for box in cash_boxes:
        # VES
        inflows_ves = (db.session.query(func.sum(Payment.amount_paid)).filter(Payment.cash_box_id == box.id, Payment.date.between(start_dt, end_dt), Payment.currency_paid == 'VES').scalar() or 0.0) + (db.session.query(func.sum(ManualFinancialMovement.amount)).filter(ManualFinancialMovement.cash_box_id == box.id, ManualFinancialMovement.date.between(start_dt, end_dt), ManualFinancialMovement.movement_type == 'Ingreso', ManualFinancialMovement.currency == 'VES', ManualFinancialMovement.status == 'Aprobado').scalar() or 0.0)
        outflows_ves = db.session.query(func.sum(ManualFinancialMovement.amount)).filter(ManualFinancialMovement.cash_box_id == box.id, ManualFinancialMovement.date.between(start_dt, end_dt), ManualFinancialMovement.movement_type == 'Egreso', ManualFinancialMovement.status == 'Aprobado', ManualFinancialMovement.currency == 'VES').scalar() or 0.0
        final_balance_ves = box.balance_ves
        initial_balance_ves = final_balance_ves - inflows_ves + outflows_ves

        # USD
        inflows_usd = (db.session.query(func.sum(Payment.amount_paid)).filter(Payment.cash_box_id == box.id, Payment.date.between(start_dt, end_dt), Payment.currency_paid == 'USD').scalar() or 0.0) + (db.session.query(func.sum(ManualFinancialMovement.amount)).filter(ManualFinancialMovement.cash_box_id == box.id, ManualFinancialMovement.date.between(start_dt, end_dt), ManualFinancialMovement.movement_type == 'Ingreso', ManualFinancialMovement.currency == 'USD', ManualFinancialMovement.status == 'Aprobado').scalar() or 0.0)
        outflows_usd = db.session.query(func.sum(ManualFinancialMovement.amount)).filter(ManualFinancialMovement.cash_box_id == box.id, ManualFinancialMovement.date.between(start_dt, end_dt), ManualFinancialMovement.movement_type == 'Egreso', ManualFinancialMovement.status == 'Aprobado', ManualFinancialMovement.currency == 'USD').scalar() or 0.0
        final_balance_usd = box.balance_usd
        initial_balance_usd = final_balance_usd - inflows_usd + outflows_usd

        cash_box_balances.append({'name': box.name, 'initial_balance_ves': initial_balance_ves, 'inflows_ves': inflows_ves, 'outflows_ves': outflows_ves, 'final_balance_ves': final_balance_ves, 'initial_balance_usd': initial_balance_usd, 'inflows_usd': inflows_usd, 'outflows_usd': outflows_usd, 'final_balance_usd': final_balance_usd})

    # --- 3. Generación de Gráfico ---
    pnl_chart_base64 = generate_pnl_chart_base64(pnl_summary, currency_symbol)

    # --- 4. Verificación de Datos y Renderizado del Template ---
    is_data_available = (
        pnl_summary['sales'] > 0 or
        pnl_summary['cogs'] > 0 or
        pnl_summary['variable_expenses'] > 0 or
        pnl_summary['fixed_expenses'] > 0 or
        pending_accounts_receivable or
        collections_in_month or
        any(b['inflows_ves'] > 0 or b['outflows_ves'] > 0 for b in bank_balances) or
        any(c['inflows_ves'] > 0 or c['outflows_ves'] > 0 or c['inflows_usd'] > 0 or c['outflows_usd'] > 0 for c in cash_box_balances)
    )

    generation_date_str = get_current_time_ve().strftime("%d/%m/%Y %H:%M:%S")

    if not is_data_available:
        # Renderiza una plantilla simple de "sin datos"
        html_string = render_template('pdf/reporte_mensual_sin_datos.html', report_period=report_period, generation_date=generation_date_str)
    else:
        # Renderiza el reporte completo
        context = {
            'report_period': report_period,
            'generation_date': generation_date_str,
            'currency_symbol': currency_symbol,
            'pnl_summary': pnl_summary,
            'pnl_chart_base64': pnl_chart_base64,
            'top_products': top_products,
            'sales_by_type': sales_by_type,
            'pending_accounts_receivable': pending_accounts_receivable,
            'collections_in_month': collections_in_month,
            'bank_balances': bank_balances,
            'cash_box_balances': cash_box_balances,
            'start_date': start_date,
            'end_date': end_date,
            'current_fallback_rate': current_fallback_rate,
        }
        html_string = render_template('pdf/reporte_mensual_pdf.html', **context)

    # --- 5. Creación del PDF y Envío de Respuesta ---
    pdf_file = HTML(string=html_string, base_url=request.base_url).write_pdf()

    response = make_response(pdf_file)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=cierre_mensual_{year}_{month:02d}.pdf'
    
    return response

# Nueva ruta para cargar productos desde un archivo de Excel
@routes_blueprint.route('/inventario/cargar_excel', methods=['GET', 'POST'])
@login_required
def cargar_excel():
    if not is_gerente(): # Superusuario and Gerente can upload excel
        flash('Acceso denegado. Solo los administradores pueden realizar esta acción.', 'danger')
        return redirect(url_for('main.inventory_list'))

    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No se ha seleccionado ningún archivo.', 'danger')
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash('No se ha seleccionado ningún archivo.', 'danger')
            return redirect(request.url)
        
        warehouse_id = request.form.get('warehouse_id', type=int)
        if not warehouse_id:
            flash('Debe seleccionar un almacén de destino.', 'danger')
            return redirect(request.url)

        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Formato de archivo no válido. Solo se aceptan archivos .xlsx.', 'danger')
            return redirect(request.url)

        # Use a temporary directory within the instance path for cross-platform compatibility
        upload_dir = os.path.join(current_app.instance_path, 'uploads')
        os.makedirs(upload_dir, exist_ok=True)
        
        filepath = os.path.join(upload_dir, file.filename)
        file.save(filepath)

        try:
            workbook = openpyxl.load_workbook(filepath, data_only=True)
            sheet = workbook.active
            
            new_products = []
            updates = []
            all_barcodes_in_file = {str(row[0]).strip() for row in sheet.iter_rows(min_row=1, values_only=True) if row and row[0]}
            
            for row in sheet.iter_rows(min_row=1, values_only=True):
                if not row[0]:
                    continue
                
                barcode = str(row[0]).strip()
                codigo_producto = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
                name = str(row[2]).strip()
                cost_usd = row[3] if row[3] is not None else 0
                price_usd = row[4] if row[4] is not None else 0
                stock_to_add = int(row[5]) if row[5] is not None else 0
                image_url = row[6] if row[6] is not None else ''
                marca = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ''
                color = str(row[8]).strip() if len(row) > 8 and row[8] is not None else ''
                talla = str(row[9]).strip() if len(row) > 9 and row[9] is not None else ''
                grupo = str(row[10]).strip() if len(row) > 10 and row[10] is not None else ''

                product = Product.query.filter_by(barcode=barcode).first()
                current_stock_in_warehouse = 0

                if product:
                    stock_entry = ProductStock.query.filter_by(product_id=product.id, warehouse_id=warehouse_id).first()
                    if stock_entry:
                        current_stock_in_warehouse = stock_entry.quantity

                    updates.append({
                        'id': product.id,
                        'name': product.name,
                        'barcode': barcode,
                        'stock_to_add': stock_to_add,
                        'old_stock': current_stock_in_warehouse,
                        'new_total_stock': current_stock_in_warehouse + stock_to_add,
                    })
                else:
                    new_products.append({
                        'barcode': barcode,
                        'codigo_producto': codigo_producto,
                        'name': name,
                        'cost_usd': float(cost_usd),
                        'price_usd': float(price_usd),
                        'stock_to_add': stock_to_add,
                        'image_url': image_url,
                        'marca': marca,
                        'color': color,
                        'size': talla,
                        'grupo': grupo,
                    })

            if updates:
                # Guardar en sesión para la página de confirmación
                session['excel_upload_data'] = {
                    'warehouse_id': warehouse_id,
                    'new_products': new_products,
                    'updates': updates
                }
                return redirect(url_for('main.cargar_excel_confirmar'))
            
            # Si solo hay productos nuevos, los procesamos directamente
            if new_products:
                # Crear el log de carga
                load_log = BulkLoadLog(user_id=current_user.id, method='Excel', warehouse_id=warehouse_id)
                db.session.add(load_log)
                db.session.flush()

                for prod_data in new_products:
                    stock_to_add = prod_data.pop('stock_to_add')
                    new_prod = Product(**prod_data)
                    db.session.add(new_prod)
                    db.session.flush() # Para obtener el ID del nuevo producto
                    
                    # Añadir stock y movimiento
                    add_stock_and_movement(new_prod.id, warehouse_id, stock_to_add, load_log.id, f"Carga Masiva #{load_log.id}")

            db.session.commit()

            log_user_activity(
                action="Realizó carga masiva de productos",
                details=f"Carga desde Excel al almacén ID {warehouse_id}. {len(new_products)} productos nuevos.",
                target_id=load_log.id,
                target_type="BulkLoadLog"
            )
            flash(f'Se han agregado {len(new_products)} productos nuevos al inventario.', 'success')
            return redirect(url_for('main.inventory_list'))

        except Exception as e:
            db.session.rollback()
            flash(f'Ocurrió un error al procesar el archivo: {str(e)}', 'danger')
            return redirect(request.url)
        finally:
            if os.path.exists(filepath):
                os.remove(filepath)
    
    warehouses = Warehouse.query.order_by(Warehouse.id).all()
    # Cargar el historial de cargas masivas para mostrarlo en la página
    load_history = BulkLoadLog.query.options(
        joinedload(BulkLoadLog.user),
        joinedload(BulkLoadLog.warehouse)
    ).order_by(BulkLoadLog.date.desc()).all()

    return render_template('inventario/cargar_excel.html', title='Cargar Inventario desde Excel', 
                           warehouses=warehouses, load_history=load_history)

def add_stock_and_movement(product_id, warehouse_id, quantity, document_id, document_type):
    """Función auxiliar para añadir stock y registrar el movimiento."""
    if quantity <= 0:
        return

    # Actualizar o crear el registro de stock
    stock_entry = ProductStock.query.filter_by(product_id=product_id, warehouse_id=warehouse_id).first()
    if stock_entry:
        stock_entry.quantity += quantity
    else:
        stock_entry = ProductStock(product_id=product_id, warehouse_id=warehouse_id, quantity=quantity)
        db.session.add(stock_entry)

    # Crear el movimiento de inventario
    movement = Movement(
        product_id=product_id,
        type='Entrada',
        warehouse_id=warehouse_id,
        quantity=quantity,
        document_id=document_id,
        document_type=document_type,
        description="Cargado mediante Excel"
    )
    db.session.add(movement)

@routes_blueprint.route('/inventario/cargar_excel_confirmar', methods=['GET', 'POST'])
@login_required
def cargar_excel_confirmar():
    if not is_gerente(): # Superusuario and Gerente can confirm excel upload
        flash('Acceso denegado. Solo los administradores pueden realizar esta acción.', 'danger')
        return redirect(url_for('main.inventory_list'))
        
    upload_data = session.get('excel_upload_data', {})
    if not upload_data:
        flash('No hay datos de carga para confirmar.', 'warning')
        return redirect(url_for('main.cargar_excel'))

    if request.method == 'POST':
        try:
            warehouse_id = upload_data['warehouse_id']
            
            # Crear el log de carga
            load_log = BulkLoadLog(user_id=current_user.id, method='Excel', warehouse_id=warehouse_id)
            db.session.add(load_log)
            db.session.flush()

            # Procesar productos nuevos
            for prod_data in upload_data.get('new_products', []):
                stock_to_add = prod_data.pop('stock_to_add')
                new_prod = Product(**prod_data)
                db.session.add(new_prod)
                db.session.flush()
                add_stock_and_movement(new_prod.id, warehouse_id, stock_to_add, load_log.id, f"Carga Masiva #{load_log.id}")

            # Procesar actualizaciones de stock para productos existentes
            for update_data in upload_data.get('updates', []):
                add_stock_and_movement(update_data['id'], warehouse_id, update_data['stock_to_add'], load_log.id, "Carga Masiva Excel")

            db.session.commit()
            log_user_activity(
                action="Confirmó carga masiva de productos",
                details=f"Carga desde Excel al almacén ID {warehouse_id}. {len(upload_data.get('new_products', []))} nuevos, {len(upload_data.get('updates', []))} actualizados.",
                target_id=load_log.id,
                target_type="BulkLoadLog"
            )

            flash(f'Carga de inventario procesada exitosamente. Se crearon {len(upload_data.get("new_products", []))} productos y se actualizó el stock de {len(upload_data.get("updates", []))} productos.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Ocurrió un error al confirmar la actualización: {str(e)}', 'danger')
        finally:
            session.pop('excel_upload_data', None)
        
        return redirect(url_for('main.inventory_list'))

    return render_template('inventario/cargar_excel_confirmar.html', 
                           title='Confirmar Actualización de Inventario',
                           updates=upload_data.get('updates', []),
                           new_products=upload_data.get('new_products', []))

@routes_blueprint.route('/inventario/carga_masiva/detalle/<int:log_id>')
@login_required
def bulk_load_detail(log_id):
    """Muestra los detalles de una carga masiva específica."""
    if not is_gerente():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.inventory_list'))

    log_entry = BulkLoadLog.query.options(
        joinedload(BulkLoadLog.user),
        joinedload(BulkLoadLog.warehouse)
    ).get_or_404(log_id)

    # Usamos la relación 'movements' que añadimos al modelo para obtener los productos afectados
    movements = log_entry.movements.options(
        joinedload(Movement.product)
    ).all()

    # Preparar una lista de IDs de productos para el botón de imprimir
    product_ids_to_print = [str(m.product.id) for m in movements]

    return render_template('inventario/detalle_carga_masiva.html',
                           title=f'Detalle de Carga #{log_entry.id}',
                           log_entry=log_entry, movements=movements,
                           product_ids_to_print=product_ids_to_print)

# Rutas de configuración de empresa
@routes_blueprint.route('/configuracion/empresa', methods=['GET', 'POST'])
@login_required
def company_settings():
    if not is_superuser(): # Only Superuser can access company settings
        flash('Acceso denegado. Solo el administrador del sistema puede realizar esta acción.', 'danger')
        return redirect(url_for('main.dashboard'))

    company_info = CompanyInfo.query.first()
    form = CompanyInfoForm(obj=company_info)

    if form.validate_on_submit():
        try:
            # Si no existe información de la empresa, se crea una nueva.
            if company_info:
                # Actualizar la empresa existente
                form.populate_obj(company_info)
            else:
                # Crear una nueva si no hay ninguna
                company_info = CompanyInfo()
                form.populate_obj(company_info)
                db.session.add(company_info)
                db.session.flush() # Para obtener el ID para el nombre del logo

            # Manejar la subida del logo
            logo_file = form.logo_file.data
            if logo_file:
                pass
                upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'logos')
                os.makedirs(upload_dir, exist_ok=True)
                
                filename = f"logo_{company_info.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{os.path.splitext(logo_file.filename)[1]}"
                filepath = os.path.join(upload_dir, filename)
                
                logo_file.save(filepath)
                company_info.logo_filename = f"uploads/logos/{filename}"
            
            db.session.commit()
            flash('Información de la empresa guardada exitosamente.', 'success')
            return redirect(url_for('main.company_settings'))
        except Exception as e:
            db.session.rollback()
            flash(f'Ocurrió un error al guardar la información: {str(e)}', 'danger')

    return render_template('configuracion/empresa.html', title='Configuración de Empresa', form=form, company_info=company_info)

# Rutas de Estructura de Costos
@routes_blueprint.route('/costos/lista')
@login_required
def cost_list():
    if not is_gerente(): # Superusuario and Gerente can view cost list
        flash('Acceso denegado. Solo los administradores pueden ver esta sección.', 'danger')
        return redirect(url_for('main.dashboard'))

    cost_structure = CostStructure.query.first()
    if not cost_structure:
        flash('Por favor, configure la estructura de costos generales primero.', 'info')
        return redirect(url_for('main.cost_structure_config'))

    # Excluir el grupo 'Ganchos' (insumos) de la estructura de costos
    products = Product.query.filter(or_(Product.grupo != 'Ganchos', Product.grupo.is_(None))).all()
    
    total_estimated_sales = db.session.query(func.sum(Product.estimated_monthly_sales)).filter(or_(Product.grupo != 'Ganchos', Product.grupo.is_(None))).scalar() or 1
    if total_estimated_sales == 0:
        total_estimated_sales = 1

    total_fixed_costs = (cost_structure.monthly_rent or 0) + \
                        (cost_structure.monthly_utilities or 0) + \
                        (cost_structure.monthly_fixed_taxes or 0)
    
    fixed_cost_per_unit = total_fixed_costs / total_estimated_sales

    products_with_costs = []
    for product in products:
        # El precio de venta final es el que está guardado en el producto.
        selling_price = product.price_usd or 0

        # Usar gastos variables específicos o los por defecto.
        var_sales_exp_pct = product.variable_selling_expense_percent if product.variable_selling_expense_percent > 0 else cost_structure.default_sales_commission_percent
        var_marketing_pct = product.variable_marketing_percent if product.variable_marketing_percent > 0 else cost_structure.default_marketing_percent

        # Calcular el costo total por unidad basado en el precio de venta final.
        total_cost_per_unit = (product.cost_usd or 0) + \
                              (product.specific_freight_cost or 0) + \
                              fixed_cost_per_unit + \
                              (selling_price * (var_sales_exp_pct or 0)) + \
                              (selling_price * (var_marketing_pct or 0))
        
        # La utilidad es la diferencia entre el precio de venta y el costo total.
        profit = selling_price - total_cost_per_unit

        error = "El producto genera pérdidas." if profit < 0 and selling_price > 0 else None

        products_with_costs.append({
            'product': product,
            'profit': profit,
            'selling_price': selling_price,
            'error': error
        })

    return render_template('costos/lista.html',
                           title='Estructura de Costos',
                           products_data=products_with_costs)


@routes_blueprint.route('/costos/configuracion', methods=['GET', 'POST'])
@login_required
def cost_structure_config():
    if not is_superuser(): # Only Superuser can configure cost structure
        flash('Acceso denegado. Solo los administradores pueden realizar esta acción.', 'danger')
        return redirect(url_for('main.dashboard'))

    cost_structure = CostStructure.query.first()
    if not cost_structure:
        cost_structure = CostStructure()
        db.session.add(cost_structure)
        db.session.commit()

    if request.method == 'POST':
        try:
            cost_structure.monthly_rent = float(request.form.get('monthly_rent', 0))
            cost_structure.monthly_utilities = float(request.form.get('monthly_utilities', 0))
            cost_structure.monthly_fixed_taxes = float(request.form.get('monthly_fixed_taxes', 0))
            cost_structure.default_sales_commission_percent = float(request.form.get('default_sales_commission_percent', 0)) / 100
            cost_structure.default_marketing_percent = float(request.form.get('default_marketing_percent', 0)) / 100
            
            db.session.commit()
            flash('Configuración de costos guardada exitosamente.', 'success')
            return redirect(url_for('main.cost_list'))
        except (ValueError, TypeError) as e:
            db.session.rollback()
            flash(f'Error al guardar la configuración. Verifique que los valores sean números. Error: {e}', 'danger')

    usd_rate = get_cached_exchange_rate('USD')
    eur_rate = get_cached_exchange_rate('EUR')
    manual_rate_required = usd_rate is None or eur_rate is None
    if manual_rate_required:
        flash('No se pudo obtener la tasa de cambio de las APIs. Por favor, ingrese un valor manualmente.', 'warning')

    return render_template('costos/configuracion.html',
                           title='Configurar Costos Generales',
                           cost_structure=cost_structure,
                           usd_rate=usd_rate or 0.0,
                           eur_rate=eur_rate or 0.0,
                           manual_rate_required=manual_rate_required,
                           exchange_rate_info_usd=ExchangeRate.query.filter_by(currency='USD').first(),
                           exchange_rate_info_eur=ExchangeRate.query.filter_by(currency='EUR').first())


@routes_blueprint.route('/costos/update_rate', methods=['POST'])
@login_required
def update_exchange_rate():
    """
    Updates an exchange rate. Can be called from the settings page (form redirect)
    or from the new order modal (AJAX).
    """
    if not is_superuser(): # Only Superuser can update exchange rate
        # CORRECCIÓN: La llamada fetch desde el modal no es JSON, pero es AJAX.
        # Usamos 'X-Requested-With' o un campo del formulario para detectar la llamada AJAX.
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.form.get('is_ajax')
        if is_ajax:
            return jsonify(success=False, message='Acceso denegado.'), 403
        return redirect(url_for('main.new_order'))

    try:
        currency = request.form.get('currency')
        manual_rate = float(request.form.get('manual_rate'))
        store_original = request.form.get('store_original_rate') == 'true'

        if manual_rate > 0 and currency in ['USD', 'EUR']:
            exchange_rate_entry = ExchangeRate.query.filter_by(currency=currency).first()
            if exchange_rate_entry:
                if store_original:
                    session['original_exchange_rate'] = exchange_rate_entry.rate
                    session['original_rate_currency'] = currency
                exchange_rate_entry.rate = manual_rate
                exchange_rate_entry.date_updated = get_current_time_ve()
            else:
                exchange_rate_entry = ExchangeRate(currency=currency, rate=manual_rate)
                db.session.add(exchange_rate_entry)
            db.session.commit()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.form.get('is_ajax') == 'true':
                return jsonify(success=True, message='Tasa de cambio actualizada.')
            else:
                flash('Tasa de cambio actualizada manualmente.', 'success')
        else:
            raise ValueError('La tasa de cambio debe ser un número positivo.')
    except (ValueError, TypeError) as err:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.form.get('is_ajax') == 'true':
            return jsonify(success=False, message=str(err)), 400
        
        flash(f'Valor de tasa de cambio inválido: {err}', 'danger')

    # Default redirect for non-AJAX calls (e.g., from config page)
    return redirect(url_for('main.cost_structure_config'))


@routes_blueprint.route('/costos/editar/<int:product_id>', methods=['GET', 'POST'])
@login_required
def edit_product_cost(product_id):
    if not is_gerente(): # Superusuario and Gerente can edit product costs
        flash('Acceso denegado. Solo los administradores pueden realizar esta acción.', 'danger')
        return redirect(url_for('main.dashboard'))

    product = Product.query.get_or_404(product_id)
    cost_structure = CostStructure.query.first()
    
    # Calcular punto de equilibrio financiero
    break_even_data = None
    if cost_structure:
        # Calcular costos fijos totales
        total_fixed_costs = (cost_structure.monthly_rent or 0) + \
                           (cost_structure.monthly_utilities or 0) + \
                           (cost_structure.monthly_fixed_taxes or 0)
        
        # Calcular ventas mensuales estimadas totales
        total_estimated_sales = db.session.query(func.sum(Product.estimated_monthly_sales)).scalar() or 1
        if total_estimated_sales == 0:
            total_estimated_sales = 1
        
        # Calcular costos fijos por unidad
        fixed_cost_per_unit = total_fixed_costs / total_estimated_sales
        
        # Usar gastos variables específicos o los valores por defecto (asegurando que no sean None)
        var_sales_exp_pct = product.variable_selling_expense_percent if product.variable_selling_expense_percent > 0 else (cost_structure.default_sales_commission_percent or 0)
        var_marketing_pct = product.variable_marketing_percent if product.variable_marketing_percent > 0 else (cost_structure.default_marketing_percent or 0)
        
        # El precio de venta se toma directamente del producto
        selling_price = product.price_usd or 0
        base_cost = (product.cost_usd or 0) + (product.specific_freight_cost or 0) + fixed_cost_per_unit

        # Recalcular el margen de utilidad para mostrar el valor actual real
        if selling_price > 0:
            profit_margin_calc = 1 - var_sales_exp_pct - var_marketing_pct - (base_cost / selling_price)
            product.profit_margin = profit_margin_calc

            # Calcular costo variable unitario
            variable_cost_per_unit = (product.cost_usd or 0) + (product.specific_freight_cost or 0) + \
                                   (selling_price * var_sales_exp_pct) + \
                                   (selling_price * var_marketing_pct)

            # Calcular punto de equilibrio
            if selling_price > variable_cost_per_unit:
                break_even_units = total_fixed_costs / (selling_price - variable_cost_per_unit)
                break_even_amount = break_even_units * selling_price

                break_even_data = {
                    'fixed_costs_total': total_fixed_costs,
                    'fixed_cost_per_unit': fixed_cost_per_unit,
                    'variable_cost_per_unit': variable_cost_per_unit,
                    'selling_price': selling_price,
                    'break_even_units': break_even_units,
                    'break_even_amount': break_even_amount,
                    'var_sales_exp_pct': var_sales_exp_pct * 100,
                    'var_marketing_pct': var_marketing_pct * 100
                }

    if request.method == 'POST':
        try:
            # Actualizar campos del producto desde el formulario
            product.price_usd = float(request.form.get('price_usd', 0))
            product.specific_freight_cost = float(request.form.get('specific_freight_cost', 0))
            product.estimated_monthly_sales = int(request.form.get('estimated_monthly_sales', 1))
            product.variable_selling_expense_percent = float(request.form.get('variable_selling_expense_percent', 0)) / 100
            product.variable_marketing_percent = float(request.form.get('variable_marketing_percent', 0)) / 100

            if not cost_structure:
                flash('La configuración de costos generales no existe. No se puede calcular la utilidad.', 'danger')
                return redirect(url_for('main.cost_structure_config'))

            # Recalcular componentes de costo con los nuevos datos
            total_estimated_sales = db.session.query(func.sum(Product.estimated_monthly_sales)).scalar() or 1
            if total_estimated_sales == 0: total_estimated_sales = 1

            total_fixed_costs = (cost_structure.monthly_rent or 0) + (cost_structure.monthly_utilities or 0) + (cost_structure.monthly_fixed_taxes or 0)
            fixed_cost_per_unit = total_fixed_costs / total_estimated_sales
            base_cost = (product.cost_usd or 0) + product.specific_freight_cost + fixed_cost_per_unit
            
            # Recalcular y guardar el nuevo margen de utilidad
            if product.price_usd > 0:
                new_profit_margin = 1 - product.variable_selling_expense_percent - product.variable_marketing_percent - (base_cost / product.price_usd)
                product.profit_margin = new_profit_margin
                # Alertar al usuario sobre utilidad baja o negativa
                if new_profit_margin < 0:
                    flash(f'¡Atención! Con los costos y precio de venta actuales, se está generando una pérdida. Margen de utilidad: {new_profit_margin*100:.2f}%.', 'danger')
                elif new_profit_margin < 0.05: # Umbral de advertencia del 5%
                    flash(f'Advertencia: El margen de utilidad es muy bajo: {new_profit_margin*100:.2f}%.', 'warning')
            else:
                product.profit_margin = 0
                flash('El precio de venta debe ser un número positivo.', 'danger')

            log_user_activity(
                action="Actualizó costos de producto",
                details=f"Producto: {product.name}. Nuevo precio: ${product.price_usd:.2f}, Margen: {product.profit_margin*100:.2f}%",
                target_id=product.id,
                target_type="Product"
            )

            db.session.commit()
            flash(f'Costos y precio del producto "{product.name}" actualizados exitosamente.', 'success')
            return redirect(url_for('main.cost_list'))
        except ValueError as e:
            db.session.rollback()
            flash(f'Error al actualizar el producto: {e}', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Ocurrió un error inesperado: {e}', 'danger')

    return render_template('costos/editar.html',
                           title=f'Editar Costos de {product.name}',
                           product=product,
                           break_even_data=break_even_data)

@routes_blueprint.route('/ordenes/imprimir/<int:order_id>')
@login_required
def print_delivery_note(order_id):
    order = Order.query.get_or_404(order_id)
    company_info = CompanyInfo.query.first()

    # El subtotal y el IVA se calculan directamente en la plantilla para manejar devoluciones.
    order_total_with_iva = order.total_amount # This is the final amount after discount

    # Calculate total paid and change
    total_paid = sum(p.amount_ves_equivalent for p in order.payments)
    change = total_paid - order_total_with_iva if total_paid > order_total_with_iva else 0.0

    # Helper function to generate barcode
    def generate_order_barcode_base64(order_id_str):
        """Generates a Code128 barcode image and returns it as a base64 string."""
        if not order_id_str:
            return None
        try:
            barcode = createBarcodeDrawing('Code128', value=order_id_str, barHeight=10*mm, barWidth=0.3*mm)
            drawing = Drawing(barcode.width, barcode.height)
            drawing.add(barcode)
            buffer = io.BytesIO()
            try:
                # First, try the C-based backend which is faster
                from reportlab.graphics import renderPM
                renderPM.drawToFile(drawing, buffer, fmt='PNG')
            except ImportError:
                # Fallback to the pure Python backend if the C-backend is not available
                from reportlab.graphics.renderPM import draw_to_file
                draw_to_file(drawing, buffer, fmt='PNG')
            buffer.seek(0)
            return base64.b64encode(buffer.getvalue()).decode('utf-8')
        except Exception as e:
            current_app.logger.error(f"Error generating barcode for order ID {order_id_str}: {e}")
            return None

    barcode_base64 = generate_order_barcode_base64(f"{order.id:09d}")

    return render_template('ordenes/imprimir_nota.html',
                           order=order,
                           company_info=company_info,
                           change=change,
                           barcode_base64=barcode_base64)

@routes_blueprint.route('/apartados/imprimir/<int:order_id>')
@login_required
def print_reservation_receipt(order_id):
    """Genera e imprime un recibo para un apartado."""
    order = Order.query.get_or_404(order_id)
    if order.status not in ['Apartado', 'Pagada']:
        flash('Esta orden no es un apartado y no se puede imprimir un recibo.', 'warning')
        return redirect(url_for('main.order_detail', order_id=order.id))
    
    company_info = CompanyInfo.query.first()

    # Helper function to generate barcode
    def generate_order_barcode_base64(order_id_str):
        """Generates a Code128 barcode image and returns it as a base64 string."""
        if not order_id_str:
            return None
        try:
            barcode = createBarcodeDrawing('Code128', value=order_id_str, barHeight=10*mm, barWidth=0.3*mm)
            drawing = Drawing(barcode.width, barcode.height)
            drawing.add(barcode)
            buffer = io.BytesIO()
            try:
                # First, try the C-based backend which is faster
                from reportlab.graphics import renderPM
                renderPM.drawToFile(drawing, buffer, fmt='PNG')
            except ImportError:
                # Fallback to the pure Python backend if the C-backend is not available
                from reportlab.graphics.renderPM import draw_to_file
                draw_to_file(drawing, buffer, fmt='PNG')
            buffer.seek(0)
            return base64.b64encode(buffer.getvalue()).decode('utf-8')
        except Exception as e:
            current_app.logger.error(f"Error generating barcode for order ID {order_id_str}: {e}")
            return None

    barcode_base64 = generate_order_barcode_base64(f"{order.id:09d}")

    return render_template('apartados/imprimir_recibo.html',
                           order=order,
                           company_info=company_info,
                           barcode_base64=barcode_base64)

# Nueva ruta de API para obtener la tasa de cambio actual
@routes_blueprint.route('/api/product_by_barcode/<barcode>')
@login_required
def api_product_by_barcode(barcode):
    """API endpoint to get product information by barcode."""
    product = Product.query.filter_by(barcode=barcode).first()
    if product:
        return jsonify({
            'id': product.id,
            'name': product.name,
            'codigo_producto': product.codigo_producto,
            'price_usd': product.price_usd,
            'cost_usd': product.cost_usd,
            'stock': product.stock_tienda # Devolver solo el stock de la tienda para la venta
        })
    else:
        return jsonify({'error': 'Producto no encontrado'}), 404

@routes_blueprint.route('/api/exchange_rate')
def api_exchange_rate():
    # CORRECCIÓN: Usar la nueva función para obtener la tasa en USD
    rate = get_cached_exchange_rate('USD') # type: ignore
    if rate:
        return jsonify(rate=rate)
    else:
        return jsonify(error="No se pudo obtener la tasa de cambio"), 500

@routes_blueprint.route('/api/check_stock', methods=['POST'])
@login_required
def api_check_stock():
    """
    API endpoint to check stock for a list of products.
    Expects JSON: {'products': [{'id': <int>, 'quantity': <int>}]}
    """
    data = request.get_json()
    if not data or 'products' not in data:
        return jsonify({'success': False, 'error': 'Invalid request format.'}), 400

    product_requests = data['products']
    product_ids = [p.get('id') for p in product_requests if p.get('id')]

    if not product_ids:
        return jsonify({'success': True}) # No products to check

    try:
        # Fetch all products in one query
        # Comprobar stock solo en el almacén de tienda (ID 1)
        products_in_db = Product.query.filter(Product.id.in_(product_ids)).options(joinedload(Product.stock_levels)).all()
        product_stock_map = {p.id: p.stock_tienda for p in products_in_db}

        errors = []
        for req in product_requests:
            req_id = req.get('id')
            req_qty = req.get('quantity')
            
            if not req_id or not isinstance(req_qty, int) or req_qty <= 0: # type: ignore
                continue

            current_stock = product_stock_map.get(int(req_id), 0)
            if current_stock < req_qty:
                product_name = next((p.name for p in products_in_db if p.id == int(req_id)), 'Desconocido')
                errors.append({
                    'id': req_id,
                    'name': product_name,
                    'stock': current_stock,
                    'requested': req_qty
                })

        if errors:
            return jsonify({'success': False, 'errors': errors})
        else:
            return jsonify({'success': True})

    except Exception as e:
        current_app.logger.error(f"Error in /api/check_stock: {e}")
        return jsonify({'success': False, 'error': 'Internal server error.'}), 500

@routes_blueprint.route('/api/search_clients')
@login_required
def api_search_clients():
    """API endpoint to search clients by cedula/rif or name."""
    query = request.args.get('q', '').strip()
    if not query:
        return jsonify(clients=[])

    # Search by cedula_rif or name (case insensitive, partial match)
    clients = Client.query.filter(
        or_(
            Client.cedula_rif.ilike(f'%{query}%'),
            Client.name.ilike(f'%{query}%')
        )
    ).limit(10).all()

    clients_data = []
    for client in clients:
        clients_data.append({
            'id': client.id,
            'name': client.name,
            'cedula_rif': client.cedula_rif,
            'email': client.email,
            'phone': client.phone,
            'address': client.address
        })

    return jsonify(clients=clients_data)

@routes_blueprint.route('/api/clientes/nuevo', methods=['POST'])
@login_required
def api_new_client():
    """API endpoint to create a new client and return its data."""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No se proporcionaron datos'}), 400

    name = data.get('name')
    cedula_rif = data.get('cedula_rif', '').strip()
    email = data.get('email', '').strip()

    if not name:
        return jsonify({'error': 'El nombre es requerido.'}), 400

    # Treat empty email as None
    if not email:
        email = None

    try:
        # Check for duplicates
        if cedula_rif and cedula_rif.strip() and Client.query.filter_by(cedula_rif=cedula_rif).first():
            return jsonify({'error': f'La Cédula/RIF "{cedula_rif}" ya está registrada.'}), 409
        
        if email and Client.query.filter_by(email=email).first():
            return jsonify({'error': f'El email "{email}" ya está registrado.'}), 409

        new_client = Client(
            name=name,
            cedula_rif=cedula_rif,
            email=email,
            phone=data.get('phone'),
            address=data.get('address')
        )
        db.session.add(new_client)
        db.session.commit()

        client_data = { 'id': new_client.id, 'name': new_client.name, 'cedula_rif': new_client.cedula_rif, 'email': new_client.email, 'phone': new_client.phone, 'address': new_client.address }
        return jsonify(client_data), 201

    except IntegrityError as e:
        db.session.rollback()
        current_app.logger.warning(f"IntegrityError al crear cliente vía API: {e}")
        return jsonify({'error': 'Error de base de datos. Es posible que el email o Cédula/RIF ya exista.'}), 409
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creando nuevo cliente vía API: {e}")
        return jsonify({'error': 'Ocurrió un error interno en el servidor.'}), 500

@routes_blueprint.route('/api/proveedores/nuevo', methods=['POST'])
@login_required
def api_new_provider():
    """API endpoint to create a new provider from a modal and return its data."""
    if not is_administrador():
        return jsonify({'error': 'Acceso denegado'}), 403

    data = request.get_json()
    if not data:
        return jsonify({'error': 'No se proporcionaron datos'}), 400

    name = data.get('name')
    tax_id = data.get('tax_id')

    if not name:
        return jsonify({'error': 'El nombre es requerido.'}), 400

    try:
        # Check for duplicates
        if Provider.query.filter_by(name=name).first():
            return jsonify({'error': f'El proveedor con el nombre "{name}" ya está registrado.'}), 409
        if tax_id and Provider.query.filter_by(tax_id=tax_id).first():
            return jsonify({'error': f'El RIF "{tax_id}" ya está registrado.'}), 409

        new_provider = Provider(
            name=name,
            tax_id=tax_id,
            phone=data.get('phone'),
            email=data.get('email'),
            address=data.get('address')
        )
        db.session.add(new_provider)
        db.session.commit()

        provider_data = {
            'id': new_provider.id,
            'name': new_provider.name,
            'tax_id': new_provider.tax_id,
            'phone': new_provider.phone,
            'email': new_provider.email,
            'address': new_provider.address
        }
        return jsonify(provider_data), 201
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creando nuevo proveedor vía API: {e}")
        return jsonify({'error': 'Ocurrió un error interno en el servidor.'}), 500

@routes_blueprint.route('/api/search_providers')
@login_required
def api_search_providers():
    """API endpoint to search providers by name, tax_id, or contact person."""
    if not is_administrador():
        return jsonify({'error': 'Acceso denegado'}), 403

    query = request.args.get('q', '').strip()
    if not query or len(query) < 2:
        return jsonify(providers=[])

    search_pattern = f'%{query}%'
    providers = Provider.query.filter(
        or_(
            Provider.name.ilike(search_pattern),
            Provider.tax_id.ilike(search_pattern),
            Provider.contact_person_name.ilike(search_pattern)
        )
    ).limit(10).all()

    providers_data = [{
        'id': p.id,
        'name': p.name,
        'tax_id': p.tax_id,
        'phone': p.phone,
        'email': p.email
    } for p in providers]

    return jsonify(providers=providers_data)

@routes_blueprint.route('/api/productos/nuevo', methods=['POST'])
@login_required
def api_new_product():
    """API endpoint to create a new product from a modal and return its data."""
    if not is_administrador():
        return jsonify({'error': 'Acceso denegado'}), 403

    data = request.get_json()
    if not data:
        return jsonify({'error': 'No se proporcionaron datos'}), 400

    # Validar campos requeridos
    required_fields = ['name', 'barcode', 'cost_usd', 'price_usd']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'error': f'El campo "{field}" es requerido.'}), 400

    try:
        # Verificar si el código de barras ya existe
        if Product.query.filter_by(barcode=data['barcode']).first():
            return jsonify({'error': f'El código de barras "{data["barcode"]}" ya está registrado.'}), 409

        new_prod = Product(
            name=data['name'],
            barcode=data['barcode'],
            cost_usd=float(data['cost_usd']),
            price_usd=float(data['price_usd']),
            codigo_producto=data.get('codigo_producto'),
            description=data.get('description'),
            image_url=data.get('image_url'),
            size=data.get('size'),
            color=data.get('color'),
            marca=data.get('marca'),
            grupo=data.get('grupo')
        )
        db.session.add(new_prod)
        db.session.commit()

        return jsonify(id=new_prod.id, name=new_prod.name, barcode=new_prod.barcode, cost_usd=new_prod.cost_usd, stock=0), 201
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creando nuevo producto vía API: {e}")
        return jsonify({'error': 'Ocurrió un error interno en el servidor.'}), 500

@routes_blueprint.route('/api/subscribe_web_push', methods=['POST'])
@login_required
def subscribe_web_push():
    """API para que el frontend registre una suscripción de Web Push."""
    subscription_data = request.get_json()
    if not subscription_data or 'endpoint' not in subscription_data:
        return jsonify({'success': False, 'error': 'Suscripción no válida.'}), 400

    try:
        subscription_json = json.dumps(subscription_data)
        
        # Buscar si ya existe para el usuario para no duplicar
        existing_device = UserDevice.query.filter_by(user_id=current_user.id, device_type='web').first()
        if existing_device:
            existing_device.fcm_token = subscription_json
            existing_device.last_login = get_current_time_ve()
        else:
            new_device = UserDevice(user_id=current_user.id, fcm_token=subscription_json, device_type='web')
            db.session.add(new_device)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Suscripción web push guardada.'}), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error al guardar suscripción web push para {current_user.id}: {e}")
        return jsonify({'success': False, 'error': 'Error interno del servidor.'}), 500

@routes_blueprint.route('/api/register_fcm_token', methods=['POST'])
@login_required
def register_fcm_token():
    """
    API endpoint for mobile apps to register their FCM token.
    """
    data = request.get_json()
    if not data or 'token' not in data:
        return jsonify({'success': False, 'error': 'Token no proporcionado.'}), 400

    token = data['token']
    device_type = data.get('device_type', 'android')

    try:
        # Verificar si el token ya existe para evitar duplicados
        existing_device = UserDevice.query.filter_by(fcm_token=token).first()
        if existing_device:
            # Si ya existe, solo actualizamos la fecha de último login y el usuario si ha cambiado
            existing_device.user_id = current_user.id
            existing_device.last_login = get_current_time_ve()
            message = 'Token de dispositivo actualizado.'
        else:
            # Si no existe, lo creamos
            new_device = UserDevice(user_id=current_user.id, fcm_token=token, device_type=device_type)
            db.session.add(new_device)
            message = 'Token de dispositivo registrado exitosamente.'
        
        db.session.commit()
        return jsonify({'success': True, 'message': message}), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error al registrar el token FCM para el usuario {current_user.id}: {e}")
        return jsonify({'success': False, 'error': 'Error interno del servidor.'}), 500

# --- Rutas de Finanzas ---

@routes_blueprint.route('/finanzas/bancos/lista')
@login_required
def bank_list():
    banks = Bank.query.order_by(Bank.name).all()
    return render_template('finanzas/bancos_lista.html', title='Lista de Bancos', banks=banks)

@routes_blueprint.route('/finanzas/bancos/nuevo', methods=['GET', 'POST'])
@login_required
def new_bank():
    if not is_superuser(): # Only Superuser can create new banks
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.new_order'))

    if request.method == 'POST':
        try:
            name = request.form.get('name')
            account_number = request.form.get('account_number')
            initial_balance = float(request.form.get('initial_balance', 0))
            new_bank = Bank(name=name, account_number=account_number, balance=initial_balance)
            db.session.add(new_bank)
            db.session.commit()
            flash('Banco creado exitosamente!', 'success')
            return redirect(url_for('main.bank_list'))
        except (ValueError, IntegrityError):
            db.session.rollback()
            flash('Error: Ya existe un banco con ese nombre o número de cuenta.', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Ocurrió un error inesperado: {e}', 'danger')
    return render_template('finanzas/nuevo_banco.html', title='Nuevo Banco')

@routes_blueprint.route('/finanzas/bancos/movimientos')
@login_required
def bank_movements():
    if not is_administrador(): # Superusuario, Gerente, administrador can view bank movements
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.new_order'))

    banks = Bank.query.order_by(Bank.name).all()
    return render_template('finanzas/seleccionar_cuenta.html', 
                           title='Movimientos Bancarios',
                           accounts=banks,
                           account_type='bank',
                           detail_route='main.bank_movement_detail')

@routes_blueprint.route('/finanzas/bancos/movimientos/<int:bank_id>')
@login_required
def bank_movement_detail(bank_id):
    if not is_administrador(): # Superusuario, Gerente, administrador can view bank movement details
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.new_order'))

    bank = Bank.query.get_or_404(bank_id)
    
    # Get payments related to this bank (direct transfers and POS terminals)
    pos_ids = [pos.id for pos in bank.pos_terminals]
    payments_query = Payment.query.filter(or_(Payment.bank_id == bank_id, Payment.pos_id.in_(pos_ids)))
    
    # Get manual movements
    manual_movements_query = ManualFinancialMovement.query.filter_by(bank_id=bank_id)
    
    # Combine and prepare for sorting
    combined_movements = []
    
    # Process payments (always income in VES)
    for p in payments_query.all():
        description_parts = [f"Pago de Orden #{p.order_id:09d}"]
        if p.method == 'transferencia':
            if p.reference:
                description_parts.append(f"Ref: {p.reference}")
            if p.issuing_bank:
                description_parts.append(f"Bco: {p.issuing_bank}")
            if p.sender_id:
                description_parts.append(f"CI/Tlf: {p.sender_id}")

        combined_movements.append({
            'date': p.date,
            'description': ". ".join(description_parts),
            'income': p.amount_ves_equivalent,
            'expense': 0,
            'currency': 'VES'
        })
        
    # Process manual movements
    for m in manual_movements_query.all():
        if m.currency != 'VES': continue
        combined_movements.append({
            'date': m.date,
            'description': m.description,
            'income': m.amount if m.movement_type == 'Ingreso' else 0,
            'expense': m.amount if m.movement_type == 'Egreso' else 0,
            'currency': m.currency
        })
        
    # Sort all movements by date (newest first)
    combined_movements.sort(key=lambda x: x['date'], reverse=True)

    return render_template('finanzas/movimientos_bancarios.html', 
                           title=f'Movimientos de {bank.name}', 
                           bank=bank,
                           movements=combined_movements)

@routes_blueprint.route('/finanzas/puntos-venta/lista')
@login_required
def pos_list():
    points_of_sale = PointOfSale.query.order_by(PointOfSale.name).all()
    return render_template('finanzas/pos_lista.html', title='Puntos de Venta', points_of_sale=points_of_sale)

@routes_blueprint.route('/finanzas/puntos-venta/nuevo', methods=['GET', 'POST'])
@login_required
def new_pos():
    if not is_superuser(): # Only Superuser can create new POS
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.new_order'))

    banks = Bank.query.order_by(Bank.name).all()
    if request.method == 'POST':
        try:
            name = request.form.get('name')
            bank_id = request.form.get('bank_id')
            if not bank_id:
                flash('Debe seleccionar un banco asociado.', 'danger')
            else:
                new_pos = PointOfSale(name=name, bank_id=bank_id)
                db.session.add(new_pos)
                db.session.commit()
                flash('Punto de Venta creado exitosamente!', 'success')
                return redirect(url_for('main.pos_list'))
        except IntegrityError:
            db.session.rollback()
            flash('Error: Ya existe un punto de venta con ese nombre.', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Ocurrió un error inesperado: {e}', 'danger')
    return render_template('finanzas/nuevo_pos.html', title='Nuevo Punto de Venta', banks=banks)

@routes_blueprint.route('/finanzas/caja/lista')
@login_required
def cashbox_list():
    cash_boxes = CashBox.query.order_by(CashBox.name).all()
    return render_template('finanzas/caja_lista.html', title='Cajas', cash_boxes=cash_boxes)

@routes_blueprint.route('/finanzas/caja/nueva', methods=['GET', 'POST'])
@login_required
def new_cashbox():
    if not is_superuser(): # Only Superuser can create new cashboxes
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.new_order'))

    if request.method == 'POST':
        try:
            name = request.form.get('name')
            balance_ves = float(request.form.get('balance_ves', 0))
            balance_usd = float(request.form.get('balance_usd', 0))
            new_box = CashBox(name=name, balance_ves=balance_ves, balance_usd=balance_usd)
            db.session.add(new_box)
            db.session.commit()
            flash('Caja creada exitosamente!', 'success')
            return redirect(url_for('main.cashbox_list'))
        except (ValueError, IntegrityError):
            db.session.rollback()
            flash('Error: Ya existe una caja con ese nombre.', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Ocurrió un error inesperado: {e}', 'danger')
    return render_template('finanzas/nueva_caja.html', title='Nueva Caja')

@routes_blueprint.route('/finanzas/caja/movimientos')
@login_required
def cashbox_movements():
    if not is_administrador(): # Superusuario, Gerente, administrador can view cashbox movements
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.new_order'))

    cash_boxes = CashBox.query.order_by(CashBox.name).all()
    return render_template('finanzas/seleccionar_cuenta.html', 
                           title='Movimientos de Caja',
                           accounts=cash_boxes,
                           account_type='cash_box',
                           detail_route='main.cashbox_movement_detail')

@routes_blueprint.route('/finanzas/caja/movimientos/<int:cash_box_id>')
@login_required
def cashbox_movement_detail(cash_box_id):
    if not is_administrador(): # Superusuario, Gerente, administrador can view cashbox movement details
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.new_order'))

    cash_box = CashBox.query.get_or_404(cash_box_id)
    
    payments_query = Payment.query.filter_by(cash_box_id=cash_box_id)
    manual_movements_query = ManualFinancialMovement.query.filter_by(cash_box_id=cash_box_id)
    
    movements_ves = []
    for p in payments_query.filter_by(currency_paid='VES').all():
        movements_ves.append({
            'id': f'P-{p.id}', 'type': 'payment', 'obj': p,
            'date': p.date, 'description': f"Pago de Orden #{p.order.id:09d}",
            'income': p.amount_paid, 'expense': 0, 'status': 'Aprobado'
        })
    for m in manual_movements_query.filter_by(currency='VES').all():
        desc = f"{m.description} (Por: {m.created_by_user.username if m.created_by_user else 'N/A'}, Recibe: {m.received_by or 'N/A'})"
        movements_ves.append({
            'id': f'M-{m.id}', 'type': 'manual', 'obj': m,
            'date': m.date, 'description': desc,
            'income': m.amount if m.movement_type == 'Ingreso' else 0,
            'expense': m.amount if m.movement_type == 'Egreso' else 0,
            'status': m.status
        })
    
    movements_usd = []
    for p in payments_query.filter_by(currency_paid='USD').all():
        movements_usd.append({
            'id': f'P-{p.id}', 'type': 'payment', 'obj': p,
            'date': p.date, 'description': f"Pago de Orden #{p.order.id:09d}",
            'income': p.amount_paid, 'expense': 0, 'status': 'Aprobado'
        })
    for m in manual_movements_query.filter_by(currency='USD').all():
        desc = f"{m.description} (Por: {m.created_by_user.username if m.created_by_user else 'N/A'}, Recibe: {m.received_by or 'N/A'})"
        movements_usd.append({
            'id': f'M-{m.id}', 'type': 'manual', 'obj': m,
            'date': m.date, 'description': desc,
            'income': m.amount if m.movement_type == 'Ingreso' else 0,
            'expense': m.amount if m.movement_type == 'Egreso' else 0,
            'status': m.status
        })
        
    movements_ves.sort(key=lambda x: x['date'], reverse=True)
    movements_usd.sort(key=lambda x: x['date'], reverse=True)

    return render_template('finanzas/movimientos_caja.html', 
                           title=f'Movimientos de {cash_box.name}', 
                           cash_box=cash_box,
                           movements_ves=movements_ves,
                           movements_usd=movements_usd)

@routes_blueprint.route('/finanzas/movimiento/nuevo', methods=['GET', 'POST'])
@login_required
def new_financial_movement():
    if not is_administrador(): # Superusuario, Gerente, administrador can create manual financial movements
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.new_order'))

    account_type = request.args.get('account_type')
    account_id = request.args.get('account_id', type=int)

    if not account_type or not account_id:
        flash('Tipo de cuenta o ID no especificado.', 'danger')
        return redirect(url_for('main.dashboard'))

    account = None
    if account_type == 'bank':
        account = Bank.query.get_or_404(account_id)
    elif account_type == 'cash_box':
        account = CashBox.query.get_or_404(account_id)
    else:
        flash('Tipo de cuenta inválido.', 'danger')
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        try:
            description = request.form.get('description')
            amount = float(request.form.get('amount'))
            currency = request.form.get('currency')
            movement_type = request.form.get('movement_type')

            if not all([description, amount, currency, movement_type]):
                raise ValueError("Todos los campos son requeridos.")
            if amount <= 0:
                raise ValueError("El monto debe ser positivo.")

            new_mov = ManualFinancialMovement(
                description=description, amount=amount, currency=currency, movement_type=movement_type
            )

            if account_type == 'bank':
                if currency != 'VES':
                    raise ValueError("Los movimientos bancarios solo pueden ser en VES.")
                
                new_mov.bank_id = account_id
                if account.currency != 'VES':
                    raise ValueError(f"El banco '{account.name}' no opera en VES.")
                if movement_type == 'Ingreso':
                    account.balance += amount
                else:
                    account.balance -= amount
            elif account_type == 'cash_box':
                new_mov.cash_box_id = account_id
                if currency == 'VES':
                    if movement_type == 'Ingreso':
                        account.balance_ves += amount
                    else:
                        account.balance_ves -= amount
                elif currency == 'USD':
                    if movement_type == 'Ingreso':
                        account.balance_usd += amount
                    else:
                        account.balance_usd -= amount
                else:
                    raise ValueError("Moneda no válida para la caja.")

            log_user_activity(
                action="Registró movimiento manual",
                details=f"Tipo: {movement_type}, Monto: {amount:.2f} {currency}, Cuenta: {account.name}, Desc: {description}",
                target_id=new_mov.id,
                target_type="ManualFinancialMovement"
            )

            db.session.add(new_mov)
            db.session.commit()

            flash('Movimiento registrado exitosamente.', 'success')
            if account_type == 'bank':
                return redirect(url_for('main.bank_movement_detail', bank_id=account_id))
            else:
                return redirect(url_for('main.cashbox_movement_detail', cash_box_id=account_id))

        except (ValueError, TypeError) as e:
            db.session.rollback()
            flash(f'Error al registrar el movimiento: {e}', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Ocurrió un error inesperado: {e}', 'danger')

    return render_template('finanzas/nuevo_movimiento.html',
                           title='Nuevo Movimiento Manual',
                           account=account,
                           account_type=account_type)

@routes_blueprint.route('/finanzas/caja/retiro', methods=['GET', 'POST'])
@login_required
def new_cash_withdrawal():
    cash_boxes = CashBox.query.order_by(CashBox.name).all()

    if request.method == 'POST':
        try:
            cash_box_id = request.form.get('cash_box_id', type=int) # type: ignore
            amount = float(request.form.get('amount'))
            currency = request.form.get('currency')
            description = request.form.get('description')
            received_by = request.form.get('received_by')
            date_str = request.form.get('date')

            if not all([cash_box_id, amount, currency, description, received_by]):
                raise ValueError("Todos los campos son requeridos.")
            if amount <= 0:
                raise ValueError("El monto debe ser positivo.")

            cash_box = CashBox.query.get_or_404(cash_box_id)

            # Always check balance before creating request
            if currency == 'VES':
                if cash_box.balance_ves < amount:
                    raise ValueError(f"Saldo insuficiente en la caja '{cash_box.name}' para VES. Saldo actual: {cash_box.balance_ves:.2f}")
            elif currency == 'USD':
                if cash_box.balance_usd < amount:
                    raise ValueError(f"Saldo insuficiente en la caja '{cash_box.name}' para USD. Saldo actual: {cash_box.balance_usd:.2f}")
            else:
                raise ValueError("Moneda no válida para la caja.")

            is_admin = is_gerente() # Gerente and Superuser can approve their own withdrawals
            
            withdrawal_date = get_current_time_ve()
            if date_str:
                try:
                    naive_dt = datetime.strptime(date_str, '%Y-%m-%dT%H:%M')
                    withdrawal_date = VE_TIMEZONE.localize(naive_dt)
                except (ValueError, TypeError):
                    current_app.logger.warning(f"Invalid date format for withdrawal: '{date_str}'. Falling back to now.")

            new_mov = ManualFinancialMovement(
                description=description, amount=amount, currency=currency, movement_type='Egreso',
                cash_box_id=cash_box_id, received_by=received_by, created_by_user_id=current_user.id,
                status='Aprobado' if is_admin else 'Pendiente', date=withdrawal_date
            )

            if is_admin:
                # Admins approve their own withdrawals instantly
                new_mov.approved_by_user_id = current_user.id
                new_mov.date_approved = get_current_time_ve()
                # Update balance
                if currency == 'VES':
                    cash_box.balance_ves -= amount
                elif currency == 'USD':
                    cash_box.balance_usd -= amount
            
            log_user_activity(
                action="Solicitó/Aprobó retiro de efectivo",
                details=f"Monto: {amount:.2f} {currency} de la caja '{cash_box.name}'. Estado: {new_mov.status}",
                target_id=new_mov.id,
                target_type="ManualFinancialMovement"
            )

            db.session.add(new_mov)
            db.session.commit()

            if is_admin:
                flash('Retiro de efectivo registrado y aprobado exitosamente. Imprimiendo recibo...', 'success')
                return redirect(url_for('main.print_withdrawal_receipt', movement_id=new_mov.id))
            else:
                # Create notification for admins
                notification_message = f"El usuario {current_user.username} solicita aprobación para un retiro de {amount:.2f} {currency}."
                notification_link = url_for('main.pending_withdrawals')
                create_notification_for_admins(notification_message, notification_link)
                
                flash('Solicitud de retiro de efectivo enviada para aprobación.', 'info')
                return redirect(url_for('main.my_withdrawals'))

        except (ValueError, TypeError) as e:
            db.session.rollback()
            flash(f'Error al registrar el retiro: {e}', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Ocurrió un error inesperado: {e}', 'danger')

    return render_template('finanzas/retiro_caja.html', title='Retiro de Efectivo de Caja', cash_boxes=cash_boxes)

@routes_blueprint.route('/finanzas/caja/retiro/imprimir/<int:movement_id>')
@login_required
def print_withdrawal_receipt(movement_id):
    movement = ManualFinancialMovement.query.get_or_404(movement_id)
    company_info = CompanyInfo.query.first()
    if movement.movement_type != 'Egreso' or not movement.cash_box_id:
        flash('Movimiento no válido para generar recibo de retiro.', 'danger')
        return redirect(url_for('main.new_order'))
    
    # New check for approval
    if movement.status != 'Aprobado':
        flash('Este retiro aún no ha sido aprobado. No se puede imprimir el recibo.', 'warning')
        if current_user.role == 'Administrador':
            return redirect(url_for('main.pending_withdrawals'))
        else:
            return redirect(url_for('main.new_order'))

    return render_template('finanzas/imprimir_recibo_retiro.html', movement=movement, company_info=company_info)

@routes_blueprint.route('/finanzas/mis-retiros')
@login_required
def my_withdrawals():
    """Muestra las solicitudes de retiro creadas por el usuario actual."""
    movements = ManualFinancialMovement.query.filter_by(
        created_by_user_id=current_user.id,
        movement_type='Egreso'
    ).order_by(ManualFinancialMovement.date.desc()).all()
    
    return render_template('finanzas/mis_retiros.html', title='Mis Solicitudes de Retiro', movements=movements)

@routes_blueprint.route('/finanzas/retiros-pendientes')
@login_required
def pending_withdrawals():
    if not is_administrador(): # Superusuario, Gerente, administrador can view pending withdrawals
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.new_order'))
    
    pending = ManualFinancialMovement.query.filter_by(status='Pendiente', movement_type='Egreso').order_by(ManualFinancialMovement.date.desc()).all()
    
    return render_template('finanzas/retiros_pendientes.html', title='Retiros Pendientes de Aprobación', movements=pending)

@routes_blueprint.route('/finanzas/retiro/procesar/<int:movement_id>/<string:action>', methods=['POST'])
@login_required
def process_withdrawal(movement_id, action):
    if not is_administrador(): # Superusuario, Gerente, administrador can process withdrawals
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.pending_withdrawals'))

    movement = ManualFinancialMovement.query.get_or_404(movement_id)
    if movement.status != 'Pendiente':
        flash('Este retiro ya ha sido procesado.', 'warning')
        return redirect(url_for('main.pending_withdrawals'))

    try:
        if action == 'approve':
            cash_box = movement.cash_box
            if not cash_box:
                raise ValueError("El movimiento no está asociado a ninguna caja.")

            # Final balance check
            if movement.currency == 'VES':
                if cash_box.balance_ves < movement.amount:
                    raise ValueError(f"Saldo insuficiente en la caja '{cash_box.name}' para aprobar este retiro.")
                cash_box.balance_ves -= movement.amount
            elif movement.currency == 'USD':
                if cash_box.balance_usd < movement.amount:
                    raise ValueError(f"Saldo insuficiente en la caja '{cash_box.name}' para aprobar este retiro.")
                cash_box.balance_usd -= movement.amount
            
            movement.status = 'Aprobado'
            flash_message = 'Retiro aprobado y saldo de caja actualizado.'
            log_user_activity(
                action="Aprobó retiro de efectivo",
                details=f"Aprobó retiro de {movement.amount:.2f} {movement.currency} solicitado por {movement.created_by_user.username}.",
                target_id=movement.id,
                target_type="ManualFinancialMovement"
            )
            flash_category = 'success'

        elif action == 'reject':
            movement.status = 'Rechazado'
            flash_message = 'Retiro rechazado.'
            log_user_activity(
                action="Rechazó retiro de efectivo",
                details=f"Rechazó retiro de {movement.amount:.2f} {movement.currency} solicitado por {movement.created_by_user.username}.",
                target_id=movement.id,
                target_type="ManualFinancialMovement"
            )
            flash_category = 'info'
        
        else:
            raise ValueError("Acción no válida.")

        movement.approved_by_user_id = current_user.id
        movement.date_approved = get_current_time_ve()
        db.session.commit()
        flash(flash_message, flash_category)

    except (ValueError, IntegrityError) as e:
        db.session.rollback()
        flash(f'Error al procesar el retiro: {e}', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Ocurrió un error inesperado: {e}', 'danger')

    return redirect(url_for('main.pending_withdrawals'))

# --- Cierre Diario ---

@routes_blueprint.route('/finanzas/cierre-diario', methods=['GET'])
@login_required
def daily_closing():
    """Shows the page for generating the daily closing report. Accessible by administrador, Gerente, Superusuario."""
    if not is_vendedor():
        flash('No tienes permiso para acceder a esta página.', 'danger')
        return redirect(url_for('main.new_order'))
    date_str = request.args.get('date', get_current_time_ve().date().strftime('%Y-%m-%d'))
    try:
        report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        report_date = get_current_time_ve().date()
        flash('Fecha inválida. Usando la fecha de hoy.', 'warning')

    return render_template('finanzas/cierre_diario.html', title='Cierre Diario', report_date=report_date)

@routes_blueprint.route('/finanzas/reporte-mensual', methods=['GET'])
@login_required
def reporte_mensual():
    """Muestra la página dedicada a generar el reporte mensual en PDF."""
    if not is_vendedor():
        flash('No tienes permiso para acceder a esta página.', 'danger')
        return redirect(url_for('main.new_order'))

    # Pasamos la fecha de hoy para que el JavaScript en la plantilla
    # pueda pre-seleccionar el mes anterior de forma fiable.
    today = get_current_time_ve().date()
    
    return render_template('finanzas/reporte_mensual.html', title='Reporte Mensual', today=today)


@routes_blueprint.route('/finanzas/cierre-diario/imprimir', methods=['GET'])
@login_required
def print_daily_closing_report():
    
    date_str = request.args.get('date')
    try:
        report_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else get_current_time_ve().date()
    except (ValueError, TypeError):
        report_date = get_current_time_ve().date()

    start_of_day = VE_TIMEZONE.localize(datetime.combine(report_date, datetime.min.time()))
    end_of_day = VE_TIMEZONE.localize(datetime.combine(report_date, datetime.max.time()))
    
    company_info = CompanyInfo.query.first()
    current_rate_usd = get_cached_exchange_rate('USD') or 1.0

    # --- 1. Sales Summary ---
    # Se incluyen todas las órdenes (contado, crédito, apartado) para el total de ventas y CMV.
    orders_today = Order.query.filter(Order.date_created.between(start_of_day, end_of_day)).options(joinedload(Order.items)).all()
    
    sales_summary = {
        'contado': {'count': 0, 'amount_ves': 0.0, 'amount_usd': 0.0},
        'credito': {'count': 0, 'amount_ves': 0.0, 'amount_usd': 0.0},
        'apartado': {'count': 0, 'amount_ves': 0.0, 'amount_usd': 0.0},
        'total': {'count': 0, 'amount_ves': 0.0, 'amount_usd': 0.0, 'cogs_ves': 0.0, 'cogs_usd': 0.0}
    }
    for order in orders_today:
        sales_summary['total']['count'] += 1
        sales_summary['total']['amount_ves'] += order.total_amount
        sales_summary['total']['amount_usd'] += order.total_amount_usd
        
        # Calcular el costo de la mercancía vendida (CMV) para esta orden
        order_cogs_ves = 0
        order_cogs_usd = 0
        rate = order.exchange_rate_at_sale or current_rate_usd
        if rate > 0:
            for item in order.items:
                if item.cost_at_sale_ves is not None:
                    cost_item_ves = item.cost_at_sale_ves * item.quantity
                    order_cogs_ves += cost_item_ves
                    order_cogs_usd += cost_item_ves / rate

        sales_summary['total']['cogs_ves'] += order_cogs_ves
        sales_summary['total']['cogs_usd'] += order_cogs_usd

        # Clasificar por tipo de orden para el desglose
        if order.order_type == 'regular':
            sales_summary['contado']['count'] += 1
            sales_summary['contado']['amount_ves'] += order.total_amount
            sales_summary['contado']['amount_usd'] += order.total_amount_usd
        elif order.order_type == 'credit':
            sales_summary['credito']['count'] += 1
            sales_summary['credito']['amount_ves'] += order.total_amount
            sales_summary['credito']['amount_usd'] += order.total_amount_usd
        elif order.order_type == 'reservation':
            sales_summary['apartado']['count'] += 1
            sales_summary['apartado']['amount_ves'] += order.total_amount
            sales_summary['apartado']['amount_usd'] += order.total_amount_usd

    # --- 2. Payments Summary by Method ---
    payments_today = Payment.query.filter(Payment.date.between(start_of_day, end_of_day)).all()
    
    payments_summary = {
        'efectivo_ves': {'amount': 0.0},
        'efectivo_usd': {'amount': 0.0, 'amount_ves_equivalent': 0.0},
        'transferencia': {'amount': 0.0}, # Incluye Pago Móvil
        'punto_de_venta': {'amount': 0.0},
        'total_ves': 0.0,
        'total_usd': 0.0
    }
    for payment in payments_today:
        payments_summary['total_ves'] += payment.amount_ves_equivalent
        payments_summary['total_usd'] += payment.amount_usd_equivalent
        if payment.method in ['efectivo_ves', 'efectivo_usd', 'transferencia', 'punto_de_venta']:
            if payment.method == 'efectivo_usd':
                payments_summary[payment.method]['amount'] += payment.amount_paid
                payments_summary[payment.method]['amount_ves_equivalent'] += payment.amount_ves_equivalent
            else:
                payments_summary[payment.method]['amount'] += payment.amount_ves_equivalent # Usar el equivalente en VES para métodos en VES

    # --- 3. Cash Box Movements ---
    cash_boxes = CashBox.query.all()
    cash_box_movements = {}
    for box in cash_boxes:
        cash_box_movements[box.name] = {
            'income_ves': 0.0, 'expense_ves': 0.0,
            'income_usd': 0.0, 'expense_usd': 0.0,
            'initial_balance_ves': 0.0, 'initial_balance_usd': 0.0,
            'final_balance_ves': box.balance_ves, 'final_balance_usd': box.balance_usd
        }

    # Payments into cash boxes
    cash_payments = Payment.query.filter(Payment.date.between(start_of_day, end_of_day), Payment.cash_box_id.isnot(None)).all()
    for p in cash_payments:
        if p.cash_box:
            if p.currency_paid == 'VES': cash_box_movements[p.cash_box.name]['income_ves'] += p.amount_paid
            elif p.currency_paid == 'USD': cash_box_movements[p.cash_box.name]['income_usd'] += p.amount_paid

    # Manual movements for cash boxes
    manual_cash_movements = ManualFinancialMovement.query.filter(ManualFinancialMovement.date.between(start_of_day, end_of_day), ManualFinancialMovement.cash_box_id.isnot(None)).all()
    for m in manual_cash_movements:
        if m.cash_box:
            if m.currency == 'VES':
                if m.movement_type == 'Ingreso': cash_box_movements[m.cash_box.name]['income_ves'] += m.amount
                elif m.movement_type == 'Egreso' and m.status == 'Aprobado': cash_box_movements[m.cash_box.name]['expense_ves'] += m.amount
            elif m.currency == 'USD':
                if m.movement_type == 'Ingreso': cash_box_movements[m.cash_box.name]['income_usd'] += m.amount
                elif m.movement_type == 'Egreso' and m.status == 'Aprobado': cash_box_movements[m.cash_box.name]['expense_usd'] += m.amount
    
    for box_name, data in cash_box_movements.items():
        data['initial_balance_ves'] = data['final_balance_ves'] - data['income_ves'] + data['expense_ves']
        data['initial_balance_usd'] = data['final_balance_usd'] - data['income_usd'] + data['expense_usd']

    # --- 4. Bank Account Movements ---
    banks = Bank.query.all()
    bank_movements = {}
    for bank in banks:
        bank_movements[bank.name] = {'income_ves': 0.0, 'expense_ves': 0.0, 'initial_balance_ves': 0.0, 'final_balance_ves': bank.balance}

    bank_payments = Payment.query.filter(Payment.date.between(start_of_day, end_of_day), or_(Payment.bank_id.isnot(None), Payment.pos_id.isnot(None))).all()
    for p in bank_payments:
        target_bank = p.bank or (p.pos.bank if p.pos else None)
        if target_bank and target_bank.name in bank_movements: bank_movements[target_bank.name]['income_ves'] += p.amount_ves_equivalent

    manual_bank_movements = ManualFinancialMovement.query.filter(ManualFinancialMovement.date.between(start_of_day, end_of_day), ManualFinancialMovement.bank_id.isnot(None)).all()
    for m in manual_bank_movements:
        if m.bank and m.currency == 'VES':
            if m.movement_type == 'Ingreso': bank_movements[m.bank.name]['income_ves'] += m.amount
            elif m.movement_type == 'Egreso' and m.status == 'Aprobado': bank_movements[m.bank.name]['expense_ves'] += m.amount
    
    for bank_name, data in bank_movements.items():
        data['initial_balance_ves'] = data['final_balance_ves'] - data['income_ves'] + data['expense_ves']

    # --- 5. Cash Withdrawals ---
    cash_withdrawals_today = ManualFinancialMovement.query.filter(ManualFinancialMovement.date.between(start_of_day, end_of_day), ManualFinancialMovement.movement_type == 'Egreso', ManualFinancialMovement.cash_box_id.isnot(None), ManualFinancialMovement.status == 'Aprobado').options(joinedload(ManualFinancialMovement.created_by_user), joinedload(ManualFinancialMovement.cash_box)).all()

    return render_template('finanzas/imprimir_cierre_diario.html', title=f'Cierre Diario - {report_date.strftime("%d/%m/%Y")}', today=report_date, company_info=company_info, sales_summary=sales_summary, payments_summary=payments_summary, cash_box_movements=cash_box_movements, bank_movements=bank_movements, cash_withdrawals_today=cash_withdrawals_today, current_rate_usd=current_rate_usd, user=current_user)

@routes_blueprint.route('/finanzas/cierre-diario/pdf', methods=['GET'])
@login_required
def print_daily_closing_report_pdf():
    """
    Gathers all data for a specific day and generates a full A4 PDF report.
    """
    
    from flask import make_response

    date_str = request.args.get('date')
    try:
        report_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else get_current_time_ve().date()
    except (ValueError, TypeError):
        report_date = get_current_time_ve().date()

    start_dt = VE_TIMEZONE.localize(datetime.combine(report_date, datetime.min.time()))
    end_dt = VE_TIMEZONE.localize(datetime.combine(report_date, datetime.max.time()))
    
    report_period = f"para el día {report_date.strftime('%d/%m/%Y')}"
    currency_symbol = "$"

    # --- Reutilizar la lógica de cálculo del reporte de ticket para consistencia ---
    company_info = CompanyInfo.query.first()
    current_rate_usd = get_cached_exchange_rate('USD') or 1.0

    # 1. Resumen de Ventas y CMV (Cost of Merchandise Vended)
    orders_today = Order.query.filter(Order.date_created.between(start_dt, end_dt)).options(joinedload(Order.items).joinedload(OrderItem.product)).all()
    
    # Modificación: Crear un desglose de ventas por tipo, similar al reporte de ticket.
    sales_summary = {
        'contado': {'count': 0, 'amount_usd': 0.0},
        'credito': {'count': 0, 'amount_usd': 0.0},
        'apartado': {'count': 0, 'amount_usd': 0.0},
        'total': {'count': 0, 'amount_ves': 0.0, 'amount_usd': 0.0, 'cogs_ves': 0.0, 'cogs_usd': 0.0}
    }
    variable_expenses_usd = 0.0
    cost_structure = CostStructure.query.first() or CostStructure()

    for order in orders_today:
        # Clasificar por tipo de orden para el desglose
        if order.order_type == 'regular':
            sales_summary['contado']['count'] += 1
            sales_summary['contado']['amount_usd'] += order.total_amount_usd
        elif order.order_type == 'credit':
            sales_summary['credito']['count'] += 1
            sales_summary['credito']['amount_usd'] += order.total_amount_usd
        elif order.order_type == 'reservation':
            sales_summary['apartado']['count'] += 1
            sales_summary['apartado']['amount_usd'] += order.total_amount_usd

        sales_summary['total']['count'] += 1
        sales_summary['total']['amount_usd'] += order.total_amount_usd
        
        order_cogs_usd = 0
        rate = order.exchange_rate_at_sale or current_rate_usd
        if rate > 0:
            for item in order.items:
                # Calcular CMV
                if item.cost_at_sale_ves is not None:
                    order_cogs_usd += (item.cost_at_sale_ves * item.quantity) / rate
                
                # Calcular Gastos Variables
                item_revenue_usd = (item.price * item.quantity) / rate
                var_sales_exp_pct = item.product.variable_selling_expense_percent if item.product and item.product.variable_selling_expense_percent > 0 else (cost_structure.default_sales_commission_percent or 0)
                var_marketing_pct = item.product.variable_marketing_percent if item.product and item.product.variable_marketing_percent > 0 else (cost_structure.default_marketing_percent or 0)
                variable_expenses_usd += item_revenue_usd * (var_sales_exp_pct + var_marketing_pct)

        sales_summary['total']['cogs_usd'] += order_cogs_usd

    # --- P&L Summary for the day ---
    pnl_summary = {
        'sales': sales_summary['total']['amount_usd'],
        'cogs': sales_summary['total']['cogs_usd'],
        'variable_expenses': variable_expenses_usd,
        'fixed_expenses': 0,
        'gross_profit': 0,
        'net_profit': 0
    }
    monthly_fixed_costs = (cost_structure.monthly_rent or 0) + (cost_structure.monthly_utilities or 0) + (cost_structure.monthly_fixed_taxes or 0)
    pnl_summary['fixed_expenses'] = monthly_fixed_costs / 30.44 # Prorated daily
    pnl_summary['gross_profit'] = pnl_summary['sales'] - pnl_summary['cogs']
    pnl_summary['net_profit'] = pnl_summary['gross_profit'] - pnl_summary['variable_expenses'] - pnl_summary['fixed_expenses']

    # --- Cash Flow Summary (calculado hacia atrás para mejor rendimiento) ---
    banks = Bank.query.all()
    bank_balances = []
    for bank in banks:
        inflows_ves = (db.session.query(func.sum(Payment.amount_ves_equivalent)).filter(or_(Payment.bank_id == bank.id, Payment.pos.has(bank_id=bank.id)), Payment.date.between(start_dt, end_dt)).scalar() or 0.0) + (db.session.query(func.sum(ManualFinancialMovement.amount)).filter(ManualFinancialMovement.bank_id == bank.id, ManualFinancialMovement.date.between(start_dt, end_dt), ManualFinancialMovement.movement_type == 'Ingreso', ManualFinancialMovement.currency == 'VES', ManualFinancialMovement.status == 'Aprobado').scalar() or 0.0)
        outflows_ves = db.session.query(func.sum(ManualFinancialMovement.amount)).filter(ManualFinancialMovement.bank_id == bank.id, ManualFinancialMovement.date.between(start_dt, end_dt), ManualFinancialMovement.movement_type == 'Egreso', ManualFinancialMovement.status == 'Aprobado', ManualFinancialMovement.currency == 'VES').scalar() or 0.0
        final_balance_ves = bank.balance
        initial_balance_ves = final_balance_ves - inflows_ves + outflows_ves
        bank_balances.append({'name': bank.name, 'initial_balance_ves': initial_balance_ves, 'inflows_ves': inflows_ves, 'outflows_ves': outflows_ves, 'final_balance_ves': final_balance_ves})

    cash_boxes = CashBox.query.all()
    cash_box_balances = []
    for box in cash_boxes:
        inflows_ves = (db.session.query(func.sum(Payment.amount_paid)).filter(Payment.cash_box_id == box.id, Payment.date.between(start_dt, end_dt), Payment.currency_paid == 'VES').scalar() or 0.0) + (db.session.query(func.sum(ManualFinancialMovement.amount)).filter(ManualFinancialMovement.cash_box_id == box.id, ManualFinancialMovement.date.between(start_dt, end_dt), ManualFinancialMovement.movement_type == 'Ingreso', ManualFinancialMovement.currency == 'VES', ManualFinancialMovement.status == 'Aprobado').scalar() or 0.0)
        outflows_ves = db.session.query(func.sum(ManualFinancialMovement.amount)).filter(ManualFinancialMovement.cash_box_id == box.id, ManualFinancialMovement.date.between(start_dt, end_dt), ManualFinancialMovement.movement_type == 'Egreso', ManualFinancialMovement.status == 'Aprobado', ManualFinancialMovement.currency == 'VES').scalar() or 0.0
        final_balance_ves = box.balance_ves
        initial_balance_ves = final_balance_ves - inflows_ves + outflows_ves

        inflows_usd = (db.session.query(func.sum(Payment.amount_paid)).filter(Payment.cash_box_id == box.id, Payment.date.between(start_dt, end_dt), Payment.currency_paid == 'USD').scalar() or 0.0) + (db.session.query(func.sum(ManualFinancialMovement.amount)).filter(ManualFinancialMovement.cash_box_id == box.id, ManualFinancialMovement.date.between(start_dt, end_dt), ManualFinancialMovement.movement_type == 'Ingreso', ManualFinancialMovement.currency == 'USD', ManualFinancialMovement.status == 'Aprobado').scalar() or 0.0)
        outflows_usd = db.session.query(func.sum(ManualFinancialMovement.amount)).filter(ManualFinancialMovement.cash_box_id == box.id, ManualFinancialMovement.date.between(start_dt, end_dt), ManualFinancialMovement.movement_type == 'Egreso', ManualFinancialMovement.status == 'Aprobado', ManualFinancialMovement.currency == 'USD').scalar() or 0.0
        final_balance_usd = box.balance_usd
        initial_balance_usd = final_balance_usd - inflows_usd + outflows_usd

        cash_box_balances.append({'name': box.name, 'initial_balance_ves': initial_balance_ves, 'inflows_ves': inflows_ves, 'outflows_ves': outflows_ves, 'final_balance_ves': final_balance_ves, 'initial_balance_usd': initial_balance_usd, 'inflows_usd': inflows_usd, 'outflows_usd': outflows_usd, 'final_balance_usd': final_balance_usd})

    # --- Generate Chart ---
    pnl_chart_base64 = generate_pnl_chart_base64(pnl_summary, currency_symbol)

    generation_date_str = get_current_time_ve().strftime("%d/%m/%Y %H:%M:%S")

    context = {
        'report_period': report_period,
        'generation_date': generation_date_str,
        'currency_symbol': currency_symbol,
        'pnl_summary': pnl_summary,
        'pnl_chart_base64': pnl_chart_base64,
        'bank_balances': bank_balances,
        'cash_box_balances': cash_box_balances,
        'orders_today': orders_today,
        'sales_summary': sales_summary, # Pasar el desglose de ventas a la plantilla
        'report_date': report_date,
    }
    html_string = render_template('pdf/reporte_diario_pdf.html', **context)

    pdf_file = HTML(string=html_string, base_url=request.base_url).write_pdf()

    response = make_response(pdf_file)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=cierre_diario_{report_date.strftime("%Y_%m_%d")}.pdf'
    
    return response

# --- Rutas de Almacenes ---

@routes_blueprint.route('/almacenes/lista', methods=['GET', 'POST'])
@login_required
def warehouse_list():
    """Muestra la lista de almacenes y permite crear nuevos."""
    if not is_gerente():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        try:
            name = request.form.get('name')
            is_sellable = 'is_sellable' in request.form

            if not name:
                raise ValueError("El nombre del almacén es obligatorio.")

            new_warehouse = Warehouse(name=name, is_sellable=is_sellable)
            db.session.add(new_warehouse)
            db.session.commit()
            flash(f'Almacén "{name}" creado exitosamente.', 'success')
        except (ValueError, IntegrityError) as e:
            db.session.rollback()
            flash(f'Error al crear el almacén: {e}', 'danger')
        return redirect(url_for('main.warehouse_list'))

    warehouses = Warehouse.query.order_by(Warehouse.id).all()
    return render_template('almacenes/lista.html', title='Gestión de Almacenes', warehouses=warehouses)

@routes_blueprint.route('/almacenes/traslados', methods=['GET', 'POST'])
@login_required
def warehouse_transfer():
    """Permite realizar traslados de stock entre almacenes."""
    
    if request.method == 'POST':
        try:
            from_warehouse_id = request.form.get('from_warehouse_id', type=int)
            to_warehouse_id = request.form.get('to_warehouse_id', type=int)
            reason = request.form.get('reason')
            product_ids = request.form.getlist('product_id[]')
            quantities = request.form.getlist('quantity[]')
            comments = request.form.getlist('comment[]') # Capturar los comentarios

            if not all([from_warehouse_id, to_warehouse_id, reason, product_ids, quantities, comments]):
                raise ValueError("Todos los campos son obligatorios.")
            if from_warehouse_id == to_warehouse_id:
                raise ValueError("El almacén de origen y destino no pueden ser el mismo.")
            
            # Generar el código de traslado usando la nueva secuencia
            # La secuencia 'warehouse_transfer_code_seq' debe ser creada en la BD
            # y configurada para que inicie en 1807001.
            next_val = db.session.execute(text("SELECT nextval('warehouse_transfer_code_seq')")).scalar()
            transfer_code = f"TR{next_val}"

            # Crear el registro principal del traslado
            transfer = WarehouseTransfer(
                reason=reason,
                user_id=current_user.id,
                transfer_code=transfer_code
            )
            db.session.add(transfer)
            db.session.flush() # Para obtener el ID del traslado

            for i, (p_id, qty_str) in enumerate(zip(product_ids, quantities)):
                product_id = int(p_id)
                quantity = int(qty_str)
                comment = comments[i]

                if quantity <= 0: continue

                # Descontar del almacén de origen
                origin_stock = ProductStock.query.filter_by(product_id=product_id, warehouse_id=from_warehouse_id).first()
                if not origin_stock or origin_stock.quantity < quantity:
                    product_name = Product.query.get(product_id).name
                    raise ValueError(f"Stock insuficiente para '{product_name}' en el almacén de origen.")
                origin_stock.quantity -= quantity

                # Añadir al almacén de destino
                destination_stock = ProductStock.query.filter_by(product_id=product_id, warehouse_id=to_warehouse_id).first()
                if not destination_stock:
                    destination_stock = ProductStock(product_id=product_id, warehouse_id=to_warehouse_id, quantity=0)
                    db.session.add(destination_stock)
                destination_stock.quantity += quantity

                # Registrar movimientos
                movement_out = Movement(product_id=product_id, type='Salida', warehouse_id=from_warehouse_id, quantity=quantity, document_id=transfer.id, document_type='Traslado de Almacén', description=f"Hacia almacén ID {to_warehouse_id}", comment=comment)
                movement_in = Movement(product_id=product_id, type='Entrada', warehouse_id=to_warehouse_id, quantity=quantity, document_id=transfer.id, document_type='Traslado de Almacén', description=f"Desde almacén ID {from_warehouse_id}", comment=comment)
                db.session.add_all([movement_out, movement_in])

            log_user_activity(
                action="Realizó traslado de inventario",
                details=f"Traslado con código {transfer.transfer_code} desde almacén ID {from_warehouse_id} a ID {to_warehouse_id}.",
                target_id=transfer.id,
                target_type="WarehouseTransfer"
            )

            db.session.commit()
            flash('Traslado realizado exitosamente. Puede imprimir el reporte si lo necesita.', 'success')
            return redirect(url_for('main.transfer_detail', transfer_id=transfer.id))

        except (ValueError, IntegrityError) as e:
            db.session.rollback()
            flash(f'Error al procesar el traslado: {e}', 'danger')
        return redirect(url_for('main.warehouse_transfer'))

    warehouses = Warehouse.query.order_by(Warehouse.id).all()
    products = Product.query.order_by(Product.name).all()
    return render_template('almacenes/traslados.html', title='Traslado entre Almacenes', warehouses=warehouses, products=products)

@routes_blueprint.route('/almacenes/traslados/<int:transfer_id>')
@login_required
def transfer_detail(transfer_id):
    """Muestra los detalles de un traslado específico."""
    
    transfer = WarehouseTransfer.query.options(
        joinedload(WarehouseTransfer.user)
    ).get_or_404(transfer_id)

    # Agrupar movimientos para obtener la lista de productos y cantidades
    movements = Movement.query.filter_by(document_id=transfer.id, document_type='Traslado de Almacén', type='Salida').options(
        joinedload(Movement.product),
        joinedload(Movement.warehouse)
    ).all()

    if not movements:
        flash('No se encontraron detalles para este traslado.', 'warning')
        return redirect(url_for('main.warehouse_transfer'))

    from_warehouse = movements[0].warehouse
    # Asumimos que el movimiento de entrada tiene el mismo product_id y cantidad
    to_warehouse_id = int(movements[0].description.split()[-1])
    to_warehouse = Warehouse.query.get(to_warehouse_id)

    return render_template('almacenes/detalle_traslado.html',
                           title=f'Detalle de Traslado #{transfer.id}',
                           transfer=transfer,
                           movements=movements,
                           from_warehouse=from_warehouse,
                           to_warehouse=to_warehouse)

@routes_blueprint.route('/almacenes/traslados/imprimir/<int:transfer_id>')
@login_required
def print_transfer_report(transfer_id):
    """Genera un PDF con los detalles de un traslado."""
    
    transfer = WarehouseTransfer.query.options(
        joinedload(WarehouseTransfer.user)
    ).get_or_404(transfer_id)

    movements = Movement.query.filter_by(document_id=transfer.id, document_type='Traslado de Almacén', type='Salida').options(
        joinedload(Movement.product),
        joinedload(Movement.warehouse)
    ).all()

    if not movements:
        flash('No se encontraron detalles para este traslado.', 'warning')
        return redirect(url_for('main.transfer_detail', transfer_id=transfer.id))

    from_warehouse = movements[0].warehouse
    to_warehouse_id = int(movements[0].description.split()[-1])
    to_warehouse = Warehouse.query.get(to_warehouse_id)
    generation_date = get_current_time_ve().strftime('%d/%m/%Y %H:%M:%S')
    company_info = CompanyInfo.query.first()
    logo_path = None
    if company_info and company_info.logo_filename:
        # Construir la ruta absoluta y convertirla a una URI de archivo compatible
        absolute_path = Path(current_app.root_path) / 'static' / company_info.logo_filename
        logo_path = absolute_path.as_uri()

    # Calcular el costo total del traslado
    total_cost_usd = sum(m.quantity * (m.product.cost_usd or 0) for m in movements)

    html_string = render_template('pdf/transfer_report.html',
                                  transfer=transfer, movements=movements,
                                  from_warehouse=from_warehouse, to_warehouse=to_warehouse,
                                  generation_date=generation_date,
                                  company_info=company_info,
                                  total_cost_usd=total_cost_usd,
                                  logo_path=logo_path)
    
    pdf_file = HTML(string=html_string, base_url=request.base_url).write_pdf()
    return Response(pdf_file, mimetype='application/pdf', headers={'Content-Disposition': f'inline; filename=reporte_traslado_{transfer.id}.pdf'})

@routes_blueprint.route('/almacenes/traslados/historial')
@login_required
def transfer_history():
    """Muestra una lista de todos los traslados de almacén realizados."""

    transfers = WarehouseTransfer.query.options(
        joinedload(WarehouseTransfer.user)
    ).order_by(WarehouseTransfer.date.desc()).all()

    return render_template('almacenes/historial_traslados.html',
                           title='Historial de Traslados de Almacén',
                           transfers=transfers)

@routes_blueprint.route('/api/warehouse_stock/<int:warehouse_id>')
@login_required
def api_warehouse_stock(warehouse_id):
    """API para obtener el stock de todos los productos en un almacén específico."""
    
    stocks = ProductStock.query.filter_by(warehouse_id=warehouse_id).all()
    stock_map = {stock.product_id: stock.quantity for stock in stocks}
    return jsonify(stocks=stock_map)

@routes_blueprint.route('/api/product_by_barcode_for_transfer/<barcode>')
@login_required
def api_product_by_barcode_for_transfer(barcode):
    """API para obtener un producto por código de barras y su stock en un almacén específico."""
    
    warehouse_id = request.args.get('warehouse_id', type=int)
    if not warehouse_id:
        return jsonify({'error': 'ID de almacén no proporcionado'}), 400

    product = Product.query.filter_by(barcode=barcode).first()
    if not product:
        return jsonify({'error': 'Producto no encontrado'}), 404

    stock_entry = ProductStock.query.filter_by(product_id=product.id, warehouse_id=warehouse_id).first()
    stock_quantity = stock_entry.quantity if stock_entry else 0

    # Devolvemos el mismo formato que la búsqueda por nombre para reutilizar addProductToTable
    product_data = {'id': product.id, 'name': product.name, 'barcode': product.barcode, 'stock': stock_quantity}
    
    return jsonify(product_data)

@routes_blueprint.route('/api/search_products_for_transfer')
@login_required
def api_search_products_for_transfer():
    """API para buscar productos por nombre o código de barras para traslados."""

    query_str = request.args.get('q', '').strip().lower()
    from_warehouse_id = request.args.get('from_warehouse_id', type=int)

    if len(query_str) < 2 or not from_warehouse_id:
        return jsonify(products=[])

    search_pattern = f'%{query_str}%'
    
    # Subconsulta para obtener el stock del almacén de origen
    stock_subquery = db.session.query(ProductStock.quantity).filter(ProductStock.product_id == Product.id, ProductStock.warehouse_id == from_warehouse_id).as_scalar()

    products = db.session.query(Product, func.coalesce(stock_subquery, 0).label('stock_in_warehouse')).filter(
        or_(
            Product.name.ilike(search_pattern), 
            Product.barcode.ilike(search_pattern),
            Product.codigo_producto.ilike(search_pattern) # Añadido: buscar por código de producto
        )
    ).limit(10).all()

    results = [{'id': p.id, 'name': p.name, 'barcode': p.barcode, 'codigo_producto': p.codigo_producto, 'stock': stock} for p, stock in products]
    
    return jsonify(products=results)

@routes_blueprint.route('/api/search_products_for_purchase')
@login_required
def api_search_products_for_purchase():
    """
    API para buscar productos por nombre, código de barras o código de producto para compras.
    Devuelve todos los productos que coincidan, sin importar el stock.
    """
    if not is_administrador():
        return jsonify({'error': 'Acceso denegado'}), 403

    query_str = request.args.get('q', '').strip()

    if len(query_str) < 2:
        return jsonify(products=[])

    search_pattern = f'%{query_str}%'
    
    products = Product.query.filter(
        or_(
            Product.name.ilike(search_pattern), 
            Product.barcode.ilike(search_pattern),
            Product.codigo_producto.ilike(search_pattern)
        )
    ).limit(10).all()

    results = [{
        'id': p.id, 
        'name': p.name, 
        'barcode': p.barcode, 
        'codigo_producto': p.codigo_producto, 
        'cost_usd': p.cost_usd
    } for p in products]
    
    return jsonify(products=results)

@routes_blueprint.route('/api/search_products_for_sale')
@login_required
def api_search_products_for_sale():
    """
    API para buscar productos por nombre, código de barras o código de producto para la venta.
    Solo devuelve productos con stock en el almacén principal (Tienda, ID=1).
    """
    query_str = request.args.get('q', '').strip()

    if len(query_str) < 2:
        return jsonify(products=[])

    search_pattern = f'%{query_str}%'
    
    # Subconsulta para obtener el stock del almacén de tienda (ID=1)
    stock_subquery = db.session.query(ProductStock.quantity).filter(ProductStock.product_id == Product.id, ProductStock.warehouse_id == 1).as_scalar()

    products_with_stock = db.session.query(Product, func.coalesce(stock_subquery, 0).label('stock_in_warehouse')).filter(
        or_(
            Product.name.ilike(search_pattern), 
            Product.barcode.ilike(search_pattern),
            Product.codigo_producto.ilike(search_pattern)
        ),
        func.coalesce(stock_subquery, 0) > 0  # Solo productos con stock en tienda
    ).limit(10).all()

    results = [{'id': p.id, 'name': p.name, 'barcode': p.barcode, 'codigo_producto': p.codigo_producto, 'price_usd': p.price_usd, 'stock': stock} for p, stock in products_with_stock]
    
    return jsonify(products=results)

@routes_blueprint.route('/manual_usuario')
def user_manual():
    """Muestra la página del manual de usuario."""
    return render_template('manual_usuario.html')

# --- Rutas de Gestión de Usuarios (Solo Superusuario) ---

@routes_blueprint.route('/configuracion/usuarios', methods=['GET'])
@login_required
def user_management():
    if not is_superuser():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    show_inactive = request.args.get('show_inactive') == 'true'

    # Query users except the current superuser
    query = User.query.filter(User.id != current_user.id)

    if not show_inactive:
        query = query.filter(User.is_active_status == True)
    
    users = query.order_by(User.username).all()
    
    return render_template('configuracion/usuarios.html', title='Gestión de Usuarios', users=users, show_inactive=show_inactive)

@routes_blueprint.route('/configuracion/entregas_pendientes', methods=['GET'])
@login_required
def pending_dispatches():
    """Muestra las entregas especiales pendientes de aprobación."""
    if not is_gerente():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.dashboard'))

    pending_orders = Order.query.filter_by(
        status='Pendiente de Aprobación',
        order_type='special_dispatch'
    ).options(
        joinedload(Order.client)
    ).order_by(Order.date_created.desc()).all()

    return render_template('configuracion/entregas_pendientes.html',
                           title='Entregas Especiales por Aprobar',
                           orders=pending_orders)

@routes_blueprint.route('/configuracion/entregas/procesar/<int:order_id>/<string:action>', methods=['POST'])
@login_required
def process_dispatch(order_id, action):
    """Procesa la aprobación o rechazo de una entrega especial."""
    if not is_gerente():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.dashboard'))

    order = Order.query.get_or_404(order_id)
    if order.status != 'Pendiente de Aprobación':
        flash('Esta entrega ya ha sido procesada.', 'warning')
        return redirect(url_for('main.pending_dispatches'))

    try:
        if action == 'approve':
            # 1. Check stock and deduct from inventory
            for item in order.items:
                stock_entry = ProductStock.query.filter_by(product_id=item.product_id, warehouse_id=1).first()
                if not stock_entry or stock_entry.quantity < item.quantity:
                    raise ValueError(f"Stock insuficiente para '{item.product.name}' para aprobar la entrega.")
                
                stock_entry.quantity -= item.quantity

                # 2. Create inventory movement
                movement = Movement(
                    product_id=item.product_id, type='Salida', warehouse_id=1,
                    quantity=item.quantity, document_id=order.id, document_type='Entrega Especial',
                    description=f"Aprobación de entrega para cliente #{order.client_id}",
                    related_party_id=order.client_id, related_party_type='Cliente', date=get_current_time_ve()
                )
                db.session.add(movement)

            # 3. Update order status
            order.status = 'Completada'
            flash(f'Entrega Especial #{order.id} aprobada y stock actualizado.', 'success')

        elif action == 'reject':
            order.status = 'Anulada'
            flash(f'Entrega Especial #{order.id} ha sido rechazada.', 'info')
        
        else:
            raise ValueError("Acción no válida.")

        db.session.commit()
    except (ValueError, IntegrityError) as e:
        db.session.rollback()
        flash(f'Error al procesar la entrega: {e}', 'danger')

    return redirect(url_for('main.pending_dispatches'))

@routes_blueprint.route('/configuracion/usuarios/nuevo', methods=['POST'])
@login_required
def add_user():
    if not is_superuser():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.user_management'))
    
    username = request.form.get('username')
    password = request.form.get('password')
    role = request.form.get('role')

    if not all([username, password, role]):
        flash('Todos los campos son requeridos.', 'danger')
        return redirect(url_for('main.user_management'))

    if User.query.filter_by(username=username).first():
        flash(f'El nombre de usuario "{username}" ya existe.', 'danger')
        return redirect(url_for('main.user_management'))

    hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
    new_user = User(username=username, password=hashed_password, role=role)
    db.session.add(new_user)
    db.session.commit()

    flash(f'Usuario "{username}" creado exitosamente.', 'success')
    return redirect(url_for('main.user_management'))

@routes_blueprint.route('/configuracion/usuarios/editar/<int:user_id>', methods=['POST'])
@login_required
def edit_user(user_id):
    if not is_superuser():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.user_management'))

    user_to_edit = User.query.get_or_404(user_id)
    
    user_to_edit.role = request.form.get('role')
    new_password = request.form.get('password')

    if new_password:
        user_to_edit.password = bcrypt.generate_password_hash(new_password).decode('utf-8')
    
    db.session.commit()
    flash(f'Usuario "{user_to_edit.username}" actualizado exitosamente.', 'success')
    return redirect(url_for('main.user_management'))

@routes_blueprint.route('/configuracion/usuarios/toggle_status/<int:user_id>', methods=['POST'])
@login_required
def toggle_user_status(user_id):
    """Toggles the is_active status of a user."""
    if not is_superuser():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.user_management'))

    user_to_toggle = User.query.get_or_404(user_id)
    
    # Toggle the is_active status
    user_to_toggle.is_active_status = not user_to_toggle.is_active_status
    db.session.commit()

    status = "activado" if user_to_toggle.is_active_status else "desactivado"
    flash(f'El usuario "{user_to_toggle.username}" ha sido {status}.', 'success')
    return redirect(url_for('main.user_management'))


@routes_blueprint.route('/ordenes/devoluciones/lista')
@login_required
def return_list():
    """Displays a history of all returns and cancellations."""
    if not is_administrador():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.dashboard'))

    returns = OrderReturn.query.options(
        joinedload(OrderReturn.user),
        joinedload(OrderReturn.order)
    ).order_by(OrderReturn.date.desc()).all()

    return render_template('ordenes/lista_devoluciones.html', title='Historial de Devoluciones', returns=returns)

@routes_blueprint.route('/ordenes/devoluciones/detalle/<int:return_id>')
@login_required
def return_detail(return_id):
    """Shows the details of a specific return."""
    if not is_administrador():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('main.dashboard'))

    # Usamos subqueryload para cargar los items y sus productos relacionados de forma eficiente.
    # Es una alternativa a joinedload que puede ser más performante para colecciones.
    return_record = OrderReturn.query.options(
        subqueryload(OrderReturn.items).joinedload(OrderReturnItem.product),
        joinedload(OrderReturn.order).joinedload(Order.client),
        joinedload(OrderReturn.user)
    ).filter_by(id=return_id).first_or_404()
    return render_template('ordenes/detalle_devolucion.html', title=f'Detalle de Devolución {return_record.return_code}', return_record=return_record)
# Nueva ruta para movimientos de inventario
@routes_blueprint.route('/ordenes/devolucion', methods=['GET', 'POST']) # type: ignore
@login_required
def return_order():
    """
    Handles the cancellation (anulación) or return (devolución) of products from a sales order.
    """
    if not is_administrador():
        flash('Acceso denegado. Solo los administradores pueden realizar esta acción.', 'danger')
        return redirect(url_for('main.dashboard'))

    order = None
    banks = Bank.query.order_by(Bank.name).all()
    cash_boxes = CashBox.query.order_by(CashBox.name).all()


    if request.method == 'POST':
        order_id = request.form.get('order_id')
        action = request.form.get('action')
        order = Order.query.get_or_404(order_id)

        if order.status in ['Anulada']:
            flash(f'La orden #{order.id} ya ha sido procesada para anulación o devolución.', 'warning')
            return redirect(url_for('main.return_order'))

        try:
            # Generate a unique code for the return
            today_str = get_current_time_ve().strftime('%y%m%d')
            count_today = OrderReturn.query.filter(OrderReturn.return_code.like(f"DEV{today_str}%")).count()
            return_code = f"DEV{today_str}-{count_today + 1}"

            if action == 'anulacion_total':
                # 1. Create the main return record
                total_value_ves = order.total_amount
                return_record = OrderReturn(
                    return_code=return_code,
                    order_id=order.id,
                    user_id=current_user.id,
                    return_type='Anulación Total',
                    reason='Anulación total de la orden.',
                    total_refund_value_ves=total_value_ves
                )
                db.session.add(return_record)
                db.session.flush()

                # Return all items to stock
                for item in order.items:
                    # Create return item record
                    OrderReturnItem.create(order_return_id=return_record.id, order_item_id=item.id, product_id=item.product_id, quantity=item.quantity, price_at_return_ves=item.price)
                    warehouse_id_origen = 1 # Las ventas siempre salen de la tienda principal

                    stock_entry = ProductStock.query.filter_by(product_id=item.product_id, warehouse_id=warehouse_id_origen).first()
                    if stock_entry:
                        stock_entry.quantity += item.quantity
                    else:
                        stock_entry = ProductStock(product_id=item.product_id, warehouse_id=warehouse_id_origen, quantity=item.quantity)
                        db.session.add(stock_entry)
                    
                    # Create OrderReturnItem
                    return_item_record = OrderReturnItem(
                        order_return_id=return_record.id,
                        order_item_id=item.id,
                        product_id=item.product_id,
                        quantity=item.quantity,
                        price_at_return_ves=item.price
                    )
                    db.session.add(return_item_record)

                    # 2. Crear movimiento de inventario de entrada
                    movement = Movement(
                        product_id=item.product_id, type='Entrada', warehouse_id=warehouse_id_origen,
                        quantity=item.quantity, document_id=order.id, document_type='Anulación de Venta',
                        description=f"Anulación total de la orden de venta #{order.id}",
                        related_party_id=order.client_id, related_party_type='Cliente'
                    )
                    db.session.add(movement)

                # 3. Revertir los pagos creando movimientos financieros de egreso
                for payment in order.payments:
                    refund_movement = ManualFinancialMovement(
                        description=f"Reverso por anulación de orden #{order.id}. Pago original ID: {payment.id}",
                        amount=payment.amount_paid,
                        currency=payment.currency_paid,
                        movement_type='Egreso',
                        status='Aprobado',
                        created_by_user_id=current_user.id,
                        approved_by_user_id=current_user.id,
                        date_approved=get_current_time_ve(),
                        bank_id=payment.bank_id or (payment.pos.bank.id if payment.pos and payment.pos.bank else None),
                        order_return_id=return_record.id,
                        cash_box_id=payment.cash_box_id
                    )
                    db.session.add(refund_movement)

                    # Actualizar saldos de cuentas
                    if refund_movement.bank_id:
                        bank = Bank.query.get(refund_movement.bank_id)
                        if bank: bank.balance -= payment.amount_ves_equivalent
                    elif refund_movement.cash_box_id:
                        cash_box = CashBox.query.get(refund_movement.cash_box_id)
                        if cash_box:
                            if payment.currency_paid == 'VES': cash_box.balance_ves -= payment.amount_paid
                            elif payment.currency_paid == 'USD': cash_box.balance_usd -= payment.amount_paid

                # 4. Actualizar estado de la orden y registrar actividad
                order.status = 'Anulada'
                log_user_activity(action="Anuló orden de venta", details=f"Anulación total de la orden #{order.id}", target_id=order.id, target_type="Order")
                db.session.commit()
                flash(f'Orden #{order.id} anulada exitosamente. El stock y los saldos financieros han sido restaurados.', 'success')
                return redirect(url_for('main.return_detail', return_id=return_record.id))

            elif action == 'devolucion_parcial':
                return_items = request.form.getlist('return_item_id')
                return_quantities = request.form.getlist('return_quantity')
                return_reason = request.form.get('return_reason', '').strip()
                refund_payments_json = request.form.get('refund_payments_data') # This can be None
                refund_payments = json.loads(refund_payments_json) if refund_payments_json else []


                items_returned_count = 0
                total_refund_value_ves = 0.0

                for item_id, qty_str in zip(return_items, return_quantities):
                    if (int(qty_str) if qty_str else 0) > 0:
                        items_returned_count += 1
                
                if items_returned_count == 0:
                    flash('No se seleccionaron productos para devolver.', 'warning')
                    return redirect(url_for('main.return_order', order_id=order.id))

                if not return_reason:
                    raise ValueError("El motivo de la devolución es obligatorio.")


                for item_id, qty_str in zip(return_items, return_quantities):
                    quantity_to_return = int(qty_str) if qty_str else 0
                    if quantity_to_return <= 0:
                        continue

                    order_item = OrderItem.query.get(item_id)
                    if not order_item or order_item.order_id != order.id: # type: ignore
                        continue
                    
                    # Check if trying to return more than what's left
                    if quantity_to_return > (order_item.quantity - (order_item.returned_quantity or 0)):
                        raise ValueError(f"Intento de devolver más unidades de las disponibles para '{order_item.product.name}'.")

                    if quantity_to_return > order_item.quantity:
                        raise ValueError(f"No se puede devolver más de la cantidad vendida para el producto '{order_item.product.name}'.")

                    total_refund_value_ves += quantity_to_return * order_item.price

                # Create the main return record
                return_record = OrderReturn(
                    return_code=return_code,
                    order_id=order.id,
                    user_id=current_user.id,
                    return_type='Devolución Parcial',
                    reason=return_reason,
                    total_refund_value_ves=total_refund_value_ves
                )
                db.session.add(return_record)
                db.session.flush()

                for item_id, qty_str in zip(return_items, return_quantities):
                    quantity_to_return = int(qty_str) if qty_str else 0
                    if quantity_to_return <= 0: continue
                    order_item = OrderItem.query.get(item_id)
                    warehouse_id_origen = 1 # Las ventas siempre salen de la tienda principal


                    stock_entry = ProductStock.query.filter_by(product_id=order_item.product_id, warehouse_id=warehouse_id_origen).first()
                    if stock_entry:
                        stock_entry.quantity += quantity_to_return
                    else:
                        stock_entry = ProductStock(product_id=order_item.product_id, warehouse_id=warehouse_id_origen, quantity=quantity_to_return)
                        db.session.add(stock_entry)
                    
                    # Update the returned quantity on the original order item
                    order_item.returned_quantity = (order_item.returned_quantity or 0) + quantity_to_return

                    # Create OrderReturnItem record
                    return_item_record = OrderReturnItem(
                        order_return_id=return_record.id, order_item_id=order_item.id,
                        product_id=order_item.product_id, quantity=quantity_to_return,
                        price_at_return_ves=order_item.price)
                    db.session.add(return_item_record)

                    movement = Movement(
                        product_id=order_item.product_id, type='Entrada', warehouse_id=warehouse_id_origen,
                        quantity=quantity_to_return, document_id=order.id, document_type='Devolución de Venta',
                        description=f"Devolución de la orden #{order.id}. Motivo: {return_reason}",
                        related_party_id=order.client_id, related_party_type='Cliente'
                    )
                    db.session.add(movement)

                # Procesar los reembolsos financieros
                total_refunded_ves = 0.0 # Initialize here
                for payment_info in refund_payments:
                    amount = float(payment_info['amount'])
                    currency = payment_info['currency']
                    amount_ves_equivalent = float(payment_info['amount_ves_equivalent'])
                    total_refunded_ves += amount_ves_equivalent

                    refund_movement = ManualFinancialMovement(
                        description=f"Reembolso por devolución de orden #{order.id}. Motivo: {return_reason}",
                        amount=amount, currency=currency, movement_type='Egreso', status='Aprobado',
                        created_by_user_id=current_user.id, approved_by_user_id=current_user.id,
                        date_approved=get_current_time_ve(),
                        bank_id=payment_info.get('bank_id'), cash_box_id=payment_info.get('cash_box_id'),
                        order_return_id=return_record.id
                    )
                    db.session.add(refund_movement)

                    # Actualizar saldos de cuentas
                    if refund_movement.bank_id:
                        bank = Bank.query.get(refund_movement.bank_id)
                        if bank: bank.balance -= amount_ves_equivalent
                    elif refund_movement.cash_box_id:
                        cash_box = CashBox.query.get(refund_movement.cash_box_id)
                        if cash_box:
                            if currency == 'VES': cash_box.balance_ves -= amount
                            elif currency == 'USD': cash_box.balance_usd -= amount

                # Validar que el monto reembolsado coincida con el valor de los productos devueltos
                if refund_payments and abs(total_refunded_ves - total_refund_value_ves) > 0.01:
                    raise ValueError(f"El monto reembolsado (Bs. {total_refunded_ves:.2f}) no coincide con el valor de los productos devueltos (Bs. {total_refund_value_ves:.2f}).")

                order.status = 'Devolución Parcial'
                log_user_activity(action="Procesó devolución de venta", details=f"Devolución parcial de la orden #{order.id}. Motivo: {return_reason}", target_id=order.id, target_type="Order")
                db.session.commit()
                flash(f'Devolución para la orden #{order.id} procesada. El stock y los saldos financieros han sido actualizados.', 'success')
                return redirect(url_for('main.return_detail', return_id=return_record.id))

        except (ValueError, IntegrityError) as e:
            db.session.rollback()
            flash(f'Error al procesar la devolución: {e}', 'danger')
            return redirect(url_for('main.return_order', order_id=order.id)) # type: ignore
        
    order_id_query = request.args.get('order_id_search', type=int)
    if order_id_query:
        order = Order.query.options(joinedload(Order.items).joinedload(OrderItem.product)).get(order_id_query)
        if not order:
            flash('La orden especificada no fue encontrada.', 'danger')
            # Redirect if order is not found to prevent rendering with a None object
            return redirect(url_for('main.return_order'))

    return render_template('ordenes/devolucion.html', title='Anulación / Devolución de Venta', order=order, banks=banks, cash_boxes=cash_boxes)